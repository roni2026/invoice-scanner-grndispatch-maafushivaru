# maafushivaru_hub.py
# Maafushivaru - Document Processing Hub
# v5.0 - Watchdog auto-ingest, desktop notifications, confidence scoring,
#         supplier learning, scroll fix, live status bar, professional UI

import os
import re
import io
import sys
import json
import shutil
import queue
import logging
import threading
import traceback
import concurrent.futures
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Tuple
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pytesseract
import fitz
import cv2
import numpy as np
from PIL import Image, ImageTk, Image as PILImage
from rapidfuzz import process, fuzz
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ---------------------------------------------------------------------------
# OPTIONAL DEPENDENCIES
# ---------------------------------------------------------------------------
try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
    WATCHDOG_AVAILABLE = True
except ImportError:
    WATCHDOG_AVAILABLE = False

try:
    from plyer import notification as plyer_notification
    PLYER_AVAILABLE = True
except ImportError:
    PLYER_AVAILABLE = False
# New matching strategies
try:
    from supplier_matcher import dispatch_match, STRATEGY_LABELS, STRATEGY_FUNCTIONS
    SUPPLIER_MATCHER_AVAILABLE = True
except ImportError:
    SUPPLIER_MATCHER_AVAILABLE = False
    STRATEGY_LABELS = {}
    STRATEGY_FUNCTIONS = {}

# AI supplier detection
try:
    from ai_supplier_matcher import AISupplierMatcher, PROVIDERS, STATUS_CONNECTED, \
        STATUS_DISCONNECTED, STATUS_LOW_CREDIT, STATUS_OFFLINE
    AI_MATCHER_AVAILABLE = True
except ImportError:
    AI_MATCHER_AVAILABLE = False
    PROVIDERS = {}
    STATUS_CONNECTED    = "connected"
    STATUS_DISCONNECTED = "disconnected"
    STATUS_LOW_CREDIT   = "low_credit"
    STATUS_OFFLINE      = "offline"
# OCR word corrector
try:
    from ocr_word_corrector import correct_supplier_in_text, correct_invoice_number
    OCR_CORRECTOR_AVAILABLE = True
except ImportError:
    OCR_CORRECTOR_AVAILABLE = False
    def correct_supplier_in_text(text, suppliers, aliases):
        return text, None
    def correct_invoice_number(raw, supplier):
        return raw

# Smart cross-matcher
try:
    from smart_cross_matcher import infer_invoice_from_supplier, infer_supplier_from_invoice
    CROSS_MATCHER_AVAILABLE = True
except ImportError:
    CROSS_MATCHER_AVAILABLE = False
    def infer_invoice_from_supplier(raw, supplier):
        return raw
    def infer_supplier_from_invoice(text, suppliers, aliases):
        return None, 0.0

# ---------------------------------------------------------------------------
# CONSTANTS / COLORS
# ---------------------------------------------------------------------------
APP_TITLE   = "Maafushivaru - Document Processing Hub"
APP_VERSION = "v5.0"

# Color palette — professional dark theme
BG       = "#0A0F1E"          # deepest background
PANEL    = "#111827"          # card / panel background
PANEL2   = "#1F2937"          # slightly lighter panel
PANEL3   = "#374151"          # hover / border
TEXT     = "#F9FAFB"          # primary text
MUTED    = "#9CA3AF"          # secondary text
ACCENT   = "#3B82F6"          # primary blue
ACCENT_H = "#2563EB"          # hover blue
ACCENT2  = "#8B5CF6"          # purple
SUCCESS  = "#10B981"          # green
WARNING  = "#F59E0B"          # amber
ERROR    = "#EF4444"          # red
BORDER   = "#1F2937"          # subtle border
WATCHER  = "#06B6D4"          # cyan for watcher badge

ENGINE_LABELS  = {"tesseract": "Tesseract OCR", "paddleocr": "PaddleOCR", "easyocr": "EasyOCR"}
ENGINE_INSTALL = {
    "tesseract": "Built-in binary",
    "paddleocr": "pip install paddlepaddle paddleocr",
    "easyocr":   "pip install easyocr",
}
ENGINE_MODULES = {"tesseract": "pytesseract", "paddleocr": "paddleocr", "easyocr": "easyocr"}
_ENGINE_CACHE  = {}
NUMBER_FMT = "#,##0.00"

ZONE_OCR_REGIONS = {
    # Each entry is (top_pct, bottom_pct, left_pct, right_pct)
    # These cover the areas where GRN, PO, supplier, date and invoice
    # data consistently appear in Birchstreet receiving reports.
    "header_top":    (0.00, 0.08, 0.00, 1.00),   # "Receiving Record #" and "Purchase Order #" line
    "header_meta":   (0.08, 0.22, 0.00, 1.00),   # Received by / date / PO status block
    "invoice_block": (0.22, 0.38, 0.00, 1.00),   # Invoice number / subtotal / total block
    "supplier_col":  (0.22, 0.38, 0.60, 1.00),   # Right column: Supplier name area
}

# Minimum characters a zone must return before we accept it as valid.
_ZONE_MIN_CHARS = 20

# How many zones must succeed for us to trust zone results over full-page OCR.
_ZONE_MIN_SUCCESS = 2

# ---------------------------------------------------------------------------
# THREAD-LOCAL SUPPLIER TRACKING
# ---------------------------------------------------------------------------
_thread_local = threading.local()

def _get_last_good_supplier():
    return getattr(_thread_local, "last_good_supplier", None)

def _set_last_good_supplier(value):
    _thread_local.last_good_supplier = value

def _reset_last_good_supplier():
    _thread_local.last_good_supplier = None

# ---------------------------------------------------------------------------
# CURRENCY MAPS
# ---------------------------------------------------------------------------
_CURRENCY_MAP = {
    "$": "USD", "US$": "USD", "USD": "USD",
    "MVR": "MVR", "RF": "MVR", "MRF": "MVR",
    "MYR": "MVR",   # OCR sometimes reads MVR as MYR — treat as MVR
    "EUR": "EUR", "€": "EUR",
    "GBP": "GBP", "£": "GBP",
    "SGD": "SGD", "S$": "SGD",
}
_INVOICE_CURRENCY_TOKENS = {
    "USD", "MVR", "RF", "MRF", "EUR", "GBP", "SGD", "MYR",
    "US$", "S$", "$", "€", "£"
}

_REPORT_INVOICE_TOTAL_PAT = re.compile(
    r"\bI?NVOICE\s*TOTAL\s*[:\-]?\s*"
    r"(?:(USD|MVR|MYR|RF|MRF|EUR|GBP|SGD|US\$|S\$|\$|€|£)\s*)?"
    r"([\d,]+(?:\.\d{1,2})?)"
    r"(?:\s*(USD|MVR|MYR|RF|MRF|EUR|GBP|SGD|US\$|S\$|\$|€|£))?",
    re.IGNORECASE,
)


def _fix_ocr_currency_amount(raw: str) -> str:
    """
    Fix OCR errors in currency+amount strings.

    Handles cases like:
      - MVRI,380.14   -> MVR1,380.14
      - MYR1,380.14   -> MVR1,380.14
      - MYRI,380.14   -> MVR1,380.14
      - MVR I,380.14  -> MVR 1,380.14
      - USD I23.50    -> USD123.50
      - EUR l,250.00  -> EUR1,250.00

    Returns the corrected string.
    """
    if not raw:
        return raw

    fixed = raw

    # Normalize MYR -> MVR when it appears as OCR currency for Maldivian invoices
    fixed = re.sub(r"\bMYR\b", "MVR", fixed, flags=re.IGNORECASE)

    # Cases like MVRI,380.14 or MYRI,380.14 -> MVR1,380.14
    fixed = re.sub(
        r"\b(MVR|MYR|USD|EUR|GBP|SGD|RF|MRF)([Il])(?=[\d,])",
        lambda m: ("MVR" if m.group(1).upper() == "MYR" else m.group(1).upper()) + "1",
        fixed,
        flags=re.IGNORECASE,
    )

    # Cases like MVR I,380.14 -> MVR 1,380.14
    fixed = re.sub(
        r"\b(MVR|MYR|USD|EUR|GBP|SGD|RF|MRF)\s+([Il])\s*(?=[\d,])",
        lambda m: ("MVR" if m.group(1).upper() == "MYR" else m.group(1).upper()) + " 1",
        fixed,
        flags=re.IGNORECASE,
    )

    return fixed

SUPPLIER_INVOICE_HINTS = {
    "BEST BUY":     [r"BB[-\/]\d{3,10}", r"\d{5,10}"],
    "BESTBUY":      [r"BB[-\/]\d{3,10}", r"\d{5,10}"],
    "STANDARD":     [r"SI[-\/]\d{5,10}", r"IS[-\/]\d{5,10}", r"\d{5,10}"],
    "ORIGIN":       [r"SI[-\/]\d{5,10}", r"IS[-\/]\d{5,10}", r"\d{5,10}"],
    "EURO MARKET":  [r"[A-Z]{2,4}[-\/]\d{4,10}", r"\d{5,10}"],
    "LYCORN":       [r"\d{5,10}", r"[A-Z]{2,4}[-\/]\d{4,10}"],
    "COSMO":        [r"[A-Z]{2,4}[-\/]\d{4,10}\/\d{4}", r"[A-Z]{2,4}[-\/]\d{4,10}"],
    "EMPARAL":      [r"\d{5,10}", r"[A-Z]{2,4}[-\/]\d{4,10}"],
    "HAPPY MARKET": [r"\d{5,10}"],
    "SAWHNEY":      [r"[A-Z]{2,4}\/\d+\/\d{4}", r"[A-Z]{2,4}[-]\d{4,10}"],
    "EASTERN":      [r"[A-Z]{2,4}[-\/]\d{4,10}", r"\d{5,10}"],
    "FOOD SPECIAL": [r"\d{5,10}", r"[A-Z]{2,4}[-\/]\d{4,10}"],
    "CHEF":         [r"[A-Z]{2,4}[-\/]\d{4,10}", r"\d{5,10}"],
    "SEAFOOD":      [r"\d{5,10}", r"[A-Z]{2,4}[-\/]\d{4,10}"],
}

_INVOICE_LABEL_PATS = [
    r"INVOICE\s*(?:NUMBER|NUM(?:BER)?|NO\.?|#|NR\.?)\s*[:\-]",
    r"INV\.?\s*(?:NO\.?|#|NUMBER)\s*[:\-]",
    r"TAX\s*INVOICE\s*(?:NO\.?|#|NUMBER)?\s*[:\-]",
    r"BILL\s*(?:NO\.?|NUMBER|#)\s*[:\-]",
    r"CREDIT\s*NOTE\s*(?:NO\.?|#|NUMBER)?\s*[:\-]",
]


_INVOICE_STOP_WORDS = [
    "DATE", "GRN", "RECEIVING", "SUPPLIER", "VENDOR", "PURCHASE ORDER",
    "SUBTOTAL", "TOTAL", "AMOUNT", "QTY", "QUANTITY", "DESCRIPTION", "TAX",
]


def _find_invoice_value(raw: str) -> Optional[str]:
    for label_pat in _INVOICE_LABEL_PATS:
        lm = re.search(label_pat, raw)
        if not lm:
            continue
        rest = raw[lm.end(): lm.end() + 200]

        # Cut off at the next line break, or the next field label found on
        # the same line (so we don't swallow "DATE: ..." etc into the value)
        cut_pos = len(rest)
        nl_pos = rest.find("\n")
        if nl_pos != -1:
            cut_pos = min(cut_pos, nl_pos)
        for sw in _INVOICE_STOP_WORDS:
            m_sw = re.search(r"\b" + sw + r"\b", rest[:cut_pos])
            if m_sw:
                cut_pos = min(cut_pos, m_sw.start())

        value = rest[:cut_pos]
        # Collapse any double/triple spaces from OCR into single spaces,
        # but keep the FULL value instead of stopping at the first token.
        value = re.sub(r"\s+", " ", value).strip(" -:.")
        if value and len(value) >= 2:
            return value
    return None


def _clean_invoice_token(token: str) -> str:
    if not token:
        return ""
    t = token.strip().upper()
    if t in _INVOICE_CURRENCY_TOKENS:
        return ""
    for cur in sorted(_INVOICE_CURRENCY_TOKENS, key=len, reverse=True):
        t = t.replace(cur, " ")
    t = re.sub(r"\s+", " ", t).strip(" -:./\\")
    return t


# ---------------------------------------------------------------------------
# OCR ROBUSTNESS HELPERS
#   - fuzzy label detection (handles staple holes / smudged characters)
#   - digit confusion fixing for RC-MAM GRN numbers
#   - invoice series-prefix correction for underline misreads (I -> L / 1)
# ---------------------------------------------------------------------------

# Letters most commonly produced when an underline merges with a digit's glyph,
# mapped back to the digit they should be. Applied ONLY to the numeric region
# of an RC-MAM code, never to the 'RC'/'MAM' prefix.
_OCR_DIGIT_FIX = str.maketrans({
    "O": "0", "Q": "0", "D": "0",
    "I": "1", "L": "1", "|": "1", "!": "1",
    "Z": "2",
    "E": "3",
    "A": "4",
    "S": "5",
    "G": "6",
    "T": "7", "J": "7",
    "B": "8",
})


def _levenshtein(a: str, b: str) -> int:
    """Plain edit distance for short tokens (label words)."""
    if a == b:
        return 0
    la, lb = len(a), len(b)
    if la == 0:
        return lb
    if lb == 0:
        return la
    prev = list(range(lb + 1))
    for i, ca in enumerate(a, 1):
        cur = [i] + [0] * lb
        for j, cb in enumerate(b, 1):
            cost = 0 if ca == cb else 1
            cur[j] = min(prev[j] + 1, cur[j - 1] + 1, prev[j - 1] + cost)
        prev = cur
    return prev[lb]


def _label_present(text: str, label: str, max_dist: int = 2) -> bool:
    """True if every word of `label` appears (in order, consecutively) inside
    `text`, allowing up to `max_dist` single-character OCR errors per word.

    This is what lets the app still recognise 'RECEIVING RECORD' when a staple
    hole or smudge makes one or more characters unreadable, e.g. 'RECEIVING
    RECORO', 'RECE1VING RECORD', 'RECEIVNG RECORD'.
    """
    words = re.findall(r"[A-Z0-9]+", (text or "").upper())
    label_words = label.upper().split()
    n = len(label_words)
    W = len(words)
    if n == 0 or W == 0:
        return False

    def _word_ok(w: str, lw: str) -> bool:
        return abs(len(w) - len(lw)) <= 1 and _levenshtein(w, lw) <= max_dist

    # Match label words in order. Each label word may match a single input word
    # OR two adjacent input words joined together - this covers staple holes that
    # split a word in two (e.g. 'RECEIV NG' -> 'RECEIVING').
    for start in range(W):
        wi = start
        ok = True
        for lw in label_words:
            if wi >= W:
                ok = False
                break
            if _word_ok(words[wi], lw):
                wi += 1
            elif wi + 1 < W and _word_ok(words[wi] + words[wi + 1], lw):
                wi += 2
            else:
                ok = False
                break
        if ok:
            return True
    return False


def _fix_invoice_prefix(token: str, fixes: dict) -> str:
    """Correct an OCR-misread invoice *series prefix* using a config map.

    Underlined invoice codes like 'MSI-282910' are frequently read as
    'MSL-282910' (the underline turns the 'I' into an 'L') or 'MS1-282910'.
    `fixes` maps the wrong prefix to the right one, e.g. {'MSL': 'MSI'}.
    Only the leading alphabetic/numeric prefix before the first separator is
    touched, so the digits and everything else are preserved exactly.
    """
    if not token or not fixes:
        return token
    t = token.strip()
    m = re.match(r"^([A-Z0-9]{2,6})([\-/ ].*)?$", t.upper())
    if not m:
        return t
    prefix = m.group(1)
    rest = m.group(2) or ""
    fixed = fixes.get(prefix)
    if fixed:
        return fixed + rest
    return t


def _build_unique_word_index(suppliers: list, aliases: dict) -> dict:
    from collections import Counter
    all_cands = list(suppliers)
    for sub in aliases.values():
        all_cands.extend(sub)

    word_freq = Counter()
    cand_words = {}
    for c in all_cands:
        words = set(re.sub(r"[^A-Z0-9 ]", " ", c.upper()).split())
        words = {w for w in words if len(w) >= 3}
        cand_words[c] = words
        word_freq.update(words)

    unique_index = {}
    for c, words in cand_words.items():
        unique_index[c] = {w for w in words if word_freq[w] == 1}

    return unique_index


def _fmt_currency(val) -> str:
    if val in ("", None):
        return ""
    try:
        f = float(str(val).replace(",", ""))
        return f"{f:.2f}"
    except (ValueError, TypeError):
        return str(val)


# ---------------------------------------------------------------------------
# DESKTOP NOTIFICATION HELPER
# ---------------------------------------------------------------------------
def send_desktop_notification(title: str, message: str, timeout: int = 8):
    """Send a system notification. Silently no-ops if plyer not installed."""
    if not PLYER_AVAILABLE:
        return
    try:
        plyer_notification.notify(
            title=title,
            message=message,
            app_name="Maafushivaru Hub",
            timeout=timeout,
        )
    except Exception as e:
        logging.warning(f"Desktop notification failed: {e}")


# ---------------------------------------------------------------------------
# WATCHDOG FILE HANDLER
# ---------------------------------------------------------------------------
class ScannedFolderHandler:
    """Wraps watchdog logic; calls `on_new_pdf(path)` when a new PDF is stable."""

    def __init__(self, on_new_pdf, debounce_seconds: float = 3.0):
        self._on_new_pdf = on_new_pdf
        self._debounce = debounce_seconds
        self._pending: Dict[str, threading.Timer] = {}
        self._lock = threading.Lock()

    def _schedule(self, path: str):
        with self._lock:
            if path in self._pending:
                self._pending[path].cancel()
            t = threading.Timer(self._debounce, self._fire, args=(path,))
            self._pending[path] = t
            t.start()

    def _fire(self, path: str):
        with self._lock:
            self._pending.pop(path, None)
        if os.path.exists(path) and path.lower().endswith(".pdf"):
            self._on_new_pdf(path)

    def dispatch(self, event):
        """Called by watchdog observer for any filesystem event."""
        if event.is_directory:
            return
        src = getattr(event, "src_path", "")
        if src.lower().endswith(".pdf"):
            self._schedule(src)


if WATCHDOG_AVAILABLE:
    class _WatchdogEventAdapter(FileSystemEventHandler):
        def __init__(self, handler: ScannedFolderHandler):
            super().__init__()
            self._h = handler

        def on_created(self, event):
            self._h.dispatch(event)

        def on_modified(self, event):
            self._h.dispatch(event)

        def on_moved(self, event):
            # treat the destination as new
            class _Fake:
                def __init__(self, p):
                    self.src_path = p
                    self.is_directory = False
            self._h.dispatch(_Fake(getattr(event, "dest_path", "")))


# ---------------------------------------------------------------------------
# OCR / EXTRACTION WORKER MIXIN
# ---------------------------------------------------------------------------
class OCRWorkerMixin:

    # ------------- ENGINE HELPERS -------------
    @staticmethod
    def _check_engine(key):
        try:
            __import__(ENGINE_MODULES.get(key, key))
            return True
        except ImportError:
            return False

    def _get_engine_instance(self, key):
        if key in _ENGINE_CACHE:
            return _ENGINE_CACHE[key]
        if key == "paddleocr":
            from paddleocr import PaddleOCR
            _ENGINE_CACHE[key] = PaddleOCR(use_angle_cls=True, lang="en", show_log=False)
        elif key == "easyocr":
            import easyocr
            _ENGINE_CACHE[key] = easyocr.Reader(["en"], verbose=False)
        return _ENGINE_CACHE.get(key)

    # ------------- IMAGE / OCR -------------
    def _preprocess_image(self, pix):
        img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
        if pix.n == 4:
            img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        elif pix.n == 3:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        else:
            gray = img.squeeze() if img.ndim == 3 else img

        if self.cfg["app_settings"].get("enhance_images", True):
            clahe = cv2.createCLAHE(clipLimit=2.5, tileGridSize=(8, 8))
            gray = clahe.apply(gray)
            gray = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
            kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]], dtype=np.float32)
            gray = cv2.filter2D(gray, -1, kernel)

        _, out = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        return out

    def _correct_rotation(self, arr: np.ndarray) -> np.ndarray:
        """
        Correct both 90/180/270 rotations (via Tesseract OSD) and
        small deskew angles (tilted scans) via Hough line analysis.
        Falls back gracefully if either step fails.
        """
        # --- Step 1: coarse rotation via Tesseract OSD ---
        try:
            osd = pytesseract.image_to_osd(arr, config="--psm 0 -c min_characters_to_try=5")
            m = re.search(r"Rotate:\s*(\d+)", osd)
            angle = int(m.group(1)) if m else 0
            if angle == 90:
                arr = cv2.rotate(arr, cv2.ROTATE_90_COUNTERCLOCKWISE)
            elif angle == 180:
                arr = cv2.rotate(arr, cv2.ROTATE_180)
            elif angle == 270:
                arr = cv2.rotate(arr, cv2.ROTATE_90_CLOCKWISE)
        except Exception:
            # Fallback: if image is landscape, assume it needs 90° rotation
            h, w = arr.shape[:2]
            if w > h * 1.3:
                arr = cv2.rotate(arr, cv2.ROTATE_90_CLOCKWISE)

        # --- Step 2: fine deskew for small tilt angles (±15°) ---
        try:
            arr = self._deskew(arr)
        except Exception as e:
            logging.debug(f"Deskew skipped: {e}")

        return arr

    def _deskew(self, arr: np.ndarray) -> np.ndarray:
        """
        Detect and correct small tilt angles in a grayscale image.
        Uses morphological operations + Hough line transform.
        Only applies correction for angles in the range ±15° to avoid
        false corrections on legitimate landscape content.
        """
        # Work on a copy; ensure grayscale
        gray = arr.copy()
        if gray.ndim == 3:
            gray = cv2.cvtColor(gray, cv2.COLOR_BGR2GRAY)

        # Invert if background is mostly white (typical scanned document)
        mean_val = np.mean(gray)
        if mean_val > 127:
            gray = cv2.bitwise_not(gray)

        # Dilate horizontally to connect text into lines
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (30, 1))
        dilated = cv2.dilate(gray, kernel, iterations=2)

        # Find contours of the text "blocks"
        contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        angles = []
        for cnt in contours:
            if cv2.contourArea(cnt) < 500:
                continue
            rect = cv2.minAreaRect(cnt)
            angle = rect[2]
            # minAreaRect returns angles in [-90, 0); normalise to (-45, 45]
            if angle < -45:
                angle += 90
            angles.append(angle)

        if not angles:
            return arr  # nothing to correct

        # Use median to be robust against outliers
        skew_angle = float(np.median(angles))

        # Only correct if tilt is meaningful but not a full rotation
        if abs(skew_angle) < 0.3 or abs(skew_angle) > 15:
            return arr

        h, w = arr.shape[:2]
        center = (w // 2, h // 2)
        M = cv2.getRotationMatrix2D(center, skew_angle, 1.0)

        # Use white (255) border fill for scanned docs
        border_val = 255 if mean_val > 127 else 0
        rotated = cv2.warpAffine(
            arr, M, (w, h),
            flags=cv2.INTER_LINEAR,
            borderMode=cv2.BORDER_CONSTANT,
            borderValue=border_val,
        )
        return rotated
    def _ocr_zone(self, page, scale: float, top_pct: float, bottom_pct: float,
                  left_pct: float, right_pct: float) -> str:
        """
        OCR a rectangular sub-region of a fitz Page.
        Coordinates are given as fractions of the page dimensions (0.0–1.0).
        Returns upper-cased text, or empty string on failure.
        """
        try:
            rect = page.rect
            clip = fitz.Rect(
                rect.width  * left_pct,
                rect.height * top_pct,
                rect.width  * right_pct,
                rect.height * bottom_pct,
            )
            if clip.width < 10 or clip.height < 10:
                return ""
            pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), clip=clip)
            if pix.w == 0 or pix.h == 0:
                return ""
            arr = self._preprocess_image(pix)
            arr = self._correct_rotation(arr)
            return self._run_ocr(arr).upper()
        except Exception as e:
            logging.debug(f"Zone OCR failed: {e}")
            return ""

    def _extract_text_zone(self, pdf_path: str) -> str:
        """
        Zone-based text extraction for Birchstreet receiving report PDFs.

        Strategy:
        1. For each page, try native PDF text first (fast, free).
        2. If native text is thin, OCR only the 4 targeted header zones
           (top ~38% of page) instead of the full page.
        3. Concatenate zone texts — this gives all the data we need
           (GRN, PO, supplier, date, invoice number, totals) while
           skipping the large line-item table at the bottom.
        4. Fall back to full-page OCR if zone results are too sparse.

        Returns the combined upper-cased text for all pages.
        """
        full = ""
        scale = self.cfg["app_settings"].get("image_scale_factor", 2)
        use_native_flag = self.cfg["app_settings"].get("extract_text_before_ocr", True)

        try:
            doc = fitz.open(pdf_path)
            for page in doc:
                page_text = ""

                # --- Step 1: try native PDF text ---
                if use_native_flag:
                    native = page.get_text("text").upper()
                    if len(native.strip()) >= 50:
                        page_text = native
                        full += page_text + "\n"
                        continue   # Native text is good — skip OCR entirely

                # --- Step 2: Zone OCR (targeted regions only) ---
                zone_texts = []
                success_count = 0
                for zone_name, (top, bottom, left, right) in ZONE_OCR_REGIONS.items():
                    zt = self._ocr_zone(page, scale, top, bottom, left, right)
                    if len(zt.strip()) >= _ZONE_MIN_CHARS:
                        success_count += 1
                    zone_texts.append(zt)

                combined_zones = "\n".join(zone_texts)

                # --- Step 3: Fall back to full-page OCR if zones are sparse ---
                if success_count < _ZONE_MIN_SUCCESS:
                    logging.debug(f"Zone OCR insufficient ({success_count} zones) — falling back to full page")
                    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
                    arr = self._preprocess_image(pix)
                    arr = self._correct_rotation(arr)
                    page_text = self._run_ocr(arr).upper()
                else:
                    page_text = combined_zones

                full += page_text + "\n"
            doc.close()
        except Exception as e:
            logging.error(f"Zone text extraction failed [{pdf_path}]: {e}")
        return full

    def _run_ocr(self, arr):
        eng = self.cfg["app_settings"].get("ocr_engine", "tesseract")
        fb = self.cfg["app_settings"].get("ocr_fallback_to_tesseract", True)
        if eng == "paddleocr":
            try:
                r = self._get_engine_instance("paddleocr").ocr(arr, cls=True)
                return " ".join(l[1][0] for l in r[0]).upper() if r and r[0] else ""
            except Exception as e:
                logging.warning(f"PaddleOCR failed:{e}")
                if not fb:
                    return ""
        elif eng == "easyocr":
            try:
                return " ".join(self._get_engine_instance("easyocr").readtext(arr, detail=0)).upper()
            except Exception as e:
                logging.warning(f"EasyOCR failed:{e}")
                if not fb:
                    return ""
        psm = self.cfg["app_settings"].get("ocr_psm", 6)
        oem = self.cfg["app_settings"].get("ocr_oem", 3)
        return pytesseract.image_to_string(arr, config=f"--oem {oem} --psm {psm}").upper()

    def _extract_text(self, pdf_path: str) -> str:
        """
        Route to zone OCR or full OCR depending on the ocr_mode setting.
        'zone'  → fast targeted extraction (new)
        'full'  → original full-page extraction (legacy behaviour)
        """
        ocr_mode = self.cfg["app_settings"].get("ocr_mode", "full")
        if ocr_mode == "zone":
            return self._extract_text_zone(pdf_path)

        # ---- original full-page logic (unchanged) ----
        full = ""
        scale = self.cfg["app_settings"].get("image_scale_factor", 2)
        mode = self.cfg["app_settings"].get("extraction_source", "auto")
        use_native_flag = self.cfg["app_settings"].get("extract_text_before_ocr", True)

        try:
            doc = fitz.open(pdf_path)
            for page in doc:
                pt = ""
                use_native = (mode in ("auto", "pdf_text_only")) and use_native_flag
                if use_native:
                    pt = page.get_text("text").upper()

                need_ocr = False
                if mode == "image_only":
                    need_ocr = True
                elif mode == "auto":
                    if len(pt.strip()) < 50:
                        need_ocr = True
                elif mode == "pdf_text_only":
                    need_ocr = False

                if need_ocr:
                    scan_enabled = self.cfg["app_settings"].get("page_scan_region_enabled", False)
                    scan_pct     = self.cfg["app_settings"].get("page_scan_region_percent", 100)
                    if scan_enabled and scan_pct < 100:
                        rect  = page.rect
                        clip  = fitz.Rect(0, 0, rect.width, rect.height * scan_pct / 100.0)
                        pix   = page.get_pixmap(matrix=fitz.Matrix(scale, scale), clip=clip)
                    else:
                        pix   = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
                    arr = self._preprocess_image(pix)
                    arr = self._correct_rotation(arr)
                    pt = self._run_ocr(arr).upper()

                full += (pt or "") + "\n"
            doc.close()
        except Exception as e:
            logging.error(f"Text extraction failed [{pdf_path}]: {e}")
        return full

    def _normalize_text(self, t):
        return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9 ]", " ", t.upper())).strip()

    def _resolve_alias(self, name, aliases):
        for main, lst in aliases.items():
            if name == main or name in lst:
                return main
        return name
    def _append_ai_log(self, message: str):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        line = f"[{ts}] {message}"
        with self._ai_log_lock:
            self._ai_log_lines.append(line)

        logging.info(f"[AI-LOG] {message}")

        if self._ai_log_text is not None:
            try:
                self.after(0, lambda l=line: self._append_ai_log_to_widget(l))
            except Exception:
                pass

    def _append_ai_log_to_widget(self, line: str):
        if self._ai_log_text is None:
            return
        self._ai_log_text.configure(state="normal")
        self._ai_log_text.insert("end", line + "\n")
        self._ai_log_text.see("end")
        self._ai_log_text.configure(state="disabled")

    def _open_ai_log_window(self):
        if self._ai_log_window is not None and self._ai_log_window.winfo_exists():
            self._ai_log_window.lift()
            return

        win = tk.Toplevel(self)
        win.title("AI Supplier Matching Log")
        win.geometry("980x620")
        win.configure(bg=BG)

        self._ai_log_window = win

        top = tk.Frame(win, bg=PANEL2, height=46)
        top.pack(fill=tk.X)
        top.pack_propagate(False)

        tk.Label(
            top,
            text="AI Supplier Matching Log",
            bg=PANEL2,
            fg=TEXT,
            font=("Segoe UI", 11, "bold"),
        ).pack(side=tk.LEFT, padx=14, pady=12)

        btns = tk.Frame(top, bg=PANEL2)
        btns.pack(side=tk.RIGHT, padx=10, pady=6)

        ttk.Button(btns, text="Clear Log", command=self._clear_ai_log).pack(side=tk.LEFT, padx=4)
        ttk.Button(btns, text="Refresh", command=self._refresh_ai_log_window).pack(side=tk.LEFT, padx=4)

        body = tk.Frame(win, bg=BG)
        body.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)

        txt = tk.Text(
            body,
            bg=PANEL,
            fg=TEXT,
            insertbackground=TEXT,
            relief="flat",
            wrap="word",
            font=("Consolas", 9),
        )
        ys = ttk.Scrollbar(body, orient="vertical", command=txt.yview)
        xs = ttk.Scrollbar(body, orient="horizontal", command=txt.xview)
        txt.configure(yscrollcommand=ys.set, xscrollcommand=xs.set, wrap="none")

        txt.grid(row=0, column=0, sticky="nsew")
        ys.grid(row=0, column=1, sticky="ns")
        xs.grid(row=1, column=0, sticky="ew")
        body.rowconfigure(0, weight=1)
        body.columnconfigure(0, weight=1)

        self._ai_log_text = txt
        self._refresh_ai_log_window()

        def _on_close():
            self._ai_log_text = None
            self._ai_log_window = None
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_close)

    def _refresh_ai_log_window(self):
        if self._ai_log_text is None:
            return
        self._ai_log_text.configure(state="normal")
        self._ai_log_text.delete("1.0", "end")
        with self._ai_log_lock:
            for line in self._ai_log_lines:
                self._ai_log_text.insert("end", line + "\n")
        self._ai_log_text.see("end")
        self._ai_log_text.configure(state="disabled")

    def _clear_ai_log(self):
        with self._ai_log_lock:
            self._ai_log_lines.clear()
        if self._ai_log_text is not None:
            self._ai_log_text.configure(state="normal")
            self._ai_log_text.delete("1.0", "end")
            self._ai_log_text.configure(state="disabled")
        self._append_ai_log("AI log cleared by user.")

    # ------------- SUPPLIER MATCHING WITH CONFIDENCE -------------
    def _match_supplier_with_confidence(self, text, filename="") -> Tuple[str, float]:
        """Returns (supplier_name, confidence_0_to_100)."""
        suppliers = self.cfg.get("suppliers", [])
        aliases = self.cfg.get("aliases", {})
        cands = suppliers + [a for sub in aliases.values() for a in sub]
        thr = self.cfg["app_settings"].get("fuzzy_match_threshold", 95)

        if not cands:
            return "UNKNOWN SUPPLIER", 0.0

        upper = text.upper()
        unique_index = _build_unique_word_index(suppliers, aliases)

        best_unique_name = None
        best_unique_score = 0.0

        for c in cands:
            unique_words = unique_index.get(c, set())
            if not unique_words:
                continue
            hits = sum(
                1 for w in unique_words
                if re.search(r"\b" + re.escape(w) + r"\b", upper)
            )
            if hits > 0:
                score = hits / len(unique_words)
                if score > best_unique_score:
                    best_unique_score = score
                    best_unique_name = c

        if best_unique_name and best_unique_score >= 0.5:
            conf = min(100.0, best_unique_score * 100.0)
            return self._resolve_alias(best_unique_name, aliases), conf

        if filename:
            hint = re.sub(r"[_\-]", " ", re.split(r"GRN|RC-MAM", filename.upper())[0]).strip()
            m = process.extractOne(hint, cands, scorer=fuzz.token_sort_ratio)
            if m and m[1] >= thr:
                return self._resolve_alias(m[0], aliases), float(m[1])

        for c in cands:
            cn = self._normalize_text(c)
            if cn and cn in upper:
                return self._resolve_alias(c, aliases), 95.0

        word_freq_global = {}
        for c in cands:
            for w in re.sub(r"[^A-Z0-9 ]", " ", c.upper()).split():
                if len(w) >= 3:
                    word_freq_global[w] = word_freq_global.get(w, 0) + 1

        best_line_score = 0.0
        best_line_name = None
        for line in upper.splitlines():
            line = line.strip()
            if len(line) < 4:
                continue
            m = process.extractOne(line, cands, scorer=fuzz.partial_ratio)
            if m and m[1] >= 88:
                candidate = m[0]
                c_words = set(re.sub(r"[^A-Z0-9 ]", " ", candidate.upper()).split())
                uncommon_hits = sum(
                    1 for w in c_words
                    if len(w) >= 3 and word_freq_global.get(w, 0) == 1 and w in upper
                )
                if uncommon_hits > 0 or len(c_words) <= 2:
                    if m[1] > best_line_score:
                        best_line_score = m[1]
                        best_line_name = candidate

        if best_line_name:
            return self._resolve_alias(best_line_name, aliases), float(best_line_score)

        return "UNKNOWN SUPPLIER", 0.0

    def _match_supplier(self, text, filename="") -> str:
        name, _ = self._match_supplier_with_confidence(text, filename)
        return name

    def _extract_supplier_from_field(self, text):
        suppliers = self.cfg.get("suppliers", [])
        aliases = self.cfg.get("aliases", {})
        cands = suppliers + [a for sub in aliases.values() for a in sub]
        thr = self.cfg["app_settings"].get("fuzzy_match_threshold", 85)

        if not cands:
            return None

        upper = text.upper()

        patterns = [
            r"SUPPLIE(?:R)?\s*(?:NAME)?\s*[:\-]?",
            r"VENDOR\s*(?:NAME)?\s*[:\-]?",
            r"BILL\s*(?:TO|FROM)\s*[:\-]?",
            r"SOLD\s*(?:BY|TO)\s*[:\-]?",
        ]
        stop_words = [
            "INVOICE", "RECEIVING", "RECEIPT", "GRN", "DATE", "BILL",
            "AMOUNT", "SOURCE DOCUMENT", "SOURCE", "DIRCET", "DIRECT",
            "NO NOTES", "NOTES", "PURCHASE ORDER", "PURCHASE", "PO",
        ]

        raw_extracted_name = None

        for pat in patterns:
            m = re.search(pat, upper)
            if not m:
                continue

            chunk = upper[m.end(): m.end() + 800]

            for s in stop_words:
                pos = chunk.find(s)
                if 0 < pos < len(chunk):
                    chunk = chunk[:pos]

            chunk = re.sub(r"[^A-Z0-9\s]", " ", chunk)
            chunk = re.sub(r"\s+", " ", chunk).strip()

            if len(chunk) < 2:
                continue

            raw_extracted_name = chunk.strip()
            extracted_words = [w for w in chunk.split() if len(w) >= 2]
            if not extracted_words:
                continue

            first_extracted_word = extracted_words[0]
            if first_extracted_word.isdigit():
                for w in extracted_words[1:]:
                    if not w.isdigit():
                        first_extracted_word = w
                        break

            for c in cands:
                c_words = [w for w in re.sub(r"[^A-Z0-9 ]", " ", c.upper()).split() if len(w) >= 2]
                if not c_words:
                    continue
                first_cand_word = c_words[0]
                if first_extracted_word == first_cand_word:
                    return self._resolve_alias(c, aliases)
                if fuzz.ratio(first_extracted_word, first_cand_word) >= 88:
                    return self._resolve_alias(c, aliases)

            for c in cands:
                cn = self._normalize_text(c)
                if cn and cn in chunk:
                    return self._resolve_alias(c, aliases)

            mt = process.extractOne(chunk[:240], cands, scorer=fuzz.token_set_ratio)
            if mt and mt[1] >= thr:
                return self._resolve_alias(mt[0], aliases)

        if raw_extracted_name:
            cleaned = re.sub(r"\s+", " ", raw_extracted_name).strip()
            words = cleaned.split()[:6]
            return " ".join(words) if words else None

        return None

    def _extract_company_from_invoice_pages(self, pdf_path):
        suppliers = self.cfg.get("suppliers", [])
        aliases = self.cfg.get("aliases", {})
        cands = suppliers + [a for sub in aliases.values() for a in sub]
        thr = self.cfg["app_settings"].get("fuzzy_match_threshold", 85)
        scale = self.cfg["app_settings"].get("image_scale_factor", 2)
        use = self.cfg["app_settings"].get("extract_text_before_ocr", True)

        if not cands:
            return None

        try:
            doc = fitz.open(pdf_path)
            for page in doc:
                pt = page.get_text("text").upper() if use else ""
                if len(pt.strip()) < 50:
                    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
                    pt = self._run_ocr(self._preprocess_image(pix))

                if (
                    "RECEIVING REPORT" not in pt
                    and "RECEIVING RECORD" not in pt
                    and "INVOICE" not in pt
                ):
                    continue

                rect = page.rect
                h_limit = rect.height * 0.25
                third_w = rect.width / 3.0

                try:
                    words = page.get_text("words")
                    words_left   = [w[4] for w in words if w[1] < h_limit and w[0] < third_w]
                    words_center = [w[4] for w in words if w[1] < h_limit and third_w <= w[0] < 2 * third_w]
                    words_right  = [w[4] for w in words if w[1] < h_limit and w[0] >= 2 * third_w]
                    words_all    = [w[4] for w in words if w[1] < h_limit]
                except Exception:
                    words_left = words_center = words_right = words_all = []

                zones = [
                    ("above-left",   fitz.Rect(0, 0, third_w, h_limit),              " ".join(words_left)),
                    ("above-center", fitz.Rect(third_w, 0, 2 * third_w, h_limit),    " ".join(words_center)),
                    ("above-right",  fitz.Rect(2 * third_w, 0, rect.width, h_limit), " ".join(words_right)),
                    ("above-full",   fitz.Rect(0, 0, rect.width, h_limit),            " ".join(words_all)),
                ]

                for zone_label, clip, native_zone_text in zones:
                    zone_ocr_text = ""
                    try:
                        px = page.get_pixmap(matrix=fitz.Matrix(scale, scale), clip=clip)
                        if px.h > 0 and px.w > 0:
                            zone_ocr_text = self._run_ocr(self._preprocess_image(px))
                    except Exception as e:
                        logging.warning(f"Zone render error [{zone_label}]:{e}")

                    combined = (zone_ocr_text + "\n" + native_zone_text).upper()
                    for line in combined.splitlines():
                        line = line.strip()
                        if len(line) < 4:
                            continue
                        mt = process.extractOne(line, cands, scorer=fuzz.token_sort_ratio)
                        if mt and mt[1] >= thr:
                            doc.close()
                            return self._resolve_alias(mt[0], aliases)

            doc.close()
        except Exception as e:
            logging.error(f"Invoice page extraction failed [{pdf_path}]:{e}", exc_info=True)

        return None

    # ------------- RECEIVING REPORT / GRN / PO / DATE -------------
    def _is_receiving_report_text(self, text: str) -> bool:
        u = (text or "").upper()
        if (
            "RECEIVING REPORT" in u
            or "RECEIVING RECORD" in u
            or ("RECEIVING" in u and "RECORD" in u)
        ):
            return True
        # Tolerant fallback for OCR-damaged labels (staple holes, smudges) where
        # one or more characters of the label are unreadable.
        if self.cfg.get("app_settings", {}).get("receiving_label_fuzzy", True):
            return (
                _label_present(u, "RECEIVING RECORD", max_dist=2)
                or _label_present(u, "RECEIVING REPORT", max_dist=2)
            )
        return False

    def _collect_receiving_report_pages(self, pdf_path: str) -> List[Dict]:
        pages = []
        mode = self.cfg["app_settings"].get("extraction_source", "auto")
        scale = self.cfg["app_settings"].get("image_scale_factor", 2)
        use_native_flag = self.cfg["app_settings"].get("extract_text_before_ocr", True)

        try:
            doc = fitz.open(pdf_path)
            for i, page in enumerate(doc):
                native = ""
                ocr_text = ""

                use_native = (mode in ("auto", "pdf_text_only")) and use_native_flag
                if use_native:
                    native = page.get_text("text").upper()

                combined = native

                need_ocr = False
                if mode == "image_only":
                    need_ocr = True
                elif mode == "auto":
                    if not self._is_receiving_report_text(native):
                        need_ocr = True
                elif mode == "pdf_text_only":
                    need_ocr = False

                if need_ocr:
                    rect = page.rect
                    header_clip = fitz.Rect(0, 0, rect.width, rect.height * 0.28)
                    try:
                        pix_header = page.get_pixmap(matrix=fitz.Matrix(scale, scale), clip=header_clip)
                        if pix_header.w > 0 and pix_header.h > 0:
                            arr_header = self._preprocess_image(pix_header)
                            arr_header = self._correct_rotation(arr_header)
                            ocr_header = self._run_ocr(arr_header).upper()
                        else:
                            ocr_header = ""
                    except Exception:
                        ocr_header = ""

                    try:
                        pix_full = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
                        arr_full = self._preprocess_image(pix_full)
                        arr_full = self._correct_rotation(arr_full)
                        ocr_full = self._run_ocr(arr_full).upper()
                    except Exception:
                        ocr_full = ""

                    ocr_text = (ocr_header + "\n" + ocr_full).strip()

                combined = (native + "\n" + ocr_text).upper() if ocr_text else native

                if self._is_receiving_report_text(combined):
                    pages.append({
                        "index": i,
                        "text": combined,
                        "native_text": native,
                    })

            doc.close()
        except Exception as e:
            logging.error(f"Receiving page collection failed [{pdf_path}]:{e}", exc_info=True)

        return pages

    def _extract_date_from_receiving_text(self, text: str) -> str:
        u = text.upper()
        patterns = [
            r"RECEIVED\s*[OAU]N\s*[:\-]?\s*(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})",
            r"RECEIVED\s*DATE\s*[:\-]?\s*(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})",
            r"DATE\s*RECEIVED\s*[:\-]?\s*(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})",
            r"RECEIPT\s*DATE\s*[:\-]?\s*(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})",
        ]
        for pat in patterns:
            m = re.search(pat, u, re.IGNORECASE)
            if m:
                mm, dd, yyyy = m.group(1), m.group(2), m.group(3)
                if len(yyyy) == 2:
                    yyyy = "20" + yyyy
                return f"{dd.zfill(2)}.{mm.zfill(2)}.{yyyy}"
        return ""

    def _extract_po_from_receiving_text(self, text: str) -> str:
        u = text.upper()
        patterns = [
            r"PURCHASE\s*ORDER\s*(?:NO\.?|NUMBER|#)?\s*[:\-]?\s*(MAM[-\s]?\d+)",
            r"\bP\.?\s*O\.?\s*(?:NO\.?|NUMBER|#)?\s*[:\-]?\s*(MAM[-\s]?\d+)",
            r"PURCHASE\s*ORDER\s*(?:NO\.?|NUMBER|#)?\s*[:\-]?\s*(\d{5,})",
            r"\bP\.?\s*O\.?\s*(?:NO\.?|NUMBER|#)?\s*[:\-]?\s*(\d{5,})",
        ]
        for pat in patterns:
            m = re.search(pat, u, re.IGNORECASE)
            if m:
                val = re.sub(r"\s+", "", m.group(1))
                if re.fullmatch(r"\d{5,}", val):
                    return f"MAM-{val.zfill(9)}"
                if val.startswith("MAM"):
                    return val.replace(" ", "")
        m = re.search(r"\bMAM-\d+\b", u)
        if m:
            return m.group(0)
        return ""

    def _reconstruct_rcmam_digits(self, raw_digits: str) -> Optional[str]:
        pats = self.cfg.get("patterns", {})
        target_len = int(pats.get("grn_digits", 9))
        min_len = int(pats.get("grn_min_digits", 4))
        max_len = int(pats.get("grn_max_digits", 12))
        digits = re.sub(r"\D", "", raw_digits or "")

        if not digits or len(digits) < min_len:
            return None
        if len(digits) > max_len:
            digits = digits[:max_len]
        if len(digits) >= target_len:
            return digits[:target_len]
        if digits.startswith("0000") and len(digits) < target_len:
            return digits.ljust(target_len, "0")
        return digits.zfill(target_len)

    def _extract_grn_candidates_from_receiving_text(self, text: str) -> List[int]:
        nums: List[int] = []
        u = (text or "").upper()

        clean_lines = []
        for raw_line in u.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            clean_lines.append(line)

        for line in clean_lines:
            if "RECEIVING" in line and "RECORD" in line:
                m = re.search(
                    r"RECEIVING\s*RECORD\s*#?\s*[:\-]?\s*RC[-\s]*MAM[-\s]*([0-9]{4,12})",
                    line
                )
                if m:
                    d = self._reconstruct_rcmam_digits(m.group(1))
                    if d:
                        nums.append(int(d))
                    continue

        for line in clean_lines:
            if "RECEIVING" in line and "RECORD" in line:
                cut = line
                po_pos = len(cut)
                for token in ["PURCHASE ORDER", "P.O", "PO NO", "PO#", "PO "]:
                    pos = cut.find(token)
                    if pos != -1:
                        po_pos = min(po_pos, pos)
                cut = cut[:po_pos]
                m = re.search(r"RC[-\s]*MAM[-\s]*([0-9]{4,12})", cut)
                if m:
                    d = self._reconstruct_rcmam_digits(m.group(1))
                    if d:
                        nums.append(int(d))

        for line in clean_lines:
            if "RECEIVING" not in line or "RECORD" not in line:
                continue
            cut = line
            po_pos = len(cut)
            for token in ["PURCHASE ORDER", "P.O", "PO NO", "PO#", "PO "]:
                pos = cut.find(token)
                if pos != -1:
                    po_pos = min(po_pos, pos)
            cut = cut[:po_pos]
            cut_ocr = cut.replace("O", "0")
            for m in re.finditer(r"R[C0][-\s]*M[A4]M[-\s]*([0-90]{4,12})", cut_ocr):
                raw = m.group(1).replace("O", "0")
                d = self._reconstruct_rcmam_digits(raw)
                if d:
                    nums.append(int(d))

        for line in clean_lines:
            if "RECEIVING" not in line or "RECORD" not in line:
                continue
            if "RC" not in line and "MAM" not in line:
                continue
            cut = line
            po_pos = len(cut)
            for token in ["PURCHASE ORDER", "P.O", "PO NO", "PO#", "PO "]:
                pos = cut.find(token)
                if pos != -1:
                    po_pos = min(po_pos, pos)
            cut = cut[:po_pos]
            m = re.search(r"(?:RC|R0)[-\s]*(?:MAM|MA4M|M4M)[-\s]*([0-9]{4,12})", cut)
            if m:
                d = self._reconstruct_rcmam_digits(m.group(1))
                if d:
                    nums.append(int(d))

        # Tolerant fallback: if the gated passes above found nothing (e.g. the
        # 'RECEIVING RECORD' label was damaged by a staple hole, or a digit in
        # the code was misread as a letter), scan the whole text for an RC-MAM
        # token with OCR-confusion repair on the numeric part.
        if not nums:
            nums.extend(self._extract_rcmam_tolerant(u))

        return sorted(set(nums))

    def _extract_rcmam_tolerant(self, text: str) -> List[int]:
        """Best-effort RC-MAM number recovery that tolerates OCR damage.

        Finds 'RC...MAM' (allowing common prefix confusions) and repairs the
        following number region by mapping look-alike letters back to digits
        (O->0, I/L->1, S->5, B->8, ...). Used only as a fallback so it never
        overrides a clean match.
        """
        out: List[int] = []
        u = (text or "").upper()
        for m in re.finditer(r"R[C0G6]\s*[-\s]?\s*M[A4]M", u):
            tail = u[m.end(): m.end() + 18]
            tail = re.sub(r"^[\s:\-#.]+", "", tail)
            region = tail[:14]
            # Stop the number region at a clear word break (2+ spaces) or a
            # following field label, so we don't slurp unrelated characters.
            region = re.split(r"\s{2,}|PURCHASE|\bPO\b|P\.O", region)[0]
            fixed = region.translate(_OCR_DIGIT_FIX)
            digits = re.sub(r"\D", "", fixed)
            d = self._reconstruct_rcmam_digits(digits)
            if d:
                out.append(int(d))
        return out

    def _extract_grn_from_filename_old(self, p):
        base = os.path.splitext(os.path.basename(p))[0].upper()
        m = re.search(r"(RC-MAM-\d+(?:-\d+)*)", base)
        return m.group(1) if m else ""

    def _build_grn_chain(self, nums: List[int]) -> str:
        if not nums:
            return ""
        pats = self.cfg.get("patterns", {})
        prefix = pats.get("grn_prefix", "RC-MAM-")
        target_len = int(pats.get("grn_digits", 9))
        nums = sorted(set(nums))
        first = str(nums[0]).zfill(target_len)
        if len(nums) == 1:
            return f"{prefix}{first}"
        tail = "-".join(str(n) for n in nums[1:])
        return f"{prefix}{first}-{tail}"

    def _extract_grn_full(self, pdf_path: str, text: str) -> str:
        rr_pages = self._collect_receiving_report_pages(pdf_path)
        nums: List[int] = []
        for pg in rr_pages:
            nums.extend(self._extract_grn_candidates_from_receiving_text(pg.get("text", "")))
        if nums:
            return self._build_grn_chain(nums)
        fn_grn = self._extract_grn_from_filename_old(pdf_path)
        if fn_grn:
            return fn_grn
        return ""

    def _extract_receiving_report_fields(self, pdf_path: str) -> Dict:
        pages = self._collect_receiving_report_pages(pdf_path)
        return self._parse_receiving_report_pages(pages)

    def _extract_receiving_report_fields_from_text(self, text: str) -> Dict:
        """Fast path: parse receiving-report fields straight from text that has
        ALREADY been OCR'd (e.g. OCR.space output) without re-opening the PDF or
        running local OCR again. Used by the AI Extract / API auto-ingest flow.

        This is what makes API extraction fast: the previous code called
        _collect_receiving_report_pages(), which re-rendered every page and ran
        local OCR twice per page (header + full) at the configured scale - and it
        was invoked twice per file. Here we reuse the OCR.space text instead.
        """
        raw = text or ""
        # Use form-feed page breaks if the extractor provided them, else 1 page.
        chunks = raw.split("\f") if "\f" in raw else [raw]
        pages: List[Dict] = []
        for i, chunk in enumerate(chunks):
            up = chunk.upper()
            if up.strip():
                pages.append({"index": i, "text": up, "native_text": up})
        # Mirror offline behaviour: prefer pages detected as receiving reports,
        # but fall back to all pages if the detector matches none (OCR variance).
        rr_pages = [p for p in pages if self._is_receiving_report_text(p["text"])]
        return self._parse_receiving_report_pages(rr_pages if rr_pages else pages)

    def _parse_receiving_report_pages(self, pages: List[Dict]) -> Dict:
        all_nums: List[int] = []
        best_date = ""
        best_po = ""
        sums = {"USD": 0.0, "MVR": 0.0, "EUR": 0.0, "GBP": 0.0, "SGD": 0.0}

        def _apply(cur_pre, amount_str, cur_post):
            cur = (cur_pre or cur_post or "").upper().strip()
            currency = _CURRENCY_MAP.get(cur)
            if not currency:
                return
            try:
                val = float(amount_str.replace(",", ""))
                sums[currency] += val
            except ValueError:
                pass

        seen_totals = set()

        for pg in pages:
            txt = (pg.get("text") or "").upper()
            page_index = pg.get("index", 0)

            nums = self._extract_grn_candidates_from_receiving_text(txt)
            all_nums.extend(nums)

            if not best_date:
                best_date = self._extract_date_from_receiving_text(txt)
            if not best_po:
                best_po = self._extract_po_from_receiving_text(txt)

            # Apply OCR currency correction before pattern matching
            txt_fixed = _fix_ocr_currency_amount(txt)

            for m in _REPORT_INVOICE_TOTAL_PAT.finditer(txt_fixed):
                cur_pre = m.group(1)
                amount_str = m.group(2)
                cur_post = m.group(3)
                cur = (cur_pre or cur_post or "").upper().strip()
                # Normalize MYR → MVR
                if cur == "MYR":
                    cur = "MVR"
                currency = _CURRENCY_MAP.get(cur)
                if not currency:
                    continue
                dedup_key = (page_index, currency, amount_str.replace(",", ""))
                if dedup_key in seen_totals:
                    continue
                seen_totals.add(dedup_key)
                try:
                    val = float(amount_str.replace(",", ""))
                    sums[currency] += val
                except ValueError:
                    pass

        formatted_totals = {}
        for k, v in sums.items():
            formatted_totals[k] = f"{v:.2f}" if v > 0 else ""

        return {
            "pages_found": len(pages),
            "grn": self._build_grn_chain(all_nums),
            "date": best_date,
            "po": best_po,
            "totals": formatted_totals,
        }

    # ------------- INVOICE EXTRACTION -------------
    def _extract_invoice(self, text, supplier_hint=""):
        raw = text.upper()
        if supplier_hint:
            hint_upper = supplier_hint.upper()
            for key, patterns in SUPPLIER_INVOICE_HINTS.items():
                if key.upper() in hint_upper:
                    for pat_str in patterns:
                        for ctx in [
                            rf"INVOICE\s*(?:NUMBER|NUM(?:BER)?|NO\.?|#)?\s*[:\-]\s*({pat_str})",
                            rf"(?:INV|BILL)\s*(?:NO\.?|#)?\s*[:\-]\s*({pat_str})",
                        ]:
                            try:
                                m = re.search(ctx, raw)
                                if m:
                                    v = _clean_invoice_token(m.group(1))
                                    if v:
                                        v = _fix_invoice_prefix(v, self.cfg.get("invoice_prefix_fixes", {}))
                                        return v.replace("/", " ").replace("\\", " ").strip()
                            except re.error:
                                pass
                    break

        v = _find_invoice_value(raw)
        v = _clean_invoice_token(v) if v else ""
        if v:
            v = _fix_invoice_prefix(v, self.cfg.get("invoice_prefix_fixes", {}))
            return v.replace("/", " ").replace("\\", " ").strip()
        return None

    def _extract_invoice_old(self, text):
        raw = text.upper()
        v = _find_invoice_value(raw)
        v = _clean_invoice_token(v) if v else ""
        return v or ""

    # ------------- LEGACY HELPERS -------------
    def _legacy_load_suppliers(self):
        base = self.dirs.get("base", "")
        txt = os.path.join(base, "suppliers.txt")
        if os.path.exists(txt):
            with open(txt, "r", encoding="utf-8") as f:
                sups = [x.strip().upper() for x in f if x.strip()]
        else:
            sups = [s.upper() for s in self.cfg.get("suppliers", [])]
        return sups, self.cfg.get("aliases", {})

    def _legacy_normalize(self, t):
        return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9 ]", " ", t.upper())).strip()

    def _legacy_extract_text(self, pdf_path):
        full = ""
        try:
            doc = fitz.open(pdf_path)
            for page in doc:
                rect = page.rect
                crop = fitz.Rect(0, 0, rect.width, min(rect.height, 650))
                pix = page.get_pixmap(matrix=fitz.Matrix(3, 3), clip=crop)
                img = PILImage.open(io.BytesIO(pix.tobytes("png")))
                full += pytesseract.image_to_string(img, config="--psm 6 --oem 3") + "\n"
                img.close()
            doc.close()
        except Exception as e:
            logging.error(f"Legacy text extraction [{pdf_path}]:{e}")
        return full.upper()

    def _legacy_best_supplier_match(self, text, cands, aliases):
        text = self._legacy_normalize(text)
        words = set(text.split())
        best_name = None
        best_score = 0.0
        for c in cands:
            cn = self._legacy_normalize(c)
            if not cn:
                continue
            if cn in text:
                return self._resolve_alias(c, aliases)
            cn_w = set(cn.split())
            overlap = len(words & cn_w) / max(len(cn_w), 1)
            fuzzy = SequenceMatcher(None, text[:2000], cn).ratio()
            score = max(overlap, fuzzy)
            for alias in aliases.get(c, []):
                an = self._legacy_normalize(alias)
                if an and an in text:
                    return self._resolve_alias(c, aliases)
                an_w = set(an.split())
                score = max(
                    score,
                    len(words & an_w) / max(len(an_w), 1),
                    SequenceMatcher(None, text[:2000], an).ratio(),
                )
            if score > best_score:
                best_score = score
                best_name = c
        return self._resolve_alias(best_name, aliases) if best_score >= 0.25 and best_name else None

    def _legacy_extract_supplier(self, text, cands, aliases):
        m = re.search(r"SUPPLIE(?:R)?\s*[:\-]?\s*", text, re.IGNORECASE)
        if m:
            chunk = text[m.end(): m.end() + 600]
            for kw in ["INVOICE", "RECEIVING", "GRN", "DATE", "BILL", "AMOUNT", "SOURCE", "DIRCET", "DIRECT"]:
                pos = chunk.find(kw)
                if 0 < pos < len(chunk):
                    chunk = chunk[:pos]
            chunk = self._legacy_normalize(chunk.replace("\n", " "))
            r = self._legacy_best_supplier_match(chunk, cands, aliases)
            if r:
                return r
        return self._legacy_best_supplier_match(text, cands, aliases)

    # ------------- PROCESSING MODES -------------
    def _process_file_legacy(self, pdf_path):
        cands, aliases = self._legacy_load_suppliers()
        text = self._legacy_extract_text(pdf_path)
        supplier = self._legacy_extract_supplier(text, cands, aliases) or "UNKNOWN SUPPLIER"
        grn = self._extract_grn_full(pdf_path, text) or "NO-GRN"
        invoice = self._extract_invoice_old(text) or "NO-INVOICE"
        if supplier != "UNKNOWN SUPPLIER":
            _set_last_good_supplier(supplier)
        elif _get_last_good_supplier():
            supplier = _get_last_good_supplier()
        return {"supplier": supplier, "grn": grn, "invoice": invoice, "confidence": 70.0}

    def _process_file_custom(self, pdf_path, text):
        method = self.cfg.get("app_settings", {}).get("supplier_extraction_method", "both")
        supplier = None
        confidence = 0.0
        if method in ("receiving_report_field", "both"):
            supplier = self._extract_supplier_from_field(text)
            if supplier:
                confidence = 85.0
        if not supplier and method in ("header_company_name", "both"):
            supplier = self._extract_company_from_invoice_pages(pdf_path)
            if supplier:
                confidence = 80.0
        if not supplier:
            supplier, confidence = self._match_supplier_with_confidence(text, os.path.basename(pdf_path))
        if not supplier or supplier == "UNKNOWN SUPPLIER":
            supplier = _get_last_good_supplier() or "UNKNOWN SUPPLIER"
            confidence = 0.0
        else:
            _set_last_good_supplier(supplier)
        grn = self._extract_grn_full(pdf_path, text) or "NO-GRN"
        invoice = self._extract_invoice(text, supplier_hint=supplier) or "NO-INVOICE"
        return {"supplier": supplier, "grn": grn, "invoice": invoice, "confidence": confidence}

    def _process_file_mixed(self, pdf_path, text):
        method = self.cfg.get("app_settings", {}).get("supplier_extraction_method", "both")
        supplier = None
        confidence = 0.0
        if method in ("receiving_report_field", "both"):
            supplier = self._extract_supplier_from_field(text)
            if supplier:
                confidence = 85.0
        if not supplier and method in ("header_company_name", "both"):
            supplier = self._extract_company_from_invoice_pages(pdf_path)
            if supplier:
                confidence = 80.0
        if not supplier:
            supplier, confidence = self._match_supplier_with_confidence(text, os.path.basename(pdf_path))
        if not supplier or supplier == "UNKNOWN SUPPLIER":
            supplier = _get_last_good_supplier() or "UNKNOWN SUPPLIER"
            confidence = 0.0
        else:
            _set_last_good_supplier(supplier)
        grn = self._extract_grn_full(pdf_path, text) or "NO-GRN"
        legacy_text = self._legacy_extract_text(pdf_path)
        invoice = self._extract_invoice_old(legacy_text) or "NO-INVOICE"
        return {"supplier": supplier, "grn": grn, "invoice": invoice, "confidence": confidence}

    # ------------- FILENAME / SAFE NAME -------------
    @staticmethod
    def _safe_filename(n):
        n = re.sub(r"[/\\]", " ", n)
        n = re.sub(r"\s+", " ", n)
        return re.sub(r'[<>:"|?*]', "", n).strip()

    @staticmethod
    def _extract_scan_number(path: str) -> int:
        """
        Extract numeric order from filenames like:
        SCAN_0040.pdf -> 40
        SCAN-0039.pdf -> 39

        Files without a scan number go to the end.
        """
        name = os.path.basename(path).upper()
        m = re.search(r"SCAN[_\- ]*(\d+)", name)
        if m:
            try:
                return int(m.group(1))
            except ValueError:
                pass

        m = re.search(r"(\d+)", name)
        if m:
            try:
                return int(m.group(1))
            except ValueError:
                pass

        return -1

    def _get_pdf_files_strict_order(self, folder: str) -> List[str]:
        """
        Return PDFs in strict descending filename-number order:
        SCAN_0040, SCAN_0039, ..., SCAN_0001
        """
        if not folder or not os.path.isdir(folder):
            return []

        files = [
            os.path.join(folder, f)
            for f in os.listdir(folder)
            if f.lower().endswith(".pdf")
        ]

        files.sort(
            key=lambda p: (self._extract_scan_number(p), os.path.basename(p).upper()),
            reverse=True,
        )
        return files

    def _extract_totals_from_pdf_old(self, pdf_path):
        rr = self._extract_receiving_report_fields(pdf_path)
        return rr.get("totals", {"USD": "", "MVR": "", "EUR": "", "GBP": "", "SGD": ""})

    # ------------- DUPLICATE INVOICE DETECTION -------------
    def _find_duplicate_invoice(self, invoice: str, processed_folder: str) -> List[str]:
        if not invoice or invoice == "NO-INVOICE":
            return []
        dupes = []
        try:
            inv_clean = invoice.strip().upper()
            for fn in os.listdir(processed_folder):
                if fn.lower().endswith(".pdf") and inv_clean in fn.upper():
                    dupes.append(fn)
        except Exception:
            pass
        return dupes

    # ------------- RENAME SINGLE FILE -------------
    def _rename_single_file(self, pdf_path, processed, failed, mode):
        fname = os.path.basename(pdf_path)
        rd = {
            "file": fname, "new_name": "", "supplier": "UNKNOWN SUPPLIER",
            "grn": "", "invoice": "", "status": "error",
            "dest_path": "", "duplicate_warning": "", "confidence": 0.0,
        }
        try:
            if mode == "legacy":
                fields = self._process_file_legacy(pdf_path)
            else:
                text = self._extract_text(pdf_path)
                if mode == "mixed":
                    fields = self._process_file_mixed(pdf_path, text)
                else:
                    fields = self._process_file_custom(pdf_path, text)

            sup  = fields["supplier"] or "UNKNOWN SUPPLIER"
            grn  = fields["grn"] or "RC-MAM-0000"
            inv  = fields["invoice"] or "NO-INVOICE"
            conf = fields.get("confidence", 0.0)

            filename_invoice = f"IN {inv}" if inv and inv != "NO-INVOICE" else "NO-INVOICE"

            dupes = self._find_duplicate_invoice(inv, processed)
            if dupes:
                rd["duplicate_warning"] = f"Invoice already exists: {dupes[0]}"
                logging.warning(f"Duplicate invoice [{inv}] found for [{fname}]: {dupes}")

            dry_run = self.cfg.get("app_settings", {}).get("dry_run", False)
            dest = os.path.join(processed, self._safe_filename(f"{sup} GRN {grn} {filename_invoice}.pdf"))

            cnt = 1
            base, ext = os.path.splitext(dest)
            while os.path.exists(dest) and not dry_run:
                dest = f"{base}_{cnt}{ext}"
                cnt += 1

            if not dry_run:
                shutil.move(pdf_path, dest)
                rd["status"] = "success"
                rd["dest_path"] = dest
            else:
                rd["status"] = "simulated"
                rd["dest_path"] = pdf_path

            rd["new_name"]   = os.path.basename(dest)
            rd["supplier"]   = sup
            rd["grn"]        = grn
            rd["invoice"]    = inv if inv != "NO-INVOICE" else ""
            rd["confidence"] = conf

        except Exception as e:
            logging.error(f"Rename failed [{fname}]:{e}", exc_info=True)
            rd["status"] = "error"
            try:
                shutil.move(pdf_path, os.path.join(failed, fname))
            except Exception:
                pass
        return rd

    # ------------- DISPATCH SINGLE FILE -------------
    def _dispatch_single_file(self, pdf_path, scan_index=0, mode=None):
        fn = os.path.basename(pdf_path)
        if mode is None:
            mode = self.cfg.get("app_settings", {}).get("processing_mode", "legacy")
        try:
            core = self._extract_core_fields_for_file(pdf_path, mode)
            return {
                "doc_id": f"{fn}|dispatch|{scan_index}",
                "file": fn,
                "date": core.get("date", ""),
                "supplier": core.get("supplier", "UNKNOWN SUPPLIER"),
                "po": core.get("po", "") or "MAM-0000",
                "invoice": core.get("invoice_dispatch", ""),
                "usd": core.get("usd", ""),
                "mvr": core.get("mvr", ""),
                "eur": core.get("eur", ""),
                "gbp": core.get("gbp", ""),
                "sgd": core.get("sgd", ""),
                "grn": core.get("grn", "RC-MAM-0000"),
                "confidence": core.get("confidence", 0.0),
                "is_valid": True, "errors": "",
                "raw_path": pdf_path, "scan_index": scan_index,
            }
        except Exception as e:
            logging.error(f"Dispatch error [{fn}]:{e}", exc_info=True)
            return {
                "doc_id": f"{fn}|dispatch|{scan_index}",
                "file": fn, "date": "", "supplier": "", "po": "MAM-0000",
                "invoice": "", "usd": "", "mvr": "", "eur": "", "gbp": "", "sgd": "",
                "grn": "", "confidence": 0.0, "is_valid": False, "errors": str(e),
                "raw_path": pdf_path, "scan_index": scan_index,
            }


# ---------------------------------------------------------------------------
# RESOURCE HELPER
# ---------------------------------------------------------------------------
def resource_path(relative_path):
    try:
        base = Path(sys._MEIPASS)
    except Exception:
        base = Path(__file__).parent
    return base / relative_path


# ---------------------------------------------------------------------------
# GUI APPLICATION
# ---------------------------------------------------------------------------
class MaafushivaruHub(tk.Tk, OCRWorkerMixin):
    APP_VERSION = APP_VERSION

    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        try:
            ip = resource_path("logo.ico")
            if ip.exists():
                self.iconbitmap(str(ip))
        except Exception:
            pass
        self.geometry("1440x920")
        self.minsize(1120, 720)
        self.configure(bg=BG)
        # ✅ Start maximized so nothing is clipped on scaled displays
        self.state("zoomed")  # Windows; use "normal" + wm_attributes on Linux
        # ... rest of __init__ unchanged ...

        # State
        self._worker_running   = False
        self._dispatch_running = False
        self._cancel_requested = False
        self._rename_results:   List[Dict] = []
        self._dispatch_results: List[Dict] = []
        self._rename_queue   = queue.Queue()
        self._dispatch_queue = queue.Queue()
        self._rename_row_map:   Dict[str, Dict] = {}
        self._dispatch_row_map: Dict[str, Dict] = {}
        self._rename_all_rows:   List[str] = []
        self._dispatch_all_rows: List[str] = []

        # Watchdog
        self._watcher_observer: Optional["Observer"] = None
        self._watcher_handler:  Optional[ScannedFolderHandler] = None
        self._watcher_queue = queue.Queue()

        self.config_path = self._find_config()
        self.cfg = self._load_config()
        self.dirs = self._resolve_dirs()
        self._ensure_dirs()
        self._configure_tesseract()
        self._setup_logging()

        # Tk variables
        self._status_var = tk.StringVar(value="Ready — waiting for files.")
        self._engine_badge_var = tk.StringVar(value="")
        self._watcher_badge_var = tk.StringVar(value="")
        self._threads_var = tk.BooleanVar(
            value=self.cfg.get("app_settings", {}).get("enable_multi_threading", True)
        )
        self._dry_run_var = tk.BooleanVar(
            value=self.cfg.get("app_settings", {}).get("dry_run", False)
        )
        self._notify_var = tk.BooleanVar(
            value=self.cfg.get("app_settings", {}).get("desktop_notifications", True)
        )
        # Two mutually-exclusive auto-ingest modes (replace the legacy single
        # "auto_ingest_watcher" flag):
        #   offline -> OCR rename + GRN dispatch, fully local
        #   api     -> AI Extract (OCR.space) with size-check + auto compress
        self._watcher_offline_var = tk.BooleanVar(
            value=self.cfg.get("app_settings", {}).get("auto_ingest_offline", False)
        )
        self._watcher_api_var = tk.BooleanVar(
            value=self.cfg.get("app_settings", {}).get("auto_ingest_api", False)
        )
        # One-shot hook the AI Extract poller calls when an auto run finishes.
        self._aix_on_complete = None
        self._conf_threshold_var = tk.IntVar(
            value=self.cfg.get("app_settings", {}).get("confidence_warn_threshold", 80)
        )

        # AI supplier matcher + log store
        self._ai_matcher = None
        self._ai_log_lines: List[str] = []
        self._ai_log_lock = threading.Lock()
        self._ai_log_window = None
        self._ai_log_text = None
        if AI_MATCHER_AVAILABLE:
            try:
                from ai_supplier_matcher import AISupplierMatcher
                self._ai_matcher = AISupplierMatcher(self.cfg, logger_func=self._append_ai_log)
            except Exception as e:
                logging.warning(f"AI matcher init failed: {e}")
        self._ai_status_var = tk.StringVar(value="")
        self._build_style()
        self._build_layout()
        self._refresh_engine_badge()
        self._refresh_dashboard_stats()

        # Start watcher if either auto-ingest mode was enabled on last run
        if self._watcher_offline_var.get() or self._watcher_api_var.get():
            self.after(500, self._start_watcher)

        # Poll watcher queue
        self.after(500, self._poll_watcher_queue)

    # ------------------------------------------------------------------
    # CONFIG / PATH / LOGGING
    # ------------------------------------------------------------------
    def _find_config(self):
        cwd = Path.cwd() / "config.json"
        scr = Path(__file__).parent / "config.json"
        return cwd if cwd.exists() else scr

    def _load_config(self):
        default = {
            "folders": {
                "base": ".",
                "scanned": "SCANNED",
                "processed": "PROCESSED",
                "archive": "ARCHIVE",
                "failed": "FAILED",
                "logs": "LOGS"
            },
            "suppliers": [],
            "aliases": {},
            "invoice_prefix_fixes": {
                "MSL": "MSI",
                "MS1": "MSI"
            },
            "patterns": {
                "grn_prefix": "RC-MAM-",
                "grn_digits": 9,
                "grn_min_digits": 4,
                "grn_max_digits": 12,
                "po_prefix": "MAM-",
                "po_min_digits": 5
            },
            "app_settings": {
                "processing_mode": "custom",
                "supplier_extraction_method": "both",
                "dry_run": False,
                "desktop_notifications": True,
                "auto_ingest_watcher": False,
                "auto_ingest_offline": False,
                "auto_ingest_api": False,
                "receiving_label_fuzzy": True,
                "delete_temp_after_send": True,
                "confidence_warn_threshold": 80,
                "page_scan_region_enabled": False,
                "page_scan_region_percent": 100,
                "supplier_match_strategy": "combined",
                "ocr_mode": "zone",
                "ocr_word_correction_enabled": False,
                "smart_cross_match_enabled": False,
                "ocr_engine": "tesseract",
                "extract_text_before_ocr": True,
                "enhance_images": True,
                "ocr_fallback_to_tesseract": True,
                "enable_multi_threading": True,
                "max_threads": 4,
                "fuzzy_match_threshold": 85,
                "image_scale_factor": 2,
                "extraction_source": "auto"
            },
            "ai_settings": {
                "enabled": False,
                "provider": "openai",
                "model": "gpt-4o-mini",
                "api_key": "",
                "custom_base_url": "",
                "timeout_seconds": 20
            },
            "ocr_space": {
                "api_key": "K88109865088957",
                "language": "eng",
                "isOverlayRequired": False,
                "detectOrientation": True,
                "scale": True,
                "OCREngine": 2,
                "isTable": False,
                "filetype": "PDF",
                "timeout_seconds": 30,
                "max_upload_mb": 1.0
            }
        }

        if not self.config_path.exists():
            return default

        try:
            with open(self.config_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
        except Exception as e:
            logging.warning(f"Config load failed, using defaults: {e}")
            return default

        for key, value in default.items():
            if key not in cfg:
                cfg[key] = value
            elif isinstance(value, dict):
                for subkey, subvalue in value.items():
                    if subkey not in cfg[key]:
                        cfg[key][subkey] = subvalue

        # --- Backward-compat migration -------------------------------------
        # The old single "auto_ingest_watcher" flag is replaced by two
        # explicit modes. If the legacy flag was on and neither new mode is
        # set, default it to the OFFLINE pipeline (its previous behaviour).
        s = cfg.setdefault("app_settings", {})
        if s.get("auto_ingest_watcher") and not s.get("auto_ingest_offline") and not s.get("auto_ingest_api"):
            s["auto_ingest_offline"] = True
        # Safety: the two modes are mutually exclusive. If both somehow ended
        # up enabled, API wins and offline is cleared.
        if s.get("auto_ingest_offline") and s.get("auto_ingest_api"):
            s["auto_ingest_offline"] = False

        return cfg


    def _save_config(self):
        with open(self.config_path, "w", encoding="utf-8") as f:
            json.dump(self.cfg, f, indent=4)
        self._set_status("Settings saved.")

    def _resolve_base(self):
        base = self.cfg.get("folders", {}).get("base", ".")
        cd = self.config_path.parent.resolve()
        if not base or str(base).strip() in (".", "./", ".\\"):
            return str(cd)
        p = Path(base)
        return str(p) if p.is_absolute() else str(cd / base)

    def _resolve_dirs(self):
        base = self._resolve_base()
        f = self.cfg.get("folders", {})
        return {
            "base":      base,
            "scanned":   os.path.join(base, f.get("scanned",   "SCANNED")),
            "processed": os.path.join(base, f.get("processed", "PROCESSED")),
            "archive":   os.path.join(base, f.get("archive",   "ARCHIVE")),
            "failed":    os.path.join(base, f.get("failed",    "FAILED")),
            "logs":      os.path.join(base, f.get("logs",      "LOGS")),
        }

    def _ensure_dirs(self):
        for k in ("scanned", "processed", "archive", "failed", "logs"):
            os.makedirs(self.dirs[k], exist_ok=True)

    def _configure_tesseract(self):
        cmd = self.cfg.get("tesseract_cmd")
        if cmd:
            pytesseract.pytesseract.tesseract_cmd = cmd

    def _setup_logging(self):
        os.makedirs(self.dirs["logs"], exist_ok=True)
        logging.basicConfig(
            filename=os.path.join(self.dirs["logs"], "maafushivaru_hub.log"),
            level=logging.INFO,
            format="%(asctime)s [%(levelname)s] %(message)s",
        )
        logging.info("Application started.")

    # ------------------------------------------------------------------
    # WATCHDOG AUTO-INGEST
    # ------------------------------------------------------------------
    def _start_watcher(self):
        if not WATCHDOG_AVAILABLE:
            messagebox.showwarning(
                "Watchdog Not Installed",
                "Auto-ingest requires the watchdog library.\n\nInstall it with:\n  pip install watchdog"
            )
            self._watcher_offline_var.set(False)
            self._watcher_api_var.set(False)
            return

        scanned = self.dirs.get("scanned", "")
        if not os.path.isdir(scanned):
            messagebox.showerror("Watcher Error", f"SCANNED folder not found:\n{scanned}")
            self._watcher_offline_var.set(False)
            self._watcher_api_var.set(False)
            return

        if self._watcher_observer and self._watcher_observer.is_alive():
            return  # already running

        def on_new_pdf(path: str):
            logging.info(f"[WATCHER] New PDF detected: {path}")
            self._watcher_queue.put(path)

        self._watcher_handler = ScannedFolderHandler(on_new_pdf=on_new_pdf, debounce_seconds=3.0)
        adapter = _WatchdogEventAdapter(self._watcher_handler)

        self._watcher_observer = Observer()
        self._watcher_observer.schedule(adapter, scanned, recursive=False)
        self._watcher_observer.start()

        self._watcher_badge_var.set(f"● AUTO-INGEST: {self._active_ingest_mode().upper()}")
        self._set_status(
            f"Auto-ingest ({self._active_ingest_mode()}) watching: {scanned}"
        )
        logging.info(
            f"[WATCHER] Started watching {scanned} in {self._active_ingest_mode()} mode"
        )

    def _stop_watcher(self):
        if self._watcher_observer:
            try:
                self._watcher_observer.stop()
                self._watcher_observer.join(timeout=3)
            except Exception as e:
                logging.warning(f"[WATCHER] Stop error: {e}")
            self._watcher_observer = None
        self._watcher_badge_var.set("")
        self._set_status("Auto-ingest stopped.")

    def _active_ingest_mode(self) -> str:
        """Return the currently selected auto-ingest mode: 'api', 'offline' or 'off'."""
        if self._watcher_api_var.get():
            return "api"
        if self._watcher_offline_var.get():
            return "offline"
        return "off"

    def _poll_watcher_queue(self):
        """Check for new files detected by watchdog, then route to the active mode.

        OFFLINE -> local OCR rename + GRN dispatch (_start_extract_and_process)
        API     -> AI Extract (OCR.space): size-check, auto-compress >threshold,
                   then auto send-to-tabs. Originals are never modified; they are
                   only renamed when results are pushed to the tabs.
        """
        try:
            while True:
                pdf_path = self._watcher_queue.get_nowait()
                if not os.path.exists(pdf_path):
                    continue

                mode = self._active_ingest_mode()
                busy = self._worker_running or self._dispatch_running or getattr(self, "_aix_running", False)
                if busy:
                    # A run is already in progress. Re-queue this file and try
                    # again on the next poll so nothing is dropped.
                    self._watcher_queue.put(pdf_path)
                    break

                fname = os.path.basename(pdf_path)
                try:
                    size_mb = os.path.getsize(pdf_path) / (1024 * 1024)
                except OSError:
                    size_mb = 0.0

                if mode == "api":
                    self._set_status(
                        f"[AUTO-INGEST API] Detected: {fname} ({size_mb:.2f} MB) — extracting...",
                        ACCENT,
                    )
                    self.after(200, self._auto_ingest_api)
                elif mode == "offline":
                    self._set_status(
                        f"[AUTO-INGEST OFFLINE] Detected: {fname} — processing...",
                        ACCENT,
                    )
                    self.after(200, self._start_extract_and_process)
                # mode == "off": ignore (watcher should not be running, but be safe)
        except queue.Empty:
            pass
        self.after(1500, self._poll_watcher_queue)

    def _auto_ingest_api(self):
        """Unattended AI Extract run for the API auto-ingest mode.

        Delegates to the AI Extract tab's processor, which already:
          - copies files <= the configured size limit untouched, and
          - compresses larger files into TEMP API PDFS first,
        always leaving the original SCANNED PDF untouched. When OCR finishes,
        the one-shot _aix_on_complete hook auto-pushes results to the tabs,
        which is where originals are renamed.
        """
        if getattr(self, "_aix_running", False) or self._worker_running or self._dispatch_running:
            return
        try:
            from aiextracttab import _aix_start_process, _aix_log, _aix_maybe_prompt_sheet
        except Exception as e:
            logging.error(f"[AUTO-INGEST API] Could not import AI Extract functions: {e}", exc_info=True)
            self._set_status(f"[AUTO-INGEST API] AI Extract unavailable: {e}", ERROR)
            return

        _aix_log(self, "SYSTEM", "Auto-ingest (API) triggered by folder watcher.")
        # In API mode we ONLY extract and ACCUMULATE results in the AI Extract
        # result tree - we do NOT auto-push them to the OCR Renamer / GRN
        # Dispatch tabs. The user reviews / edits / renames in the result tree
        # first, then clicks "Send to Tabs" manually. Results persist across
        # tab switches and further auto-ingests until the user presses "Clear".
        # The one-shot hook only offers to generate the sheet at the 30 mark.
        self._aix_on_complete = lambda: _aix_maybe_prompt_sheet(self)
        _aix_start_process(self, auto=True)

    def _on_watcher_offline_toggle(self):
        enabled = self._watcher_offline_var.get()
        if enabled:
            # Mutually exclusive with API mode.
            self._watcher_api_var.set(False)
        self._persist_ingest_modes()
        self._apply_watcher_state()

    def _on_watcher_api_toggle(self):
        enabled = self._watcher_api_var.get()
        if enabled:
            # Mutually exclusive with offline mode.
            self._watcher_offline_var.set(False)
        self._persist_ingest_modes()
        self._apply_watcher_state()

    def _persist_ingest_modes(self):
        s = self.cfg.setdefault("app_settings", {})
        s["auto_ingest_offline"] = bool(self._watcher_offline_var.get())
        s["auto_ingest_api"] = bool(self._watcher_api_var.get())
        # Keep the legacy key roughly in sync so older code/exports still work.
        s["auto_ingest_watcher"] = s["auto_ingest_offline"] or s["auto_ingest_api"]
        self._save_config()

    def _apply_watcher_state(self):
        """Start or stop the folder observer based on the active mode."""
        if self._active_ingest_mode() != "off":
            if self._watcher_observer and self._watcher_observer.is_alive():
                # Already running — just refresh the badge to the new mode.
                self._watcher_badge_var.set(f"● AUTO-INGEST: {self._active_ingest_mode().upper()}")
                self._set_status(
                    f"Auto-ingest mode set to {self._active_ingest_mode()}.", ACCENT
                )
            else:
                self._start_watcher()
        else:
            self._stop_watcher()

    # Backward-compat alias (kept in case other code references the old name).
    def _on_watcher_toggle(self):
        self._apply_watcher_state()

    # ------------------------------------------------------------------
    # QUEUE POLLING
    # ------------------------------------------------------------------
    def _poll_rename_queue(self):
        try:
            while True:
                item = self._rename_queue.get_nowait()
                if item is None:
                    self._worker_running = False
                    self._set_buttons_state("normal")
                    self._refresh_dashboard_stats()
                    return
                kind, data = item
                if kind == "row":
                    self._add_rename_tree_row(data)
                elif kind == "progress":
                    self._progress.configure(value=data)
                elif kind == "status":
                    self._status_var.set(data)
        except queue.Empty:
            pass
        if self._worker_running:
            self.after(60, self._poll_rename_queue)

    def _poll_dispatch_queue(self):
        try:
            while True:
                item = self._dispatch_queue.get_nowait()
                if item is None:
                    self._dispatch_running = False
                    self._set_buttons_state("normal")
                    return
                kind, data = item
                if kind == "row":
                    self._add_dispatch_tree_row(data)
                elif kind == "progress":
                    self._dispatch_progress.configure(value=data)
                elif kind == "status":
                    self._status_var.set(data)
        except queue.Empty:
            pass
        if self._dispatch_running:
            self.after(60, self._poll_dispatch_queue)

    # ------------------------------------------------------------------
    # BUTTON STATE MANAGEMENT
    # ------------------------------------------------------------------
    def _set_buttons_state(self, state: str):
        """Enable or disable process-trigger buttons."""
        btns = getattr(self, "_process_buttons", [])
        for btn in btns:
            try:
                btn.configure(state=state)
            except Exception:
                pass

    # ------------------------------------------------------------------
    # STYLE
    # ------------------------------------------------------------------
    def _build_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        # Base
        style.configure("TFrame",         background=BG)
        style.configure("Panel.TFrame",   background=PANEL)
        style.configure("Panel2.TFrame",  background=PANEL2)

        # Labels
        style.configure("TLabel",        background=BG,    foreground=TEXT,  font=("Segoe UI", 10))
        style.configure("Muted.TLabel",  background=BG,    foreground=MUTED, font=("Segoe UI", 9))
        style.configure("Title.TLabel",  background=BG,    foreground=TEXT,  font=("Segoe UI", 22, "bold"))
        style.configure("Subtitle.TLabel", background=BG,  foreground=MUTED, font=("Segoe UI", 10))
        style.configure("Panel.TLabel",  background=PANEL, foreground=TEXT,  font=("Segoe UI", 10))
        style.configure("Panel2.TLabel", background=PANEL2,foreground=TEXT,  font=("Segoe UI", 10))
        style.configure("CardTitle.TLabel", background=PANEL, foreground=MUTED,  font=("Segoe UI", 9,  "bold"))
        style.configure("CardValue.TLabel", background=PANEL, foreground=TEXT,   font=("Segoe UI", 30, "bold"))
        style.configure("Author.TLabel",    background=BG,    foreground=ACCENT, font=("Segoe UI", 13, "bold"))
        style.configure("SectionHead.TLabel", background=PANEL, foreground=TEXT, font=("Segoe UI", 12, "bold"))

        # Standard button
        style.configure(
            "TButton", font=("Segoe UI", 9, "bold"), padding=(12, 7),
            background=PANEL2, foreground=TEXT, borderwidth=0, focuscolor="none",
            relief="flat",
        )
        style.map("TButton",
            background=[("active", PANEL3), ("disabled", PANEL)],
            foreground=[("disabled", MUTED)],
        )

        # Accent button
        style.configure(
            "Accent.TButton", font=("Segoe UI", 9, "bold"), padding=(14, 6),
            background=ACCENT, foreground="white", borderwidth=0,
        )
        style.map("Accent.TButton", background=[("active", ACCENT_H), ("disabled", PANEL3)])

        # Success button
        style.configure(
            "Success.TButton", font=("Segoe UI", 9, "bold"), padding=(12, 7),
            background=SUCCESS, foreground="white", borderwidth=0,
        )
        style.map("Success.TButton", background=[("active", "#059669")])

        # Warning button
        style.configure(
            "Warning.TButton", font=("Segoe UI", 9, "bold"), padding=(12, 7),
            background=WARNING, foreground="#1a1a1a", borderwidth=0,
        )
        style.map("Warning.TButton", background=[("active", "#D97706")])

        # Notebook
        style.configure("TNotebook", background=BG, borderwidth=0, tabmargins=[12, 8, 0, 0])
        style.configure(
            "TNotebook.Tab", padding=(22, 10), font=("Segoe UI", 10, "bold"),
            background=PANEL2, foreground=MUTED, borderwidth=0, focuscolor="none",
        )
        style.map("TNotebook.Tab",
            background=[("selected", ACCENT)],
            foreground=[("selected", "white")],
        )

        # Treeview
        style.configure(
            "Treeview", background=PANEL, fieldbackground=PANEL, foreground=TEXT,
            rowheight=32, borderwidth=0, font=("Segoe UI", 9),
        )
        style.map("Treeview",
            background=[("selected", ACCENT)],
            foreground=[("selected", "white")],
        )
        style.configure(
            "Treeview.Heading", background=PANEL2, foreground=MUTED,
            font=("Segoe UI", 9, "bold"), borderwidth=0, padding=(0, 7),
        )
        style.map("Treeview.Heading", background=[("active", PANEL3)])

        # Progressbar
        style.configure(
            "Horizontal.TProgressbar", background=ACCENT,
            troughcolor=PANEL2, borderwidth=0, thickness=6,
        )

        # Scrollbar
        style.configure("TScrollbar", background=PANEL2, troughcolor=PANEL, borderwidth=0, arrowcolor=MUTED)
        style.map("TScrollbar", background=[("active", PANEL3)])

        # Radiobutton / Checkbutton
        style.configure("TRadiobutton", background=PANEL, foreground=TEXT, font=("Segoe UI", 10), focuscolor="none")
        style.configure("TCheckbutton", background=PANEL, foreground=TEXT, font=("Segoe UI", 10), focuscolor="none")
        style.map("TRadiobutton", background=[("active", PANEL2)])
        style.map("TCheckbutton", background=[("active", PANEL2)])

    # ------------------------------------------------------------------
    # LAYOUT
    # ------------------------------------------------------------------
    def _build_layout(self):
        self._build_header()

        # Build status bar FIRST so it anchors to the bottom before the notebook
        self._build_status_bar()

        # Notebook fills all remaining space between header and status bar
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 0))

        self._build_dashboard_tab()
        self._build_renamer_tab()
        self._build_dispatch_tab()
        self._build_settings_tab()
        self._build_about_tab()

        try:
            from scan_tab import add_scan_tab
            add_scan_tab(self)
        except Exception as e:
            logging.error(f"Scan tab failed to load: {e}")
        try:
            from aiextracttab import add_ai_extract_tab
            add_ai_extract_tab(self)
        except Exception as e:
            logging.error(f"AI Extract tab failed to load: {e}", exc_info=True)
            messagebox.showerror("AI Extract Tab Load Error", str(e))

        self._process_buttons = [
            getattr(self, "_btn_start_rename",   None),
            getattr(self, "_btn_start_dispatch", None),
            getattr(self, "_btn_extract_all",    None),
        ]
        self._process_buttons = [b for b in self._process_buttons if b]

    def _build_header(self):
        header = tk.Frame(self, bg=PANEL2, height=64)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        left = tk.Frame(header, bg=PANEL2)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(20, 0))

        lp = resource_path("logo.png")
        if lp.exists():
            try:
                img = Image.open(lp)
                img.thumbnail((44, 44))
                self._header_logo_img = ImageTk.PhotoImage(img)
                tk.Label(left, image=self._header_logo_img, bg=PANEL2).pack(side=tk.LEFT, padx=(0, 14), pady=10)
            except Exception:
                pass

        tb = tk.Frame(left, bg=PANEL2)
        tb.pack(side=tk.LEFT, fill=tk.Y)

        tk.Label(
            tb,
            text="OUTRIGGER MAAFUSHIVARU",
            bg=PANEL2,
            fg=TEXT,
            font=("Segoe UI", 15, "bold"),
        ).pack(anchor="w", pady=(12, 0))

        tk.Label(
            tb,
            text=f"Document Processing Hub  ·  {self.APP_VERSION}",
            bg=PANEL2,
            fg=MUTED,
            font=("Segoe UI", 9),
        ).pack(anchor="w")

        right = tk.Frame(header, bg=PANEL2)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=20)

        self._watcher_badge_lbl = tk.Label(
            right,
            textvariable=self._watcher_badge_var,
            bg=PANEL2,
            fg=WATCHER,
            font=("Segoe UI", 9, "bold"),
        )
        self._watcher_badge_lbl.pack(side=tk.RIGHT, padx=(12, 0), pady=20)

        tk.Label(
            right,
            textvariable=self._engine_badge_var,
            bg=PANEL2,
            fg=ACCENT,
            font=("Segoe UI", 9, "bold"),
        ).pack(side=tk.RIGHT, pady=20)

    def _build_status_bar(self):
        bar = tk.Frame(self, bg=PANEL2, height=44)
        bar.pack(fill=tk.X)
        bar.pack_propagate(False)

        self._status_indicator = tk.Canvas(bar, width=10, height=10, bg=PANEL2, highlightthickness=0)
        self._status_indicator.pack(side=tk.LEFT, padx=(14, 6), pady=13)
        self._status_dot = self._status_indicator.create_oval(1, 1, 9, 9, fill=SUCCESS, outline="")

        tk.Label(bar, textvariable=self._status_var, bg=PANEL2, fg=TEXT,
                 font=("Segoe UI", 9), anchor="w").pack(side=tk.LEFT, pady=8)

        self._btn_extract_all = ttk.Button(
            bar, text="⚡  Extract & Process",
            style="Accent.TButton", command=self._start_extract_and_process,
        )
        self._btn_extract_all.pack(side=tk.RIGHT, padx=16, pady=4)

    def _refresh_engine_badge(self):
        eng = self.cfg.get("app_settings", {}).get("ocr_engine", "tesseract")
        self._engine_badge_var.set(f"OCR: {ENGINE_LABELS.get(eng, eng)}")

    # ------------------------------------------------------------------
    # TAB HELPERS
    # ------------------------------------------------------------------
    def _make_tab(self, title):
        """Non-scrollable tab (for Renamer/Dispatch which have internal trees)."""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text=f"  {title}  ")
        return frame

    def _make_scrollable_tab(self, title):
        outer = ttk.Frame(self.notebook)
        self.notebook.add(outer, text=f"  {title}  ")
        canvas = tk.Canvas(outer, bg=BG, highlightthickness=0)
        sb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        wid = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_configure(_):
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner.bind("<Configure>", _on_configure)

        def _on_canvas_configure(e):
            canvas.itemconfigure(wid, width=e.width)
        canvas.bind("<Configure>", _on_canvas_configure)
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        def _bind_mw(_):
            canvas.bind_all("<MouseWheel>", lambda ev: canvas.yview_scroll(int(-1 * (ev.delta / 120)), "units"))
        def _unbind_mw(_):
            canvas.unbind_all("<MouseWheel>")
        canvas.bind("<Enter>", _bind_mw)
        canvas.bind("<Leave>", _unbind_mw)
        return inner

    def _section(self, parent, title, subtitle=""):
        box = tk.Frame(parent, bg=PANEL, bd=0)
        box.pack(fill=tk.X, padx=16, pady=(0, 14))
        hdr = tk.Frame(box, bg=PANEL)
        hdr.pack(fill=tk.X, padx=20, pady=(16, 0))
        tk.Label(hdr, text=title, bg=PANEL, fg=TEXT, font=("Segoe UI", 12, "bold")).pack(side=tk.LEFT)
        if subtitle:
            tk.Label(hdr, text=f"  —  {subtitle}", bg=PANEL, fg=MUTED, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        # Thin separator line
        sep = tk.Frame(box, bg=PANEL3, height=1)
        sep.pack(fill=tk.X, padx=20, pady=(8, 12))
        return box

    def _make_stat_card(self, parent, title, value, badge, color):
        card = tk.Frame(parent, bg=PANEL, bd=0)
        card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=6, pady=4)
        # Top accent bar
        accent_bar = tk.Frame(card, bg=color, height=3)
        accent_bar.pack(fill=tk.X)

        inner = tk.Frame(card, bg=PANEL)
        inner.pack(fill=tk.BOTH, expand=True, padx=18, pady=14)

        top = tk.Frame(inner, bg=PANEL)
        top.pack(fill=tk.X)
        tk.Label(top, text=badge, bg=PANEL, fg=color, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)

        tk.Label(inner, text=title, bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(6, 0))
        var = tk.StringVar(value=str(value))
        tk.Label(inner, textvariable=var, bg=PANEL, fg=TEXT, font=("Segoe UI", 28, "bold")).pack(anchor="w")
        return var

    # ------------------------------------------------------------------
    # SEARCH BAR HELPER (fixed mousewheel)
    # ------------------------------------------------------------------
    def _make_search_bar(self, parent, tree_ref_getter, all_rows_ref_getter):
        bar = tk.Frame(parent, bg=PANEL)
        bar.pack(fill=tk.X, padx=20, pady=(4, 8))

        tk.Label(bar, text="Search:", bg=PANEL, fg=MUTED, font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 6))
        var = tk.StringVar()
        ent = tk.Entry(
            bar, textvariable=var, bg=PANEL2, fg=TEXT, insertbackground=TEXT,
            relief="flat", font=("Segoe UI", 10), bd=0,
            highlightthickness=1, highlightbackground=PANEL3, highlightcolor=ACCENT,
        )
        ent.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=5)

        def _clear():
            var.set("")
        ttk.Button(bar, text="✕", command=_clear, width=3).pack(side=tk.LEFT, padx=(6, 0))

        def _on_search(*_):
            tree = tree_ref_getter()
            q = var.get().strip().lower()
            for iid in tree.get_children():
                tree.detach(iid)
            for iid in all_rows_ref_getter():
                if not q:
                    tree.reattach(iid, "", "end")
                else:
                    vals = tree.item(iid, "values")
                    if any(q in str(v).lower() for v in vals):
                        tree.reattach(iid, "", "end")

        var.trace_add("write", _on_search)
        return var

    # ------------------------------------------------------------------
    # BIND MOUSEWHEEL TO TREE (FIX)
    # The key fix: we bind directly to the tree widget so scrolling works
    # when the cursor is inside it, regardless of outer canvas bindings.
    # ------------------------------------------------------------------
    def _bind_tree_mousewheel(self, tree: ttk.Treeview):
        def _on_mw(event):
            tree.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"

        tree.bind("<MouseWheel>", _on_mw)
        tree.bind("<Button-4>",  lambda e: (tree.yview_scroll(-1, "units"), "break"))  # Linux
        tree.bind("<Button-5>",  lambda e: (tree.yview_scroll( 1, "units"), "break"))  # Linux

    # ------------------------------------------------------------------
    # DASHBOARD TAB
    # ------------------------------------------------------------------
    def _build_dashboard_tab(self):
        frame = self._make_scrollable_tab("Dashboard")
        frame.pack_configure(padx=0)

        # Stats row
        sr = tk.Frame(frame, bg=BG)
        sr.pack(fill=tk.X, padx=16, pady=(20, 0))
        self._stat_waiting_var   = self._make_stat_card(sr, "Waiting",   "-", "SCAN",  WARNING)
        self._stat_processed_var = self._make_stat_card(sr, "Processed", "-", "DONE",  SUCCESS)
        self._stat_archived_var  = self._make_stat_card(sr, "Archived",  "-", "ARCH",  ACCENT2)
        self._stat_failed_var    = self._make_stat_card(sr, "Failed",    "-", "FAIL",  ERROR)

        # Quick actions
        qa = self._section(frame, "Quick Actions")
        row = tk.Frame(qa, bg=PANEL)
        row.pack(fill=tk.X, padx=20, pady=(0, 16))
        ttk.Button(row, text="↺  Refresh Stats",   command=self._refresh_dashboard_stats).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(row, text="📂  Open SCANNED",   command=lambda: self._open_folder(self.dirs["scanned"])).pack(side=tk.LEFT, padx=4)
        ttk.Button(row, text="📂  Open PROCESSED", command=lambda: self._open_folder(self.dirs["processed"])).pack(side=tk.LEFT, padx=4)
        ttk.Button(row, text="📂  Open FAILED",    command=lambda: self._open_folder(self.dirs["failed"])).pack(side=tk.LEFT, padx=4)
        ttk.Button(row, text="📋  Open Logs",      command=lambda: self._open_folder(self.dirs["logs"])).pack(side=tk.LEFT, padx=4)

        # Directories
        info = self._section(frame, "System Directories")
        for k in ("base", "scanned", "processed", "archive", "failed", "logs"):
            r = tk.Frame(info, bg=PANEL)
            r.pack(fill=tk.X, padx=20, pady=3)
            tk.Label(r, text=f"{k.upper()}:", bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold"), width=12, anchor="w").pack(side=tk.LEFT)
            tk.Label(r, text=self.dirs.get(k, ""), bg=PANEL, fg=ACCENT, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        tk.Frame(info, bg=PANEL).pack(pady=8)

    # ------------------------------------------------------------------
    # OCR RENAMER TAB
    # ------------------------------------------------------------------
    def _build_renamer_tab(self):
        frame = self._make_tab("OCR Renamer")
        frame.configure(style="TFrame")

        # Top control bar
        ctrl_bar = tk.Frame(frame, bg=PANEL2, height=52)
        ctrl_bar.pack(fill=tk.X)
        ctrl_bar.pack_propagate(False)

        self._btn_start_rename = ttk.Button(
            ctrl_bar, text="▶  Start OCR Rename", style="Accent.TButton",
            command=self._start_rename_worker,
        )
        self._btn_start_rename.pack(side=tk.LEFT, padx=(16, 6), pady=8)

        ttk.Button(ctrl_bar, text="⏹  Cancel",     command=self._cancel_worker).pack(side=tk.LEFT, padx=4, pady=8)
        ttk.Button(ctrl_bar, text="♻  Retry Failed", style="Warning.TButton",
                   command=self._retry_failed_files).pack(side=tk.LEFT, padx=4, pady=8)
        ttk.Button(ctrl_bar, text="🗑  Clear",       command=self._clear_rename_results).pack(side=tk.LEFT, padx=4, pady=8)
        ttk.Button(ctrl_bar, text="🔬  Debug PDF",   command=self._test_single_pdf_debug).pack(side=tk.LEFT, padx=4, pady=8)

        # Dry run toggle on right
        dr_frame = tk.Frame(ctrl_bar, bg=PANEL2)
        dr_frame.pack(side=tk.RIGHT, padx=16, pady=8)
        ttk.Checkbutton(
            dr_frame, text="Dry Run Preview", variable=self._dry_run_var,
            command=self._on_dry_run_toggle, style="TCheckbutton",
        ).pack(side=tk.RIGHT)

        # Progress bar
        prog_bar = tk.Frame(frame, bg=PANEL, height=6)
        prog_bar.pack(fill=tk.X)
        self._progress = ttk.Progressbar(prog_bar, orient="horizontal", mode="determinate")
        self._progress.pack(fill=tk.X)

        # Results label + search
        hdr_bar = tk.Frame(frame, bg=PANEL)
        hdr_bar.pack(fill=tk.X)
        tk.Label(hdr_bar, text="Rename Results", bg=PANEL, fg=TEXT,
                 font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT, padx=20, pady=(12, 4))
        tk.Label(hdr_bar, text="Double-click Supplier / GRN / Invoice to edit  ·  Double-click File to preview PDF",
                 bg=PANEL, fg=MUTED, font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=8, pady=(12, 4))

        search_bar = tk.Frame(frame, bg=PANEL)
        search_bar.pack(fill=tk.X)
        self._make_search_bar(search_bar, lambda: self._rename_tree, lambda: self._rename_all_rows)

        # Tree frame — takes all remaining space
        tf = tk.Frame(frame, bg=PANEL)
        tf.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 16))

        rcols = ("File", "New Name", "Supplier", "Confidence", "GRN", "Invoice #", "Status", "Warning")
        cw = {
            "File": 165, "New Name": 230, "Supplier": 185, "Confidence": 85,
            "GRN": 195, "Invoice #": 125, "Status": 75, "Warning": 210,
        }
        self._rename_tree = ttk.Treeview(tf, columns=rcols, show="headings", height=18)
        for c in rcols:
            self._rename_tree.heading(c, text=c, anchor="center")
            self._rename_tree.column(c, width=cw.get(c, 100), anchor="center", stretch=False)

        yr = ttk.Scrollbar(tf, orient="vertical",   command=self._rename_tree.yview)
        xr = ttk.Scrollbar(tf, orient="horizontal", command=self._rename_tree.xview)
        self._rename_tree.configure(yscrollcommand=yr.set, xscrollcommand=xr.set)
        self._rename_tree.grid(row=0, column=0, sticky="nsew")
        yr.grid(row=0, column=1, sticky="ns")
        xr.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1)
        tf.columnconfigure(0, weight=1)

        self._bind_tree_mousewheel(self._rename_tree)

        def _rename_pre_edit(col_index, cur):
            if col_index == 4:  # GRN (shifted by Confidence column)
                return "RC-MAM-0000" if cur in ("NO-GRN", "", "NO GRN") else cur
            if col_index == 5:  # Invoice #
                c = str(cur).strip()
                return c[3:].strip() if c.upper().startswith("IN ") else c
            return cur

        self._make_tree_editable(
            self._rename_tree,
            on_edit_callback=self._on_rename_tree_edit,
            editable_cols={2, 4, 5},
            pre_edit_fn=_rename_pre_edit,
        )
        # Register PDF preview for column 0
        self._rename_tree._edit_on_preview_cb = self._on_rename_preview_by_rowid

    # ------------------------------------------------------------------
    # GRN DISPATCH TAB
    # ------------------------------------------------------------------
    def _build_dispatch_tab(self):
        frame = self._make_tab("GRN Dispatch")
        frame.configure(style="TFrame")

        # Top control bar
        ctrl_bar = tk.Frame(frame, bg=PANEL2, height=52)
        ctrl_bar.pack(fill=tk.X)
        ctrl_bar.pack_propagate(False)

        self._btn_start_dispatch = ttk.Button(
            ctrl_bar, text="▶  Run GRN Extraction", style="Accent.TButton",
            command=self._start_dispatch_worker,
        )
        self._btn_start_dispatch.pack(side=tk.LEFT, padx=(16, 6), pady=8)

        ttk.Button(ctrl_bar, text="⏹  Cancel", command=self._cancel_worker).pack(side=tk.LEFT, padx=4, pady=8)
        ttk.Button(ctrl_bar, text="✅  Validate GRN", command=self._validate_grn, style="Success.TButton").pack(side=tk.LEFT, padx=4, pady=8)
        ttk.Button(ctrl_bar, text="📊  Export Excel", command=self._export_dispatch_to_excel).pack(side=tk.LEFT, padx=4, pady=8)
        ttk.Button(ctrl_bar, text="🗑  Clear", command=self._clear_dispatch_results).pack(side=tk.LEFT, padx=4, pady=8)

        # Source folder on right
        sf_frame = tk.Frame(ctrl_bar, bg=PANEL2)
        sf_frame.pack(side=tk.RIGHT, padx=16, pady=8)
        tk.Label(sf_frame, text="Source:", bg=PANEL2, fg=MUTED, font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 6))
        self._dispatch_folder_var = tk.StringVar(value=self.dirs.get("processed", ""))
        ent = tk.Entry(
            sf_frame,
            textvariable=self._dispatch_folder_var,
            bg=PANEL,
            fg=TEXT,
            insertbackground=TEXT,
            relief="flat",
            font=("Segoe UI", 9),
            bd=0,
            highlightthickness=1,
            highlightbackground=PANEL3,
            highlightcolor=ACCENT,
            width=36,
        )
        ent.pack(side=tk.LEFT, ipady=4)
        ttk.Button(sf_frame, text="Browse", command=self._browse_dispatch_folder).pack(side=tk.LEFT, padx=(6, 0))

        # Progress
        prog_bar = tk.Frame(frame, bg=PANEL, height=6)
        prog_bar.pack(fill=tk.X)
        self._dispatch_progress = ttk.Progressbar(prog_bar, orient="horizontal", mode="determinate")
        self._dispatch_progress.pack(fill=tk.X)

        # Header
        hdr_bar = tk.Frame(frame, bg=PANEL)
        hdr_bar.pack(fill=tk.X)
        tk.Label(
            hdr_bar,
            text="GRN Dispatch Results",
            bg=PANEL,
            fg=TEXT,
            font=("Segoe UI", 11, "bold"),
        ).pack(side=tk.LEFT, padx=20, pady=(12, 4))
        tk.Label(
            hdr_bar,
            text="Double-click any cell except File to edit",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 8),
        ).pack(side=tk.LEFT, padx=8, pady=(12, 4))

        search_bar = tk.Frame(frame, bg=PANEL)
        search_bar.pack(fill=tk.X)
        self._make_search_bar(search_bar, lambda: self._dispatch_tree, lambda: self._dispatch_all_rows)

        # Tree
        tf = tk.Frame(frame, bg=PANEL)
        tf.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 16))

        dcols = ("File", "Date", "Supplier", "Confidence", "PO #", "Invoice #", "USD", "MVR", "EUR", "GBP", "SGD", "GRN")
        dcw = {
            "File": 190, "Date": 90, "Supplier": 200, "Confidence": 80,
            "PO #": 115, "Invoice #": 130, "USD": 85, "MVR": 85,
            "EUR": 85, "GBP": 85, "SGD": 85, "GRN": 220,
        }

        self._dispatch_tree = ttk.Treeview(tf, columns=dcols, show="headings", height=20)
        for c in dcols:
            self._dispatch_tree.heading(c, text=c, anchor="center")
            self._dispatch_tree.column(c, width=dcw.get(c, 100), anchor="center", stretch=False)

        ys = ttk.Scrollbar(tf, orient="vertical", command=self._dispatch_tree.yview)
        xs = ttk.Scrollbar(tf, orient="horizontal", command=self._dispatch_tree.xview)
        self._dispatch_tree.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)

        self._dispatch_tree.grid(row=0, column=0, sticky="nsew")
        ys.grid(row=0, column=1, sticky="ns")
        xs.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1)
        tf.columnconfigure(0, weight=1)

        self._bind_tree_mousewheel(self._dispatch_tree)

        self._make_tree_editable(
            self._dispatch_tree,
            on_edit_callback=self._on_dispatch_tree_edit,
            editable_cols=set(range(1, 12)),
        )

        self._dispatch_tree._edit_on_preview_cb = self._on_dispatch_preview_by_rowid

    def _on_dispatch_preview_by_rowid(self, row_id: str):
        result   = self._dispatch_row_map.get(row_id, {})
        pdf_path = result.get("raw_path", "")
        folder   = self._dispatch_folder_var.get().strip() or self.dirs.get("processed", "")
        if not pdf_path or not os.path.exists(pdf_path):
            fname    = result.get("file", "")
            pdf_path = os.path.join(folder, fname)
        self._show_pdf_preview(pdf_path)

    def _browse_dispatch_folder(self):
        f = filedialog.askdirectory(initialdir=self._dispatch_folder_var.get() or self.dirs["base"])
        if f:
            self._dispatch_folder_var.set(f)

    # ------------------------------------------------------------------
    # SETTINGS TAB
    # ------------------------------------------------------------------
    def _build_settings_tab(self):
        frame = self._make_scrollable_tab("Settings")

        # Auto-ingest watcher
        wb = self._section(frame, "Auto-Ingest (Watched Folder)",
                           subtitle="Monitors SCANNED folder for new PDFs")
        wf = tk.Frame(wb, bg=PANEL)
        wf.pack(fill=tk.X, padx=20, pady=(0, 4))

        if not WATCHDOG_AVAILABLE:
            tk.Label(
                wf,
                text="⚠  watchdog library not installed.  Run:  pip install watchdog",
                bg=PANEL, fg=WARNING, font=("Segoe UI", 9),
            ).pack(anchor="w", pady=(0, 8))

        ttk.Checkbutton(
            wf,
            text="Enable Auto-Ingest (Offline) — new PDFs run local OCR Rename + GRN Dispatch",
            variable=self._watcher_offline_var, command=self._on_watcher_offline_toggle,
        ).pack(anchor="w", pady=4)
        ttk.Checkbutton(
            wf,
            text="Enable Auto-Ingest (API) — new PDFs run AI Extract (OCR.space): "
                 "auto size-check, compress if over the limit, then send to tabs",
            variable=self._watcher_api_var, command=self._on_watcher_api_toggle,
        ).pack(anchor="w", pady=4)
        tk.Label(
            wf,
            text=(
                "Choose ONE mode — the two are mutually exclusive so a file is never "
                "processed twice. New files are debounced by 3 seconds to wait for the full "
                "copy to finish.\n"
                "• Offline: fully local OCR — no internet, no upload limit.\n"
                "• API: files at or below the OCR.space upload limit are sent as-is; "
                "larger files are compressed into 'TEMP API PDFS' first. The original PDF in "
                "SCANNED is never modified — it is only renamed when results are sent to the tabs."
            ),
            bg=PANEL, fg=MUTED, font=("Segoe UI", 9), wraplength=850, justify="left",
        ).pack(anchor="w", pady=(0, 12))

        # Desktop notifications
        nb = self._section(frame, "Desktop Notifications")
        nf = tk.Frame(nb, bg=PANEL)
        nf.pack(fill=tk.X, padx=20, pady=(0, 4))

        if not PLYER_AVAILABLE:
            tk.Label(
                nf,
                text="⚠  plyer library not installed.  Run:  pip install plyer",
                bg=PANEL, fg=WARNING, font=("Segoe UI", 9),
            ).pack(anchor="w", pady=(0, 8))

        ttk.Checkbutton(
            nf,
            text="Show desktop notifications when a batch finishes or a file fails",
            variable=self._notify_var, command=self._on_notify_toggle,
        ).pack(anchor="w", pady=4)
        tk.Label(
            nf,
            text='Summary format: "23 done, 1 failed — LYCORN invoice unmatched."',
            bg=PANEL, fg=MUTED, font=("Segoe UI", 9),
        ).pack(anchor="w", pady=(0, 12))

        # Confidence threshold
        cfb = self._section(frame, "Confidence Scoring", subtitle="Flags uncertain supplier matches")
        cf = tk.Frame(cfb, bg=PANEL)
        cf.pack(fill=tk.X, padx=20, pady=(0, 12))
        tk.Label(cf, text="Warn threshold (%):", bg=PANEL, fg=TEXT, font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Spinbox(cf, from_=0, to=100, textvariable=self._conf_threshold_var, width=6,
                    font=("Segoe UI", 10)).pack(side=tk.LEFT)
        tk.Label(cf, text="  — rows below this are highlighted in orange",
                 bg=PANEL, fg=MUTED, font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=8)
                 
        # Page scan region
        psb = self._section(frame, "Page Scan Region",
                            subtitle="Limit OCR to a % of the page from the top")
        pf = tk.Frame(psb, bg=PANEL)
        pf.pack(fill=tk.X, padx=20, pady=(0, 12))

        self._page_scan_enabled_var = tk.BooleanVar(
            value=self.cfg.get("app_settings", {}).get("page_scan_region_enabled", False)
        )
        self._page_scan_percent_var = tk.IntVar(
            value=self.cfg.get("app_settings", {}).get("page_scan_region_percent", 100)
        )

        ttk.Checkbutton(
            pf,
            text="Enable page scan region (scan only top N% of each page)",
            variable=self._page_scan_enabled_var,
            style="TCheckbutton",
        ).pack(anchor="w", pady=(0, 8))

        pct_row = tk.Frame(pf, bg=PANEL)
        pct_row.pack(anchor="w")
        tk.Label(pct_row, text="Scan top:", bg=PANEL, fg=TEXT,
                 font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Spinbox(
            pct_row, from_=5, to=100,
            textvariable=self._page_scan_percent_var,
            width=6, font=("Segoe UI", 10),
        ).pack(side=tk.LEFT)
        tk.Label(pct_row, text="% of each page  (5–100)",
                 bg=PANEL, fg=MUTED, font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=8)

        tk.Label(
            pf,
            text="Useful when supplier/invoice info always appears in the page header. "
                 "Set to 100% to scan the full page (same as disabled).",
            bg=PANEL, fg=MUTED, font=("Segoe UI", 9), wraplength=850,
        ).pack(anchor="w", pady=(8, 0))

        # ---- OCR Mode (Zone vs Full) ----
        ocr_mode_box = self._section(frame, "OCR Extraction Mode",
                                     subtitle="Zone OCR is faster; Full OCR is more thorough")
        self._ocr_mode_var = tk.StringVar(
            value=self.cfg.get("app_settings", {}).get("ocr_mode", "full")
        )
        ocr_mode_inner = tk.Frame(ocr_mode_box, bg=PANEL)
        ocr_mode_inner.pack(fill=tk.X, padx=20, pady=(0, 4))

        for val, lbl, desc in [
            (
                "full",
                "Full OCR  (Default, Recommended)",
                "Scans the entire page. Slower but catches all data including rotated or unusual layouts.",
            ),
            (
                "zone",
                "Zone OCR  (Fast)",
                "Scans only the top ~38% header region. 3–5× faster. Best for standard Birchstreet "
                "receiving reports where GRN, supplier and invoice data are always in the header.",
            ),
        ]:
            row = tk.Frame(ocr_mode_inner, bg=PANEL)
            row.pack(fill=tk.X, pady=4)
            ttk.Radiobutton(
                row, text=lbl, value=val,
                variable=self._ocr_mode_var,
                command=self._on_ocr_mode_changed,
                style="TRadiobutton",
            ).pack(anchor="w")
            tk.Label(
                row, text=f"    {desc}",
                bg=PANEL, fg=MUTED, font=("Segoe UI", 8), wraplength=820, justify="left",
            ).pack(anchor="w")

        # Info label showing current mode
        self._ocr_mode_info_var = tk.StringVar(
            value=self._ocr_mode_label(self.cfg.get("app_settings", {}).get("ocr_mode", "full"))
        )
        tk.Label(
            ocr_mode_inner,
            textvariable=self._ocr_mode_info_var,
            bg=PANEL, fg=ACCENT, font=("Segoe UI", 9, "bold"),
        ).pack(anchor="w", pady=(4, 12))
        
        # Processing mode
        pm_box = self._section(frame, "Processing Mode")
        self._processing_mode_var = tk.StringVar(
            value=self.cfg.get("app_settings", {}).get("processing_mode", "legacy")
        )
        pm_inner = tk.Frame(pm_box, bg=PANEL)
        pm_inner.pack(fill=tk.X, padx=20, pady=(0, 12))
        for val, lbl, desc in [
            ("legacy", "Legacy  (Default, Recommended)", "Original algorithm — 650px crop, 3× scale, Tesseract PSM 6."),
            ("mixed",  "Mixed",                          "Custom supplier/GRN + Legacy invoice number extraction."),
            ("custom", "Custom",                         "Enhanced multi-method extraction via fields and headers."),
        ]:
            row = tk.Frame(pm_inner, bg=PANEL)
            row.pack(fill=tk.X, pady=3)
            ttk.Radiobutton(row, text=lbl, value=val, variable=self._processing_mode_var,
                            command=self._on_processing_mode_changed, style="TRadiobutton").pack(anchor="w")
            tk.Label(row, text=f"    {desc}", bg=PANEL, fg=MUTED, font=("Segoe UI", 8)).pack(anchor="w")

        # Supplier extraction method
        mb = self._section(frame, "Supplier Extraction Method", subtitle="Custom / Mixed mode only")
        self._supplier_method_var = tk.StringVar(
            value=self.cfg.get("app_settings", {}).get("supplier_extraction_method", "both")
        )
        mb_inner = tk.Frame(mb, bg=PANEL)
        mb_inner.pack(fill=tk.X, padx=20, pady=(0, 12))
        for val, lbl in [
            ("both",                  'Both (Recommended): Check "Supplier:" / "Vendor:" field first, then header'),
            ("receiving_report_field",'Field Only: Extract from "Supplier:" or "Vendor:" in Receiving Report'),
            ("header_company_name",   "Header Only: Extract from invoice header area"),
        ]:
            ttk.Radiobutton(mb_inner, text=lbl, value=val, variable=self._supplier_method_var,
                            command=self._on_supplier_method_changed, style="TRadiobutton").pack(anchor="w", pady=4)

        # OCR engine
        eb = self._section(frame, "OCR Engine")
        self._engine_var = tk.StringVar(value=self.cfg.get("app_settings", {}).get("ocr_engine", "tesseract"))
        eb_inner = tk.Frame(eb, bg=PANEL)
        eb_inner.pack(fill=tk.X, padx=20, pady=(0, 12))
        for key, label in ENGINE_LABELS.items():
            row = tk.Frame(eb_inner, bg=PANEL)
            row.pack(fill=tk.X, pady=5)
            ttk.Radiobutton(row, text=label, value=key, variable=self._engine_var,
                            command=self._on_engine_changed, style="TRadiobutton").pack(side=tk.LEFT)
            st = "Installed" if self._check_engine(key) or key == "tesseract" else "Not installed"
            tk.Label(row, text=st, bg=PANEL, fg=SUCCESS if st == "Installed" else ERROR,
                     font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=16)
            if key != "tesseract":
                ttk.Button(row, text="How to install", command=lambda k=key: self._install_engine(k)).pack(side=tk.RIGHT)
        # ── Supplier Match Strategy ──────────────────────────────────────
        smb = self._section(
            frame,
            "Supplier Match Strategy",
            subtitle="Algorithm used to identify the supplier name from extracted text",
        )
        self._supplier_strategy_var = tk.StringVar(
            value=self.cfg.get("app_settings", {}).get("supplier_match_strategy", "combined")
        )
        smb_inner = tk.Frame(smb, bg=PANEL)
        smb_inner.pack(fill=tk.X, padx=20, pady=(0, 12))

        if SUPPLIER_MATCHER_AVAILABLE:
            for val, lbl in STRATEGY_LABELS.items():
                ttk.Radiobutton(
                    smb_inner, text=lbl, value=val,
                    variable=self._supplier_strategy_var,
                    command=self._on_supplier_strategy_changed,
                    style="TRadiobutton",
                ).pack(anchor="w", pady=3)
        else:
            tk.Label(
                smb_inner,
                text="⚠  supplier_matcher.py not found — place it in the same folder as maafushivaru_hub.py",
                bg=PANEL, fg=WARNING, font=("Segoe UI", 9),
            ).pack(anchor="w", pady=(0, 8))

        # ── AI Supplier Match ─────────────────────────────────────────────
        aib = self._section(
            frame,
            "AI Supplier Matching",
            subtitle="Use a cloud AI (Claude, GPT, Gemini) to identify suppliers",
        )
        ai_inner = tk.Frame(aib, bg=PANEL)
        ai_inner.pack(fill=tk.X, padx=20, pady=(0, 12))

        if not AI_MATCHER_AVAILABLE:
            tk.Label(
                ai_inner,
                text="⚠  ai_supplier_matcher.py not found — place it in the same folder as maafushivaru_hub.py",
                bg=PANEL,
                fg=WARNING,
                font=("Segoe UI", 9),
            ).pack(anchor="w", pady=(0, 8))
        else:
            tk.Label(
                ai_inner,
                text="AI supplier module loaded successfully. OCR.space is available as the default API OCR backend.",
                bg=PANEL,
                fg=SUCCESS,
                font=("Segoe UI", 9),
            ).pack(anchor="w", pady=(0, 8))

        ai_cfg = self.cfg.get("ai_settings", {})

        self._ai_enabled_var = tk.BooleanVar(value=bool(ai_cfg.get("enabled", False)))
        ttk.Checkbutton(
            ai_inner,
            text="Enable AI supplier matching (falls back to rule-based if AI fails)",
            variable=self._ai_enabled_var,
            command=self._on_ai_enabled_toggle,
            style="TCheckbutton",
        ).pack(anchor="w", pady=(0, 8))

        prow = tk.Frame(ai_inner, bg=PANEL)
        prow.pack(fill=tk.X, pady=3)
        tk.Label(
            prow,
            text="Provider:",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 9, "bold"),
            width=16,
            anchor="w",
        ).pack(side=tk.LEFT)

        self._ai_provider_var = tk.StringVar(value=ai_cfg.get("provider", "openai"))
        provider_names = {k: v["label"] for k, v in PROVIDERS.items()} if PROVIDERS else {
            "openai": "OpenAI GPT",
            "custom": "Custom / Local",
        }

        for val, lbl in provider_names.items():
            ttk.Radiobutton(
                prow,
                text=lbl,
                value=val,
                variable=self._ai_provider_var,
                command=self._on_ai_provider_changed,
                style="TRadiobutton",
            ).pack(side=tk.LEFT, padx=(0, 14))

        mrow = tk.Frame(ai_inner, bg=PANEL)
        mrow.pack(fill=tk.X, pady=3)
        tk.Label(
            mrow,
            text="Model:",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 9, "bold"),
            width=16,
            anchor="w",
        ).pack(side=tk.LEFT)

        self._ai_model_var = tk.StringVar(value=ai_cfg.get("model", "gpt-4o-mini"))
        self._ai_model_entry = tk.Entry(
            mrow,
            textvariable=self._ai_model_var,
            bg=PANEL2,
            fg=TEXT,
            insertbackground=TEXT,
            relief="flat",
            font=("Segoe UI", 10),
            bd=0,
            highlightthickness=1,
            highlightbackground=PANEL3,
            highlightcolor=ACCENT,
            width=36,
        )
        self._ai_model_entry.pack(side=tk.LEFT, ipady=4)

        krow = tk.Frame(ai_inner, bg=PANEL)
        krow.pack(fill=tk.X, pady=3)
        tk.Label(
            krow,
            text="API Key:",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 9, "bold"),
            width=16,
            anchor="w",
        ).pack(side=tk.LEFT)

        self._ai_key_var = tk.StringVar(value=ai_cfg.get("api_key", ""))
        tk.Entry(
            krow,
            textvariable=self._ai_key_var,
            bg=PANEL2,
            fg=TEXT,
            insertbackground=TEXT,
            relief="flat",
            font=("Segoe UI", 10),
            bd=0,
            highlightthickness=1,
            highlightbackground=PANEL3,
            highlightcolor=ACCENT,
            show="*",
            width=52,
        ).pack(side=tk.LEFT, ipady=4)

        urow = tk.Frame(ai_inner, bg=PANEL)
        urow.pack(fill=tk.X, pady=3)
        tk.Label(
            urow,
            text="Custom URL:",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 9, "bold"),
            width=16,
            anchor="w",
        ).pack(side=tk.LEFT)

        self._ai_url_var = tk.StringVar(value=ai_cfg.get("custom_base_url", ""))
        tk.Entry(
            urow,
            textvariable=self._ai_url_var,
            bg=PANEL2,
            fg=TEXT,
            insertbackground=TEXT,
            relief="flat",
            font=("Segoe UI", 10),
            bd=0,
            highlightthickness=1,
            highlightbackground=PANEL3,
            highlightcolor=ACCENT,
            width=52,
        ).pack(side=tk.LEFT, ipady=4)

        tk.Label(
            urow,
            text="  (for Custom/Local only)",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 8),
        ).pack(side=tk.LEFT, padx=6)

        trow = tk.Frame(ai_inner, bg=PANEL)
        trow.pack(fill=tk.X, pady=3)
        tk.Label(
            trow,
            text="Timeout (sec):",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 9, "bold"),
            width=16,
            anchor="w",
        ).pack(side=tk.LEFT)

        self._ai_timeout_var = tk.IntVar(value=int(ai_cfg.get("timeout_seconds", 20)))
        ttk.Spinbox(
            trow,
            from_=5,
            to=120,
            textvariable=self._ai_timeout_var,
            width=6,
            font=("Segoe UI", 10),
        ).pack(side=tk.LEFT)

        ocr_space_cfg = self.cfg.get("ocr_space", {})

        okrow = tk.Frame(ai_inner, bg=PANEL)
        okrow.pack(fill=tk.X, pady=3)
        tk.Label(
            okrow,
            text="OCR.space Key:",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 9, "bold"),
            width=16,
            anchor="w",
        ).pack(side=tk.LEFT)

        self._ocr_space_key_var = tk.StringVar(value=ocr_space_cfg.get("api_key", "K88109865088957"))
        tk.Entry(
            okrow,
            textvariable=self._ocr_space_key_var,
            bg=PANEL2,
            fg=TEXT,
            insertbackground=TEXT,
            relief="flat",
            font=("Segoe UI", 10),
            bd=0,
            highlightthickness=1,
            highlightbackground=PANEL3,
            highlightcolor=ACCENT,
            show="*",
            width=52,
        ).pack(side=tk.LEFT, ipady=4)

        oerow = tk.Frame(ai_inner, bg=PANEL)
        oerow.pack(fill=tk.X, pady=3)
        tk.Label(
            oerow,
            text="OCR.space Engine:",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 9, "bold"),
            width=16,
            anchor="w",
        ).pack(side=tk.LEFT)

        self._ocr_space_engine_var = tk.IntVar(value=int(ocr_space_cfg.get("OCREngine", 2)))
        for eng_num in (1, 2, 3):
            ttk.Radiobutton(
                oerow,
                text=str(eng_num),
                value=eng_num,
                variable=self._ocr_space_engine_var,
                style="TRadiobutton",
            ).pack(side=tk.LEFT, padx=(0, 14))

        srow = tk.Frame(ai_inner, bg=PANEL)
        srow.pack(fill=tk.X, pady=3)
        tk.Label(
            srow,
            text="Max Upload:",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 9, "bold"),
            width=16,
            anchor="w",
        ).pack(side=tk.LEFT)

        self._ocr_space_max_mb_var = tk.StringVar(value=str(ocr_space_cfg.get("max_upload_mb", 1.0)))
        tk.Entry(
            srow,
            textvariable=self._ocr_space_max_mb_var,
            bg=PANEL2,
            fg=TEXT,
            insertbackground=TEXT,
            relief="flat",
            font=("Segoe UI", 10),
            bd=0,
            highlightthickness=1,
            highlightbackground=PANEL3,
            highlightcolor=ACCENT,
            width=8,
        ).pack(side=tk.LEFT, ipady=4)

        tk.Label(
            srow,
            text="MB  (OCR upload will be compressed below this limit; recommended 1.0)",
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 8),
        ).pack(side=tk.LEFT, padx=8)

        strow = tk.Frame(ai_inner, bg=PANEL)
        strow.pack(fill=tk.X, pady=(8, 0))

        self._ai_status_lbl = tk.Label(
            strow,
            textvariable=self._ai_status_var,
            bg=PANEL,
            fg=MUTED,
            font=("Segoe UI", 9),
        )
        self._ai_status_lbl.pack(side=tk.LEFT, padx=(0, 12))

        ttk.Button(
            strow,
            text="🔌  Test Connection",
            command=self._test_ai_connection,
        ).pack(side=tk.LEFT, padx=(0, 6))

        ttk.Button(
            strow,
            text="📜  View AI Log",
            command=self._open_ai_log_window,
        ).pack(side=tk.LEFT)

        tk.Label(
            ai_inner,
            text="API keys are stored in config.json. Keep that file secure and out of version control.",
            bg=PANEL,
            fg=ERROR,
            font=("Segoe UI", 8),
            wraplength=850,
        ).pack(anchor="w", pady=(6, 0))

        # Processing parameters
        sb = self._section(frame, "Processing Parameters")
        self._extract_text_var      = tk.BooleanVar(value=self.cfg.get("app_settings", {}).get("extract_text_before_ocr", True))
        self._enhance_var           = tk.BooleanVar(value=self.cfg.get("app_settings", {}).get("enhance_images", True))
        self._fallback_var          = tk.BooleanVar(value=self.cfg.get("app_settings", {}).get("ocr_fallback_to_tesseract", True))
        self._extraction_source_var = tk.StringVar(value=self.cfg.get("app_settings", {}).get("extraction_source", "auto"))
        self._max_threads_var       = tk.IntVar(value=self.cfg.get("app_settings", {}).get("max_threads", 4))
        self._threshold_var         = tk.IntVar(value=self.cfg.get("app_settings", {}).get("fuzzy_match_threshold", 85))
        self._scale_var             = tk.IntVar(value=self.cfg.get("app_settings", {}).get("image_scale_factor", 2))

        sb_inner = tk.Frame(sb, bg=PANEL)
        sb_inner.pack(fill=tk.X, padx=20, pady=(0, 12))

        checks = [
            ("Extract native PDF text before OCR",      self._extract_text_var),
            ("Enhance images before OCR",               self._enhance_var),
            ("Fallback to Tesseract if engine fails",   self._fallback_var),
            ("Enable parallel processing (multi-threading)", self._threads_var),
        ]
        for lbl, var in checks:
            ttk.Checkbutton(sb_inner, text=lbl, variable=var, style="TCheckbutton").pack(anchor="w", pady=3)

        tk.Label(sb_inner, text="Extraction source:", bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(10, 2))
        for val, lbl in [
            ("auto",          "Auto — PDF text first, then image OCR if needed"),
            ("pdf_text_only", "PDF text only — no image OCR"),
            ("image_only",    "Image OCR only — ignore embedded text"),
        ]:
            ttk.Radiobutton(sb_inner, text=lbl, value=val, variable=self._extraction_source_var,
                            style="TRadiobutton").pack(anchor="w", padx=20, pady=2)

        g = tk.Frame(sb_inner, bg=PANEL)
        g.pack(fill=tk.X, pady=(12, 0))
        for r, (lbl, var, lo, hi) in enumerate([
            ("Max threads:",                       self._max_threads_var,  1, 12),
            ("Fuzzy match threshold:",             self._threshold_var,   50, 100),
            ("Image scale factor (Custom mode):",  self._scale_var,        1,  5),
        ]):
            tk.Label(g, text=lbl, bg=PANEL, fg=TEXT, font=("Segoe UI", 10), width=34, anchor="w").grid(row=r, column=0, sticky="w", pady=6)
            ttk.Spinbox(g, from_=lo, to=hi, textvariable=var, width=8, font=("Segoe UI", 10)).grid(row=r, column=1, sticky="w", padx=8)

        # ── OCR Word Correction ───────────────────────────────────────────
        ocr_corr_box = self._section(
            frame,
            "OCR Word Correction",
            subtitle="Fixes broken OCR words by fuzzy-matching against known supplier names",
        )
        ocr_corr_inner = tk.Frame(ocr_corr_box, bg=PANEL)
        ocr_corr_inner.pack(fill=tk.X, padx=20, pady=(0, 12))

        self._ocr_word_correction_var = tk.BooleanVar(
            value=self.cfg.get("app_settings", {}).get("ocr_word_correction_enabled", False)
        )
        ttk.Checkbutton(
            ocr_corr_inner,
            text="Enable OCR Word Correction — fix garbled supplier words (e.g. EASRN → EASTERN)",
            variable=self._ocr_word_correction_var,
            style="TCheckbutton",
        ).pack(anchor="w", pady=4)
        tk.Label(
            ocr_corr_inner,
            text=(
                "Allows 1–2 wrong characters per word (longer words allow slightly more).\n"
                "Turn OFF if you get false supplier matches. Requires ocr_word_corrector.py."
            ),
            bg=PANEL, fg=MUTED, font=("Segoe UI", 9), wraplength=850, justify="left",
        ).pack(anchor="w", pady=(0, 4))

        if not OCR_CORRECTOR_AVAILABLE:
            tk.Label(
                ocr_corr_inner,
                text="⚠  ocr_word_corrector.py not found in the application folder.",
                bg=PANEL, fg=WARNING, font=("Segoe UI", 9),
            ).pack(anchor="w")

        # ── Smart Cross-Matching ──────────────────────────────────────────
        xcm_box = self._section(
            frame,
            "Smart Supplier ↔ Invoice Cross-Matching",
            subtitle="Infer supplier from invoice format, and vice versa",
        )
        xcm_inner = tk.Frame(xcm_box, bg=PANEL)
        xcm_inner.pack(fill=tk.X, padx=20, pady=(0, 12))

        self._smart_cross_match_var = tk.BooleanVar(
            value=self.cfg.get("app_settings", {}).get("smart_cross_match_enabled", False)
        )
        ttk.Checkbutton(
            xcm_inner,
            text="Enable Smart Cross-Matching — identify supplier from invoice number pattern, and vice versa",
            variable=self._smart_cross_match_var,
            style="TCheckbutton",
        ).pack(anchor="w", pady=4)
        tk.Label(
            xcm_inner,
            text=(
                "When enabled:\n"
                "  • If the supplier is found, the invoice number is cleaned using that supplier's known format.\n"
                "  • If the supplier is NOT found, the invoice number pattern is used to identify the supplier.\n"
                "  • When OFF: the system works exactly as before — no cross-matching.\n"
                "Edit SUPPLIER_INVOICE_XREF in smart_cross_matcher.py to add your invoice formats."
            ),
            bg=PANEL, fg=MUTED, font=("Segoe UI", 9), wraplength=850, justify="left",
        ).pack(anchor="w", pady=(0, 4))

        if not CROSS_MATCHER_AVAILABLE:
            tk.Label(
                xcm_inner,
                text="⚠  smart_cross_matcher.py not found in the application folder.",
                bg=PANEL, fg=WARNING, font=("Segoe UI", 9),
            ).pack(anchor="w")

        # Base path
        pb = self._section(frame, "Base Folder Path")
        self._base_var = tk.StringVar(value=self.dirs["base"])
        pb_row = tk.Frame(pb, bg=PANEL)
        pb_row.pack(fill=tk.X, padx=20, pady=(0, 16))
        tk.Label(pb_row, text="Base folder:", bg=PANEL, fg=MUTED, font=("Segoe UI", 9, "bold"), width=14, anchor="w").pack(side=tk.LEFT)
        ent_base = tk.Entry(pb_row, textvariable=self._base_var, bg=PANEL2, fg=TEXT,
                            insertbackground=TEXT, relief="flat", font=("Segoe UI", 10), bd=0,
                            highlightthickness=1, highlightbackground=PANEL3, highlightcolor=ACCENT)
        ent_base.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=6, ipady=5)
        ttk.Button(pb_row, text="Browse", command=self._browse_base_folder).pack(side=tk.LEFT, padx=(6, 0))

        # Actions
        ab = self._section(frame, "Configuration Actions")
        ab_row = tk.Frame(ab, bg=PANEL)
        ab_row.pack(fill=tk.X, padx=20, pady=(0, 20))
        ttk.Button(ab_row, text="🔧  Test Engine",       command=self._test_current_engine).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(ab_row, text="↺  Reload Config",      command=self._reload_settings).pack(side=tk.LEFT, padx=4)
        ttk.Button(ab_row, text="💾  Save All Settings", style="Accent.TButton",
                   command=self._apply_settings).pack(side=tk.RIGHT)

        # Suppliers
        sup_box = self._section(frame, "Suppliers & Aliases")
        tk.Label(
            sup_box,
            text="Add or remove suppliers. Aliases are alternative names that map to the same supplier.",
            bg=PANEL, fg=MUTED, font=("Segoe UI", 9), wraplength=900,
        ).pack(anchor="w", padx=20, pady=(0, 6))

        # --- Search bar for supplier tree ---
        sup_search_bar = tk.Frame(sup_box, bg=PANEL)
        sup_search_bar.pack(fill=tk.X, padx=20, pady=(0, 6))
        tk.Label(
            sup_search_bar, text="Search:", bg=PANEL, fg=MUTED, font=("Segoe UI", 9),
        ).pack(side=tk.LEFT, padx=(0, 6))
        self._sup_search_var = tk.StringVar()
        sup_search_ent = tk.Entry(
            sup_search_bar,
            textvariable=self._sup_search_var,
            bg=PANEL2, fg=TEXT, insertbackground=TEXT,
            relief="flat", font=("Segoe UI", 10), bd=0,
            highlightthickness=1, highlightbackground=PANEL3, highlightcolor=ACCENT,
        )
        sup_search_ent.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=5)

        def _clear_sup_search():
            self._sup_search_var.set("")
        ttk.Button(sup_search_bar, text="✕", command=_clear_sup_search, width=3).pack(side=tk.LEFT, padx=(6, 0))

        # --- Supplier treeview ---
        sup_tf = tk.Frame(sup_box, bg=PANEL)
        sup_tf.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 8))

        sup_cols = ("Supplier Name", "Aliases (comma-separated)")
        self._sup_tree = ttk.Treeview(sup_tf, columns=sup_cols, show="headings", height=10)
        self._sup_tree.heading("Supplier Name",              text="Supplier Name",              anchor="w")
        self._sup_tree.heading("Aliases (comma-separated)",  text="Aliases (comma-separated)",  anchor="w")
        self._sup_tree.column("Supplier Name",             width=220, anchor="w")
        self._sup_tree.column("Aliases (comma-separated)", width=520, anchor="w")

        sup_ys = ttk.Scrollbar(sup_tf, orient="vertical", command=self._sup_tree.yview)
        self._sup_tree.configure(yscrollcommand=sup_ys.set)
        self._sup_tree.grid(row=0, column=0, sticky="nsew")
        sup_ys.grid(row=0, column=1, sticky="ns")
        sup_tf.rowconfigure(0, weight=1)
        sup_tf.columnconfigure(0, weight=1)
        self._bind_tree_mousewheel(self._sup_tree)

        self._populate_supplier_tree()
        self._make_tree_editable(
            self._sup_tree,
            on_edit_callback=self._on_sup_tree_edit,
            editable_cols={0, 1},
            pre_edit_fn=None,
        )

        # Supplier tree column 0 is editable, not preview-only
        self._sup_tree._edit_on_preview_cb = None

        # Wire up live search against the supplier tree
        def _on_sup_search(*_):
            q = self._sup_search_var.get().strip().lower()
            # Re-populate filtered view
            for iid in self._sup_tree.get_children():
                self._sup_tree.detach(iid)
            # Keep a master list of all row ids
            for iid in getattr(self, '_sup_all_rows', []):
                if not q:
                    self._sup_tree.reattach(iid, "", "end")
                else:
                    v = self._sup_tree.item(iid, "values")
                    if any(q in str(x).lower() for x in v):
                        self._sup_tree.reattach(iid, "", "end")

        self._sup_search_var.trace_add("write", _on_sup_search)

        sup_ctrl = tk.Frame(sup_box, bg=PANEL)
        sup_ctrl.pack(fill=tk.X, padx=20, pady=(0, 20))
        ttk.Button(sup_ctrl, text="+ Add",    command=self._add_supplier_row).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(sup_ctrl, text="✕ Remove", command=self._remove_supplier_row).pack(side=tk.LEFT, padx=4)
        ttk.Button(
            sup_ctrl, text="💾  Save", style="Accent.TButton",
            command=self._save_suppliers,
        ).pack(side=tk.RIGHT)

    # ------------------------------------------------------------------
    # ABOUT TAB
    # ------------------------------------------------------------------
    def _build_about_tab(self):
        frame = self._make_scrollable_tab("About")
        center = tk.Frame(frame, bg=BG)
        center.pack(expand=True, pady=30, fill=tk.X)

        alp = resource_path("logo2.png")
        if alp.exists():
            try:
                img = Image.open(alp)
                img.thumbnail((480, 160))
                self._about_logo_img = ImageTk.PhotoImage(img)
                tk.Label(center, image=self._about_logo_img, bg=BG).pack(pady=(0, 16))
            except Exception:
                pass

        tk.Label(center, text="Document Processing Hub", bg=BG, fg=TEXT, font=("Segoe UI", 20, "bold")).pack()
        tk.Label(center, text=self.APP_VERSION, bg=BG, fg=MUTED, font=("Segoe UI", 11)).pack(pady=(4, 4))
        tk.Label(center, text="Author: Roni", bg=BG, fg=ACCENT, font=("Segoe UI", 13, "bold")).pack(pady=(0, 20))

        items = [
            ("Auto-Ingest",       "watchdog monitors SCANNED — new PDFs process automatically (toggle in Settings)"),
            ("Notifications",     "plyer sends desktop alerts when batches complete or files fail"),
            ("Confidence Score",  "Every supplier match shows % confidence; rows below threshold are highlighted"),
            ("Supplier Learning", "Manual corrections offer to save as aliases for future auto-matching"),
            ("Dry Run Preview",   "Proposes filenames in a popup before any files are moved"),
            ("Retry Failed",      "One-click retry for all files in the FAILED folder"),
            ("Duplicate Guard",   "Warns when an invoice number already exists in PROCESSED"),
            ("Live Search",       "Filter bars above both result tables update rows as you type"),
            ("PDF Preview",       "Double-click a filename in results to see the first page thumbnail"),
            ("Linked Views",      "Corrections in Renamer sync to Dispatch and vice-versa"),
            ("Multi-GRN",         "Chains multiple GRNs: RC-MAM-000019581-19582"),
        ]

        grid = tk.Frame(center, bg=PANEL, padx=24, pady=20)
        grid.pack(fill=tk.X, padx=20)
        for i, (lbl, val) in enumerate(items):
            tk.Label(grid, text=lbl + ":", bg=PANEL, fg=ACCENT, font=("Segoe UI", 9, "bold"),
                     width=20, anchor="w").grid(row=i, column=0, sticky="w", pady=4, padx=(0, 12))
            tk.Label(grid, text=val, bg=PANEL, fg=MUTED, font=("Segoe UI", 9),
                     anchor="w").grid(row=i, column=1, sticky="w", pady=4)

        deps = [
            ("watchdog",  "pip install watchdog",             WATCHDOG_AVAILABLE),
            ("plyer",     "pip install plyer",                PLYER_AVAILABLE),
            ("easyocr",   "pip install easyocr",              self._check_engine("easyocr")),
            ("paddleocr", "pip install paddlepaddle paddleocr", self._check_engine("paddleocr")),
        ]
        dep_frame = tk.Frame(center, bg=PANEL, padx=24, pady=16)
        dep_frame.pack(fill=tk.X, padx=20, pady=(12, 0))
        tk.Label(dep_frame, text="Optional Dependencies", bg=PANEL, fg=TEXT,
                 font=("Segoe UI", 10, "bold")).pack(anchor="w", pady=(0, 8))
        for name, install, ok in deps:
            r = tk.Frame(dep_frame, bg=PANEL)
            r.pack(fill=tk.X, pady=2)
            dot_color = SUCCESS if ok else MUTED
            tk.Label(r, text="●", bg=PANEL, fg=dot_color, font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=(0, 8))
            tk.Label(r, text=name, bg=PANEL, fg=TEXT, font=("Consolas", 9), width=14, anchor="w").pack(side=tk.LEFT)
            tk.Label(r, text=install if not ok else "installed", bg=PANEL,
                     fg=MUTED if not ok else SUCCESS, font=("Consolas", 9)).pack(side=tk.LEFT)

    # ------------------------------------------------------------------
    # STATUS / FOLDERS / DASHBOARD
    # ------------------------------------------------------------------
    def _set_status(self, text: str, dot_color: str = None):
        logging.info(text)
        try:
            if hasattr(self, "_status_var"):
                self.after(0, lambda t=text: self._status_var.set(t))
            if dot_color and hasattr(self, "_status_dot"):
                self.after(0, lambda c=dot_color: self._status_indicator.itemconfig(self._status_dot, fill=c))
        except Exception:
            pass

    def _open_folder(self, path):
        try:
            os.makedirs(path, exist_ok=True)
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("Open Folder Failed", str(e))

    def _refresh_dashboard_stats(self):
        try:
            def cpdf(f):
                return len([x for x in os.listdir(f) if x.lower().endswith(".pdf")]) if f and os.path.exists(f) else 0
            v = {
                "waiting":   cpdf(self.dirs.get("scanned")),
                "processed": cpdf(self.dirs.get("processed")),
                "archived":  cpdf(self.dirs.get("archive") or self.dirs.get("archived")),
                "failed":    cpdf(self.dirs.get("failed")),
            }
            for attr, key in [
                ("_stat_waiting_var",   "waiting"),
                ("_stat_processed_var", "processed"),
                ("_stat_archived_var",  "archived"),
                ("_stat_failed_var",    "failed"),
            ]:
                if hasattr(self, attr):
                    getattr(self, attr).set(str(v[key]))
        except Exception as e:
            logging.error(f"Dashboard refresh failed: {e}")

    # ------------------------------------------------------------------
    # TREE EDITING (GENERIC)
    # ------------------------------------------------------------------
    def _make_tree_editable(self, tree, on_edit_callback=None, editable_cols=None, pre_edit_fn=None):
        """
        Bind double-click editing to a Treeview.

        Column 0 is reserved for PDF preview (handled by the caller via
        on_preview_callback).  All other editable columns open an inline
        Entry widget.

        We store the callback refs on the tree widget itself so the binding
        is never replaced by a second call.
        """
        # Attach metadata to the tree so we can retrieve it from the handler
        tree._edit_editable_cols   = editable_cols
        tree._edit_on_edit_cb      = on_edit_callback
        tree._edit_pre_edit_fn     = pre_edit_fn

        def on_double_click(event):
            region = tree.identify("region", event.x, event.y)
            if region != "cell":
                return

            col_id = tree.identify_column(event.x)
            row_id = tree.identify_row(event.y)
            if not row_id:
                return

            ci = int(col_id[1:]) - 1   # 0-based column index

            # Column 0 → PDF preview only if a preview callback is registered
            if ci == 0:
                preview_cb = getattr(tree, '_edit_on_preview_cb', None)
                if preview_cb is not None:
                    preview_cb(row_id)
                    return
                # No preview callback = treat col 0 as editable (e.g. supplier tree)
				

            # Check editable
            ec = tree._edit_editable_cols
            if ec is not None and ci not in ec:
                return

            vals = list(tree.item(row_id, "values"))
            if ci >= len(vals):
                return

            cur         = str(vals[ci])
            pre_fn      = tree._edit_pre_edit_fn
            display_val = pre_fn(ci, cur) if pre_fn else cur

            bbox = tree.bbox(row_id, col_id)
            if not bbox:
                return
            bx, by, bw, bh = bbox

            ev  = tk.StringVar(value=display_val)
            ent = tk.Entry(
                tree, textvariable=ev,
                background=PANEL2, foreground=TEXT,
                insertbackground=TEXT, relief="flat",
                font=("Segoe UI", 9, "bold"), bd=2,
                highlightthickness=1,
                highlightbackground=ACCENT, highlightcolor=ACCENT,
            )
            ent.place(x=bx, y=by, width=bw, height=bh)
            ent.focus_set()
            ent.select_range(0, tk.END)
            done = [False]

            def commit(e=None):
                if done[0]:
                    return
                done[0] = True
                nv = ev.get().strip()
                ent.destroy()
                if not nv or nv == display_val:
                    return
                cb = tree._edit_on_edit_cb
                if cb:
                    cb(row_id, ci, cur, nv)
                else:
                    vals[ci] = nv
                    tree.item(row_id, values=vals)

            def cancel(e=None):
                done[0] = True
                ent.destroy()

            ent.bind("<Return>",   commit)
            ent.bind("<KP_Enter>", commit)
            ent.bind("<Escape>",   cancel)
            ent.bind("<FocusOut>", commit)

        # Single binding — no duplicate
        tree.bind("<Double-1>", on_double_click)

    # ------------------------------------------------------------------
    # DRY RUN PREVIEW POPUP
    # ------------------------------------------------------------------
    def _show_dry_run_preview(self, previews: List[Dict]) -> bool:
        win = tk.Toplevel(self)
        win.title("Dry Run Preview — Proposed Filenames")
        win.geometry("1080x580")
        win.configure(bg=BG)
        win.grab_set()

        tk.Label(win, text="Review proposed renames below.",
                 bg=BG, fg=TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=20, pady=(16, 2))
        tk.Label(win,
                 text="Click  ✓ Proceed  to move files, or  ✕ Cancel  to abort. Orange rows have duplicate invoice warnings.",
                 bg=BG, fg=MUTED, font=("Segoe UI", 9)).pack(anchor="w", padx=20, pady=(0, 10))

        tf = tk.Frame(win, bg=BG)
        tf.pack(fill=tk.BOTH, expand=True, padx=16, pady=4)
        cols = ("Original File", "Proposed New Name", "Supplier", "Conf %", "GRN", "Invoice #", "Duplicate?")
        cw = {"Original File": 200, "Proposed New Name": 260, "Supplier": 180,
              "Conf %": 60, "GRN": 170, "Invoice #": 115, "Duplicate?": 175}
        tree = ttk.Treeview(tf, columns=cols, show="headings", height=14)
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=cw.get(c, 100), anchor="center", stretch=False)

        tree.tag_configure("dupe", foreground=WARNING)
        tree.tag_configure("ok",   foreground=SUCCESS)
        tree.tag_configure("err",  foreground=ERROR)

        ys = ttk.Scrollbar(tf, orient="vertical",   command=tree.yview)
        xs = ttk.Scrollbar(tf, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        tree.grid(row=0, column=0, sticky="nsew")
        ys.grid(row=0, column=1, sticky="ns")
        xs.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1)
        tf.columnconfigure(0, weight=1)
        self._bind_tree_mousewheel(tree)

        for p in previews:
            dupe = p.get("duplicate_warning", "")
            err  = "ERROR" in p.get("proposed_name", "")
            tag  = "err" if err else ("dupe" if dupe else "ok")
            conf_disp = f"{p.get('confidence', 0.0):.0f}%" if not err else "—"
            tree.insert("", "end", tags=(tag,), values=(
                p.get("file", ""),
                p.get("proposed_name", ""),
                p.get("supplier", ""),
                conf_disp,
                p.get("grn", ""),
                p.get("invoice", ""),
                dupe or "—",
            ))

        result = [False]

        def proceed():
            result[0] = True
            win.destroy()

        btn_row = tk.Frame(win, bg=BG)
        btn_row.pack(fill=tk.X, padx=20, pady=(8, 16))
        ttk.Button(btn_row, text="✓  Proceed — Move Files Now", style="Accent.TButton",
                   command=proceed).pack(side=tk.LEFT, padx=(0, 12))
        ttk.Button(btn_row, text="✕  Cancel", command=win.destroy).pack(side=tk.LEFT)
        n_dupe = sum(1 for p in previews if p.get("duplicate_warning"))
        tk.Label(btn_row, text=f"{len(previews)} files  ·  {n_dupe} duplicate invoice(s)",
                 bg=BG, fg=MUTED, font=("Segoe UI", 9)).pack(side=tk.RIGHT, padx=8)

        win.wait_window()
        return result[0]

    # ------------------------------------------------------------------
    # PDF PREVIEW PANEL
    # ------------------------------------------------------------------
    def _show_pdf_preview(self, pdf_path: str):
        if not pdf_path or not os.path.exists(pdf_path):
            messagebox.showwarning("Preview", f"File not found:\n{pdf_path or '(no path)'}")
            return

        win = tk.Toplevel(self)
        win.title(f"Preview — {os.path.basename(pdf_path)}")
        win.geometry("900x980")
        win.configure(bg=BG)

        top = tk.Frame(win, bg=PANEL2, height=46)
        top.pack(fill=tk.X)
        top.pack_propagate(False)

        tk.Label(
            top,
            text=os.path.basename(pdf_path),
            bg=PANEL2,
            fg=TEXT,
            font=("Segoe UI", 10, "bold"),
        ).pack(side=tk.LEFT, padx=14, pady=12)

        page_info_var = tk.StringVar(value="Loading...")
        tk.Label(
            top,
            textvariable=page_info_var,
            bg=PANEL2,
            fg=MUTED,
            font=("Segoe UI", 9),
        ).pack(side=tk.RIGHT, padx=14)

        toolbar = tk.Frame(win, bg=PANEL, height=42)
        toolbar.pack(fill=tk.X)
        toolbar.pack_propagate(False)

        canvas = tk.Canvas(win, bg=BG, highlightthickness=0, cursor="fleur")
        sb_y = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        sb_x = ttk.Scrollbar(win, orient="horizontal", command=canvas.xview)
        canvas.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)

        sb_y.pack(side=tk.RIGHT, fill=tk.Y)
        sb_x.pack(side=tk.BOTTOM, fill=tk.X)
        canvas.pack(fill=tk.BOTH, expand=True)

        state = {
            "doc": None,
            "page_index": 0,
            "page_count": 0,
            "zoom": 1.0,
            "fit_zoom": 1.0,
            "photo": None,
            "page_width": 0,
            "page_height": 0,
        }

        def render_page():
            try:
                if not state["doc"]:
                    return
                page = state["doc"][state["page_index"]]
                z = max(0.2, state["zoom"])
                pix = page.get_pixmap(matrix=fitz.Matrix(z, z), alpha=False)
                img_data = pix.tobytes("png")
                pil_img = PILImage.open(io.BytesIO(img_data))
                photo = ImageTk.PhotoImage(pil_img)

                canvas.delete("all")
                canvas.create_image(0, 0, anchor="nw", image=photo, tags="page")
                canvas.configure(scrollregion=(0, 0, pil_img.width, pil_img.height))

                state["photo"] = photo
                state["page_width"] = pil_img.width
                state["page_height"] = pil_img.height

                zoom_pct = int(state["zoom"] * 100)
                page_info_var.set(
                    f"Page {state['page_index'] + 1} of {state['page_count']}  ·  {zoom_pct}%  ·  {pil_img.width}x{pil_img.height}"
                )
            except Exception as e:
                page_info_var.set(f"Error: {e}")

        def fit_to_window():
            if not state["doc"]:
                return
            page = state["doc"][state["page_index"]]
            rect = page.rect
            canvas.update_idletasks()
            cw = max(canvas.winfo_width() - 20, 100)
            ch = max(canvas.winfo_height() - 20, 100)
            zx = cw / rect.width
            zy = ch / rect.height
            state["fit_zoom"] = min(zx, zy)
            state["zoom"] = state["fit_zoom"]
            render_page()

        def zoom_in():
            state["zoom"] = min(state["zoom"] * 1.2, 5.0)
            render_page()

        def zoom_out():
            state["zoom"] = max(state["zoom"] / 1.2, 0.2)
            render_page()

        def zoom_100():
            state["zoom"] = 1.0
            render_page()

        def zoom_fit():
            fit_to_window()

        def prev_page():
            if state["doc"] and state["page_index"] > 0:
                state["page_index"] -= 1
                fit_to_window()

        def next_page():
            if state["doc"] and state["page_index"] < state["page_count"] - 1:
                state["page_index"] += 1
                fit_to_window()

        ttk.Button(toolbar, text="Fit", command=zoom_fit).pack(side=tk.LEFT, padx=(10, 4), pady=7)
        ttk.Button(toolbar, text="100%", command=zoom_100).pack(side=tk.LEFT, padx=4, pady=7)
        ttk.Button(toolbar, text="-", command=zoom_out, width=3).pack(side=tk.LEFT, padx=4, pady=7)
        ttk.Button(toolbar, text="+", command=zoom_in, width=3).pack(side=tk.LEFT, padx=4, pady=7)
        ttk.Button(toolbar, text="Prev", command=prev_page).pack(side=tk.LEFT, padx=(18, 4), pady=7)
        ttk.Button(toolbar, text="Next", command=next_page).pack(side=tk.LEFT, padx=4, pady=7)

        def start_pan(event):
            canvas.scan_mark(event.x, event.y)

        def do_pan(event):
            canvas.scan_dragto(event.x, event.y, gain=1)

        canvas.bind("<ButtonPress-1>", start_pan)
        canvas.bind("<B1-Motion>", do_pan)

        def on_mousewheel(event):
            if event.state & 0x0004:  # Ctrl pressed
                if event.delta > 0:
                    zoom_in()
                else:
                    zoom_out()
                return "break"
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            return "break"

        canvas.bind("<MouseWheel>", on_mousewheel)
        canvas.bind("<Button-4>", lambda e: (canvas.yview_scroll(-1, "units"), "break"))
        canvas.bind("<Button-5>", lambda e: (canvas.yview_scroll(1, "units"), "break"))

        def load():
            try:
                doc = fitz.open(pdf_path)
                state["doc"] = doc
                state["page_count"] = doc.page_count
                win.after(0, fit_to_window)
            except Exception as e:
                win.after(0, lambda: page_info_var.set(f"Error: {e}"))

        def on_resize(_event):
            if state["doc"]:
                pass

        canvas.bind("<Configure>", on_resize)
        threading.Thread(target=load, daemon=True).start()

    def _on_rename_preview_by_rowid(self, row_id: str):
        result   = self._rename_row_map.get(row_id, {})
        pdf_path = result.get("dest_path", "")
        if not pdf_path or not os.path.exists(pdf_path):
            fname    = result.get("file", "")
            pdf_path = os.path.join(self.dirs.get("scanned", ""), fname)
        self._show_pdf_preview(pdf_path)

    def _on_rename_tree_double_click(self, event):
        if self._rename_tree.identify("region", event.x, event.y) != "cell":
            return
        col_id = self._rename_tree.identify_column(event.x)
        ci = int(col_id[1:]) - 1
        if ci != 0:
            return
        row_id = self._rename_tree.identify_row(event.y)
        if not row_id:
            return
        result = self._rename_row_map.get(row_id, {})
        pdf_path = result.get("dest_path", "")
        if not pdf_path or not os.path.exists(pdf_path):
            fname = result.get("file", "")
            pdf_path = os.path.join(self.dirs.get("scanned", ""), fname)
        self._show_pdf_preview(pdf_path)

    def _on_dispatch_tree_double_click(self, event):
        if self._dispatch_tree.identify("region", event.x, event.y) != "cell":
            return
        col_id = self._dispatch_tree.identify_column(event.x)
        ci = int(col_id[1:]) - 1
        if ci != 0:
            return
        row_id = self._dispatch_tree.identify_row(event.y)
        if not row_id:
            return
        result = self._dispatch_row_map.get(row_id, {})
        pdf_path = result.get("raw_path", "")
        folder = self._dispatch_folder_var.get().strip() or self.dirs.get("processed", "")
        if not pdf_path or not os.path.exists(pdf_path):
            fname = result.get("file", "")
            pdf_path = os.path.join(folder, fname)
        self._show_pdf_preview(pdf_path)

    # ------------------------------------------------------------------
    # RETRY FAILED FILES
    # ------------------------------------------------------------------
    def _retry_failed_files(self):
        failed_dir  = self.dirs.get("failed", "")
        scanned_dir = self.dirs.get("scanned", "")
        if not os.path.isdir(failed_dir):
            messagebox.showinfo("Retry Failed", "FAILED folder not found.")
            return
        failed_pdfs = [f for f in os.listdir(failed_dir) if f.lower().endswith(".pdf")]
        if not failed_pdfs:
            messagebox.showinfo("Retry Failed", "No PDF files in FAILED folder.")
            return
        preview = "\n".join(failed_pdfs[:10]) + ("\n..." if len(failed_pdfs) > 10 else "")
        if not messagebox.askyesno(
            "Retry Failed Files",
            f"Move {len(failed_pdfs)} file(s) from FAILED → SCANNED and re-process?\n\n{preview}",
        ):
            return
        moved = 0
        for fn in failed_pdfs:
            try:
                shutil.move(os.path.join(failed_dir, fn), os.path.join(scanned_dir, fn))
                moved += 1
            except Exception as e:
                logging.error(f"Retry move failed [{fn}]: {e}")
        self._set_status(f"Moved {moved} file(s) to SCANNED — starting rename...", SUCCESS)
        self._refresh_dashboard_stats()
        self.after(300, self._start_rename_worker)

    # ------------------------------------------------------------------
    # SETTINGS CALLBACKS
    # ------------------------------------------------------------------
    def _on_engine_changed(self):
        self.cfg.setdefault("app_settings", {})["ocr_engine"] = self._engine_var.get()
        self._refresh_engine_badge()

    def _install_engine(self, key):
        messagebox.showinfo("Install Engine",
                            f"To install {ENGINE_LABELS.get(key, key)}, run:\n\n  {ENGINE_INSTALL.get(key, '')}")

    def _browse_base_folder(self):
        f = filedialog.askdirectory(initialdir=self._base_var.get() or self.dirs["base"])
        if f:
            self._base_var.set(f)

    def _apply_settings(self):
        s = self.cfg.setdefault("app_settings", {})
        s["ocr_engine"] = self._engine_var.get()
        s["extract_text_before_ocr"] = bool(self._extract_text_var.get())
        s["extraction_source"] = self._extraction_source_var.get()
        s["enhance_images"] = bool(self._enhance_var.get())
        s["ocr_fallback_to_tesseract"] = bool(self._fallback_var.get())
        s["enable_multi_threading"] = bool(self._threads_var.get())
        s["max_threads"] = int(self._max_threads_var.get())
        s["fuzzy_match_threshold"] = int(self._threshold_var.get())
        s["image_scale_factor"] = int(self._scale_var.get())
        s["processing_mode"] = self._processing_mode_var.get()
        s["dry_run"] = bool(self._dry_run_var.get())
        s["desktop_notifications"] = bool(self._notify_var.get())
        s["auto_ingest_offline"] = bool(self._watcher_offline_var.get())
        s["auto_ingest_api"] = bool(self._watcher_api_var.get())
        s["auto_ingest_watcher"] = s["auto_ingest_offline"] or s["auto_ingest_api"]
        s["confidence_warn_threshold"] = int(self._conf_threshold_var.get())
        s["page_scan_region_enabled"] = bool(self._page_scan_enabled_var.get())
        s["page_scan_region_percent"] = int(self._page_scan_percent_var.get())
        s["ocr_mode"] = self._ocr_mode_var.get()
        s["supplier_extraction_method"] = self._supplier_method_var.get()
        s["ocr_word_correction_enabled"] = bool(self._ocr_word_correction_var.get())
        s["smart_cross_match_enabled"] = bool(self._smart_cross_match_var.get())

        if hasattr(self, "_supplier_strategy_var"):
            s["supplier_match_strategy"] = self._supplier_strategy_var.get()

        if hasattr(self, "_ocr_mode_info_var"):
            self._ocr_mode_info_var.set(self._ocr_mode_label(s["ocr_mode"]))

        self.cfg.setdefault("folders", {})["base"] = self._base_var.get()

        ai = self.cfg.setdefault("ai_settings", {})
        ai["enabled"] = bool(self._ai_enabled_var.get())
        ai["provider"] = self._ai_provider_var.get()
        ai["model"] = self._ai_model_var.get().strip()
        ai["api_key"] = self._ai_key_var.get().strip()
        ai["custom_base_url"] = self._ai_url_var.get().strip()
        ai["timeout_seconds"] = int(self._ai_timeout_var.get())

        ocr_space = self.cfg.setdefault("ocr_space", {})
        ocr_space["api_key"] = self._ocr_space_key_var.get().strip() or "K88109865088957"
        ocr_space["language"] = "eng"
        ocr_space["isOverlayRequired"] = False
        ocr_space["detectOrientation"] = True
        ocr_space["scale"] = True
        ocr_space["OCREngine"] = int(self._ocr_space_engine_var.get())
        ocr_space["isTable"] = False
        ocr_space["filetype"] = "PDF"
        ocr_space["timeout_seconds"] = 30
        try:
            ocr_space["max_upload_mb"] = float(self._ocr_space_max_mb_var.get())
        except Exception:
            ocr_space["max_upload_mb"] = 1.0

        self._save_config()
        self.dirs = self._resolve_dirs()
        self._ensure_dirs()
        self._configure_tesseract()
        self._refresh_engine_badge()
        self._refresh_dashboard_stats()

        if AI_MATCHER_AVAILABLE:
            try:
                from ai_supplier_matcher import AISupplierMatcher
                self._ai_matcher = AISupplierMatcher(self.cfg, logger_func=self._append_ai_log)
                self._append_ai_log("AI matcher reloaded with latest settings.")
            except Exception as e:
                logging.warning(f"AI matcher reload failed: {e}")
                self._append_ai_log(f"AI matcher reload failed: {e}")

        messagebox.showinfo("Settings Saved", "All settings saved successfully.")

    def _reload_settings(self):
        self.cfg = self._load_config()
        self.dirs = self._resolve_dirs()
        self._ensure_dirs()
        self._configure_tesseract()
        self._refresh_engine_badge()
        self._threads_var.set(self.cfg.get("app_settings", {}).get("enable_multi_threading", True))
        self._set_status("Configuration reloaded.")
        messagebox.showinfo("Reloaded", "Configuration reloaded from disk.")

    def _on_processing_mode_changed(self):
        mode = self._processing_mode_var.get()
        self.cfg.setdefault("app_settings", {})["processing_mode"] = mode
        self._save_config()

    def _on_parallel_toggle(self):
        val = bool(self._threads_var.get())
        self.cfg.setdefault("app_settings", {})["enable_multi_threading"] = val
        self._save_config()

    def _on_dry_run_toggle(self):
        val = bool(self._dry_run_var.get())
        self.cfg.setdefault("app_settings", {})["dry_run"] = val
        self._save_config()
        self._set_status("Dry Run ON — preview before committing." if val else "Dry Run OFF — files will be moved.")

    def _on_supplier_method_changed(self):
        method = self._supplier_method_var.get()
        self.cfg.setdefault("app_settings", {})["supplier_extraction_method"] = method
        self._save_config()

    def _on_notify_toggle(self):
        val = bool(self._notify_var.get())
        self.cfg.setdefault("app_settings", {})["desktop_notifications"] = val
        self._save_config()
        self._set_status("Desktop notifications: ON" if val else "Desktop notifications: OFF")

    def _ocr_mode_label(self, mode: str) -> str:
        return {
            "full": "Active: Full OCR — entire page scanned",
            "zone": "Active: Zone OCR — header region only (fast mode)",
        }.get(mode, "Active: Full OCR")

    def _on_ocr_mode_changed(self):
        mode = self._ocr_mode_var.get()
        self.cfg.setdefault("app_settings", {})["ocr_mode"] = mode
        if hasattr(self, "_ocr_mode_info_var"):
            self._ocr_mode_info_var.set(self._ocr_mode_label(mode))
        self._save_config()
        self._set_status(
            "OCR Mode set to ZONE (fast header scan)." if mode == "zone"
            else "OCR Mode set to FULL PAGE."
        )

    def _on_supplier_strategy_changed(self):
        strategy = self._supplier_strategy_var.get()
        self.cfg.setdefault("app_settings", {})["supplier_match_strategy"] = strategy
        self._save_config()
        self._set_status(f"Supplier match strategy set to: {STRATEGY_LABELS.get(strategy, strategy)}")

    def _on_ai_enabled_toggle(self):
        val = bool(self._ai_enabled_var.get())
        self.cfg.setdefault("ai_settings", {})["enabled"] = val
        self._save_config()
        self._set_status("AI matching: ON" if val else "AI matching: OFF")

    def _on_ai_provider_changed(self):
        provider = self._ai_provider_var.get()
        self.cfg.setdefault("ai_settings", {})["provider"] = provider
        # Auto-fill default model for selected provider
        if PROVIDERS and provider in PROVIDERS:
            default_model = PROVIDERS[provider].get("default_model", "")
            self._ai_model_var.set(default_model)
            self.cfg["ai_settings"]["model"] = default_model
        self._save_config()

    def _test_ai_connection(self):
        self._ai_status_var.set("Testing connection...")
        self._append_ai_log("Starting AI/OCR connection test...")

        if not AI_MATCHER_AVAILABLE or self._ai_matcher is None:
            self._ai_status_var.set("ai_supplier_matcher.py not available")
            self._append_ai_log("ai_supplier_matcher.py not available.")
            return

        def _run():
            try:
                self.cfg.setdefault("ai_settings", {}).update({
                    "enabled": bool(self._ai_enabled_var.get()),
                    "provider": self._ai_provider_var.get(),
                    "model": self._ai_model_var.get(),
                    "api_key": self._ai_key_var.get().strip(),
                    "custom_base_url": self._ai_url_var.get().strip(),
                    "timeout_seconds": int(self._ai_timeout_var.get()),
                })

                self.cfg.setdefault("ocr_space", {}).update({
                    "api_key": self._ocr_space_key_var.get().strip() or "K88109865088957",
                    "OCREngine": int(self._ocr_space_engine_var.get()),
                    "language": "eng",
                    "isOverlayRequired": False,
                    "detectOrientation": True,
                    "scale": True,
                    "isTable": False,
                    "filetype": "PDF",
                    "timeout_seconds": 30,
                })

                try:
                    max_mb = float(self._ocr_space_max_mb_var.get() or 1.0)
                except Exception:
                    max_mb = 1.0

                self.cfg["ocr_space"]["max_upload_mb"] = max_mb

                from ai_supplier_matcher import AISupplierMatcher
                self._ai_matcher = AISupplierMatcher(self.cfg, logger_func=self._append_ai_log)
                status_code, message = self._ai_matcher.test_connection()

            except Exception as e:
                status_code = STATUS_DISCONNECTED
                message = str(e)

            color_map = {
                STATUS_CONNECTED: SUCCESS,
                STATUS_DISCONNECTED: ERROR,
                STATUS_LOW_CREDIT: WARNING,
                STATUS_OFFLINE: MUTED,
            }

            color = color_map.get(status_code, MUTED)

            dot_map = {
                STATUS_CONNECTED: "●",
                STATUS_DISCONNECTED: "✗",
                STATUS_LOW_CREDIT: "⚠",
                STATUS_OFFLINE: "○",
            }

            dot = dot_map.get(status_code, "○")

            self._append_ai_log(f"Connection test result: {status_code} | {message}")
            self.after(0, lambda: self._ai_status_var.set(f"{dot}  {message}"))
            self.after(0, lambda: self._ai_status_lbl.configure(fg=color))

        threading.Thread(target=_run, daemon=True).start()

    # ------------------------------------------------------------------
    # DESKTOP NOTIFICATION HELPER (INSTANCE)
    # ------------------------------------------------------------------
    def _notify(self, title: str, message: str):
        if self.cfg.get("app_settings", {}).get("desktop_notifications", True):
            threading.Thread(
                target=send_desktop_notification, args=(title, message), daemon=True
            ).start()

    # ------------------------------------------------------------------
    # CONFIDENCE DISPLAY HELPER
    # ------------------------------------------------------------------
    def _conf_display(self, confidence: float) -> str:
        if confidence <= 0:
            return "—"
        return f"{confidence:.0f}%"

    def _conf_tag(self, confidence: float, supplier: str) -> str:
        threshold = self.cfg.get("app_settings", {}).get("confidence_warn_threshold", 80)
        if supplier == "UNKNOWN SUPPLIER" or confidence <= 0:
            return "unknown"
        if confidence < threshold:
            return "low_conf"
        return "high_conf"

    # ------------------------------------------------------------------
    # RENAME WORKER
    # ------------------------------------------------------------------
    def _start_rename_worker(self):
        if self._worker_running:
            return

        _reset_last_good_supplier()
        self._worker_running = True
        self._cancel_requested = False

        if not getattr(self, "_rename_was_cancelled", False):
            self._rename_results = []
            self._rename_row_map.clear()
            self._rename_all_rows.clear()
            for i in self._rename_tree.get_children():
                self._rename_tree.delete(i)
        self._rename_was_cancelled = False

        self._set_buttons_state("disabled")
        self._set_status("OCR renaming started...", ACCENT)

        def worker():
            n_success = 0
            n_failed = 0
            failed_names = []

            try:
                scanned = self.dirs.get("scanned")
                processed = self.dirs.get("processed")
                failed = self.dirs.get("failed")

                os.makedirs(scanned, exist_ok=True)
                os.makedirs(processed, exist_ok=True)
                os.makedirs(failed, exist_ok=True)

                files = self._get_pdf_files_strict_order(scanned)
                logging.info("Strict processing order: %s", [os.path.basename(f) for f in files])

                total = len(files)
                if total == 0:
                    self._set_status("SCANNED folder is empty — no files to process.", WARNING)
                    self._rename_queue.put(None)
                    return

                self.after(0, lambda: self._progress.configure(maximum=max(total, 1), value=0))
                mode = self.cfg["app_settings"].get("processing_mode", "legacy")
                dry_run = self.cfg["app_settings"].get("dry_run", False)

                # Force sequential processing to preserve exact file order
                use_threads = False

                if dry_run:
                    previews = []
                    for pdf_path in files:
                        fname = os.path.basename(pdf_path)
                        self._set_status(f"[DRY RUN] Scanning: {fname}...", ACCENT)
                        try:
                            if mode == "legacy":
                                fields = self._process_file_legacy(pdf_path)
                            else:
                                text = self._extract_text(pdf_path)
                                if mode == "mixed":
                                    fields = self._process_file_mixed(pdf_path, text)
                                else:
                                    fields = self._process_file_custom(pdf_path, text)

                            sup = fields.get("supplier") or "UNKNOWN SUPPLIER"
                            grn = fields.get("grn") or "NO-GRN"
                            inv = fields.get("invoice") or "NO-INVOICE"
                            conf = fields.get("confidence", 0.0)

                            inv_part = f"IN {inv}" if inv and inv != "NO-INVOICE" else "NO-INVOICE"
                            proposed = self._safe_filename(f"{sup} GRN {grn} {inv_part}.pdf")
                            dupes = self._find_duplicate_invoice(inv, processed)

                            previews.append({
                                "file": fname,
                                "proposed_name": proposed,
                                "supplier": sup,
                                "grn": grn,
                                "confidence": conf,
                                "invoice": inv if inv != "NO-INVOICE" else "",
                                "duplicate_warning": f"Already exists: {dupes[0]}" if dupes else "",
                                "_path": pdf_path,
                            })
                        except Exception as e:
                            previews.append({
                                "file": fname,
                                "proposed_name": f"ERROR: {e}",
                                "supplier": "",
                                "grn": "",
                                "confidence": 0.0,
                                "invoice": "",
                                "duplicate_warning": "",
                                "_path": pdf_path,
                            })

                    proceed = [False]

                    def _show():
                        proceed[0] = self._show_dry_run_preview(previews)

                    self.after(0, _show)

                    import time
                    deadline = time.time() + 300
                    while time.time() < deadline:
                        time.sleep(0.1)
                        try:
                            if not any(
                                isinstance(w, tk.Toplevel) and "Preview" in (w.title() or "")
                                for w in self.winfo_children()
                            ):
                                break
                        except Exception:
                            break

                    if not proceed[0]:
                        self._set_status("Dry Run cancelled — no files moved.", WARNING)
                        self._rename_queue.put(None)
                        return

                    self.cfg["app_settings"]["dry_run"] = False

                for idx, pdf_path in enumerate(files, 1):
                    if self._cancel_requested:
                        self._set_status("Rename cancelled by user.", WARNING)
                        break

                    fname = os.path.basename(pdf_path)
                    self._set_status(f"[{mode.upper()}] Processing {fname}... ({idx}/{total})", ACCENT)

                    rd = self._rename_single_file(pdf_path, processed, failed, mode)
                    self._rename_results.append(rd)

                    if rd.get("status") == "success":
                        n_success += 1
                    else:
                        n_failed += 1
                        failed_names.append(fname)

                    self._rename_queue.put(("row", rd))
                    self._rename_queue.put(("progress", idx))

                if dry_run:
                    self.cfg["app_settings"]["dry_run"] = True

                if self._cancel_requested:
                    self._rename_was_cancelled = True

                if not self._cancel_requested:
                    summary = f"Rename complete — {n_success} succeeded, {n_failed} failed."
                    self._set_status(summary, SUCCESS if n_failed == 0 else WARNING)

                    notif_msg = summary
                    if n_failed > 0 and failed_names:
                        notif_msg += f"\nFailed: {', '.join(failed_names[:3])}"
                        if len(failed_names) > 3:
                            notif_msg += f" +{len(failed_names) - 3} more"

                    self._notify("Maafushivaru — Rename Complete", notif_msg)

            except Exception as e:
                logging.error(f"Rename worker error: {e}\n{traceback.format_exc()}")
                self.after(0, lambda m=str(e): messagebox.showerror("Rename Error", f"Unexpected error:\n\n{m}"))
                self._set_status(f"Rename failed: {e}", ERROR)

            finally:
                self._worker_running = False
                self._rename_queue.put(None)
                self.after(0, self._refresh_dashboard_stats)

        threading.Thread(target=worker, daemon=True).start()
        self.after(60, self._poll_rename_queue)

    def _cancel_worker(self):
        self._cancel_requested = True
        self._set_status("Cancel requested — stopping after current file...", WARNING)

    def _add_rename_tree_row(self, result):
        s    = result.get("status", "")
        sup  = result.get("supplier", "")
        conf = result.get("confidence", 0.0)
        dupe = result.get("duplicate_warning", "")
        inv_raw = result.get("invoice", "")
        inv_display = f"IN {inv_raw}" if inv_raw and inv_raw not in ("NO-INVOICE", "") else "NO-INVOICE"

        conf_thr = self.cfg.get("app_settings", {}).get("confidence_warn_threshold", 80)

        # Tag priority: error > unknown supplier > duplicate > low confidence > success
        if "error" in s:
            tag = "error"
        elif sup == "UNKNOWN SUPPLIER":
            tag = "unknown"
        elif dupe:
            tag = "dupe_warn"
        elif conf > 0 and conf < conf_thr:
            tag = "low_conf"
        elif s == "corrected":
            tag = "corrected"
        else:
            tag = "success"

        for t, fg in [
            ("success",  SUCCESS),
            ("unknown",  WARNING),
            ("error",    ERROR),
            ("dupe_warn",WARNING),
            ("low_conf", WARNING),
            ("corrected",ACCENT2),
        ]:
            self._rename_tree.tag_configure(t, foreground=fg)

        # Ensure each rename result has a stable doc_id so it can be linked
        # to a Dispatch row later.
        if not result.get("doc_id"):
            # Simple unique id based on filename + current time
            result["doc_id"] = f"{result.get('file', '')}|rename|{datetime.now().timestamp()}"

        row_id = self._rename_tree.insert(
            "", "end", tags=(tag,),
            values=(
                result.get("file", ""),
                result.get("new_name", ""),
                sup,
                self._conf_display(conf),
                result.get("grn", ""),
                inv_display,
                result.get("status", ""),
                dupe,
            ),
        )
        self._rename_row_map[row_id] = result
        self._rename_all_rows.append(row_id)

    def _clear_rename_results(self):
        self._rename_results = []
        self._rename_row_map.clear()
        self._rename_all_rows.clear()
        for i in self._rename_tree.get_children():
            self._rename_tree.delete(i)
        self._set_status("Rename results cleared.")

    # ------------------------------------------------------------------
    # SUPPLIER CORRECTION MEMORY (OFFER ALIAS)
    # ------------------------------------------------------------------
    def _offer_save_alias(self, old_supplier: str, new_supplier: str):
        if not old_supplier or not new_supplier or old_supplier == new_supplier:
            return
        if old_supplier == "UNKNOWN SUPPLIER":
            return
        suppliers = self.cfg.get("suppliers", [])
        if new_supplier.upper() not in [s.upper() for s in suppliers]:
            return
        if messagebox.askyesno(
            "Remember Correction?",
            f'Save "{old_supplier}" as an alias for "{new_supplier}"?\n\n'
            f'Future documents containing "{old_supplier}" will be matched automatically.',
        ):
            aliases = self.cfg.setdefault("aliases", {})
            existing = aliases.get(new_supplier, [])
            if old_supplier.upper() not in [a.upper() for a in existing]:
                existing.append(old_supplier.upper())
                aliases[new_supplier] = existing
                self._save_config()
                self._populate_supplier_tree()
                self._set_status(f"Alias saved: '{old_supplier}' → '{new_supplier}'", SUCCESS)

    def _on_rename_tree_edit(self, row_id, col_index, old_val, new_val):
        vals = list(self._rename_tree.item(row_id, "values"))

        if col_index == 4:   # GRN
            nv = new_val.strip().upper()
            if nv in ("NO-GRN", "", "NO GRN"):
                new_val_disp = "RC-MAM-0000"
            else:
                nv = "RC-MAM-" + nv if not nv.startswith("RC-MAM-") else nv
                new_val_disp = nv
            vals[4] = new_val_disp

        elif col_index == 5:  # Invoice #
            raw = new_val.strip()
            if raw.upper().startswith("IN "):
                raw = raw[3:].strip()
            new_val_disp = "NO-INVOICE" if not raw or raw.upper() == "NO-INVOICE" else f"IN {raw}"
            invoice_internal = "" if new_val_disp == "NO-INVOICE" else raw
            vals[5] = new_val_disp

        elif col_index == 2:  # Supplier
            new_val_disp = new_val.strip().upper()
            vals[2] = new_val_disp

        else:
            vals[col_index] = new_val
            self._rename_tree.item(row_id, values=vals)
            return

        result = self._rename_row_map.get(row_id)
        if not result:
            self._rename_tree.item(row_id, values=vals)
            return

        dp = result.get("dest_path", "")
        if not dp or not os.path.exists(dp):
            messagebox.showwarning("Cannot Rename", f"File not found:\n{dp or '(no path)'}")
            return

        sup = vals[2]
        grn = vals[4]

        if col_index == 5:
            result["invoice"] = invoice_internal
            result["invoice_dispatch"] = invoice_internal
        else:
            existing_disp = vals[5]
            if existing_disp.upper().startswith("IN "):
                result["invoice"] = existing_disp[3:].strip()
                result["invoice_dispatch"] = existing_disp[3:].strip()
            else:
                result["invoice"] = ""
                result["invoice_dispatch"] = ""

        old_supplier = result.get("supplier", "")
        result["supplier"] = sup
        result["grn"] = grn

        internal_invoice = result.get("invoice") or ""
        fn_inv = f"IN {internal_invoice}" if internal_invoice else "NO-INVOICE"

        nd = os.path.join(
            os.path.dirname(dp),
            self._safe_filename(f"{sup} GRN {grn} {fn_inv}.pdf")
        )

        if nd != dp:
            base, ext = os.path.splitext(nd)
            cnt = 1
            while os.path.exists(nd):
                nd = f"{base}_{cnt}{ext}"
                cnt += 1
            try:
                shutil.move(dp, nd)
                result["dest_path"] = nd
                vals[0] = os.path.basename(nd)
                vals[1] = os.path.basename(nd)
            except Exception as e:
                messagebox.showerror("Rename Failed", f"Could not rename:\n\n{e}")
                return
        else:
            vals[0] = os.path.basename(dp)
            vals[1] = os.path.basename(dp)

        # --- Rebuild proposed filename from all current cell values ---
        current_sup = vals[2] or "UNKNOWN SUPPLIER"
        current_grn = vals[4] or "NO-GRN"

        # Get invoice from column 5 display value
        inv_disp = vals[5] if len(vals) > 5 else "NO-INVOICE"
        if inv_disp.upper().startswith("IN "):
            current_inv_internal = inv_disp[3:].strip()
        else:
            current_inv_internal = ""

        fn_inv_part = f"IN {current_inv_internal}" if current_inv_internal else "NO-INVOICE"
        new_proposed_filename = self._safe_filename(
            f"{current_sup} GRN {current_grn} {fn_inv_part}.pdf"
        )

        # Update column 1 (New Name) live
        vals[1] = new_proposed_filename
        # --- End live filename rebuild ---

        result["file"] = vals[0]
        result["new_name"] = vals[1]
        result["status"] = "corrected"

        self._rename_tree.item(row_id, values=vals)
        self._rename_tree.item(row_id, tags=("corrected",))
        self._rename_tree.tag_configure("corrected", foreground=ACCENT2)

        self._set_status(f"Corrected: {vals[1]}", ACCENT2)
        self._refresh_dashboard_stats()
        self._sync_rename_to_dispatch(result)
        # Mirror the correction back to the AI Extract result tree too.
        self._sync_to_ai_extract(
            result.get("doc_id", ""),
            result.get("supplier", ""),
            result.get("grn", ""),
            result.get("invoice", ""),
            result.get("file", ""),
        )

        if col_index == 2 and old_supplier != sup:
            self.after(300, lambda: self._offer_save_alias(old_supplier, sup))

    # ------------------------------------------------------------------
    # DISPATCH WORKER
    # ------------------------------------------------------------------
    def _start_dispatch_worker(self):
        if self._dispatch_running:
            return

        self._dispatch_running = True
        self._cancel_requested = False
        self._dispatch_results = []
        self._dispatch_row_map.clear()
        self._dispatch_all_rows.clear()

        for i in self._dispatch_tree.get_children():
            self._dispatch_tree.delete(i)

        self._set_buttons_state("disabled")
        self._set_status("GRN extraction started...", ACCENT)

        def worker():
            n_success = 0
            n_failed = 0
            failed_names = []

            try:
                folder = self.dirs.get("processed") or self.dirs.get("scanned")
                if hasattr(self, "_dispatch_folder_var"):
                    ch = self._dispatch_folder_var.get().strip()
                    if ch:
                        folder = ch

                files = sorted(
                    [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith(".pdf")],
                    key=lambda p: os.path.getmtime(p),
                    reverse=True,
                )

                total = len(files)
                self.after(0, lambda: self._dispatch_progress.configure(maximum=max(total, 1), value=0))

                ut = self.cfg["app_settings"].get("enable_multi_threading", True)
                mt = self.cfg["app_settings"].get("max_threads", 4)
                mode = self.cfg["app_settings"].get("processing_mode", "legacy")

                if ut and total > 1:
                    with concurrent.futures.ThreadPoolExecutor(max_workers=mt) as ex:
                        fm = {
                            ex.submit(self._dispatch_single_file, p, idx, mode): p
                            for idx, p in enumerate(files)
                        }
                        dc = 0
                        for fut in concurrent.futures.as_completed(fm):
                            if self._cancel_requested:
                                break
                            r = fut.result()
                            self._dispatch_results.append(r)
                            dc += 1

                            fname = r.get("file", "")
                            sup = r.get("supplier", "")
                            conf = r.get("confidence", 0.0)

                            self._set_status(
                                f"[GRN] {fname}  →  {sup} ({conf:.0f}%)  [{dc}/{total}]",
                                SUCCESS if r.get("is_valid") else ERROR,
                            )

                            if r.get("is_valid"):
                                n_success += 1
                            else:
                                n_failed += 1
                                failed_names.append(fname)

                            self._dispatch_queue.put(("row", r))
                            self._dispatch_queue.put(("progress", dc))
                else:
                    for idx, p in enumerate(files):
                        if self._cancel_requested:
                            break

                        fname = os.path.basename(p)
                        self._set_status(f"[GRN] Extracting {fname}... ({idx + 1}/{total})", ACCENT)

                        r = self._dispatch_single_file(p, idx, mode)
                        self._dispatch_results.append(r)

                        if r.get("is_valid"):
                            n_success += 1
                        else:
                            n_failed += 1
                            failed_names.append(fname)

                        self._dispatch_queue.put(("row", r))
                        self._dispatch_queue.put(("progress", idx + 1))

                summary = f"GRN extraction complete — {n_success} ok, {n_failed} failed."
                self._set_status(summary, SUCCESS if n_failed == 0 else WARNING)

                notif = summary
                if n_failed and failed_names:
                    notif += f"\nFailed: {', '.join(failed_names[:3])}"

                self._notify("Maafushivaru — GRN Extraction Complete", notif)

            except Exception as e:
                logging.error("Dispatch worker failed", exc_info=True)
                self.after(0, lambda m=str(e): messagebox.showerror("Dispatch Error", m))
                self._set_status(f"GRN extraction failed: {e}", ERROR)

            finally:
                self._dispatch_running = False
                self._dispatch_queue.put(None)

        threading.Thread(target=worker, daemon=True).start()
        self.after(60, self._poll_dispatch_queue)

    def _add_dispatch_tree_row(self, result):
        iv   = result.get("is_valid", False)
        conf = result.get("confidence", 0.0)
        conf_thr = self.cfg.get("app_settings", {}).get("confidence_warn_threshold", 80)
        sup = result.get("supplier", "")

        if not iv:
            tag = "invalid"
        elif sup == "UNKNOWN SUPPLIER" or conf <= 0:
            tag = "unknown"
        elif conf < conf_thr:
            tag = "low_conf"
        else:
            tag = "valid"

        for t, fg in [
            ("valid",    SUCCESS),
            ("invalid",  ERROR),
            ("unknown",  WARNING),
            ("low_conf", WARNING),
        ]:
            self._dispatch_tree.tag_configure(t, foreground=fg)

        row_id = self._dispatch_tree.insert(
            "", "end", tags=(tag,),
            values=(
                result.get("file", ""),
                result.get("date", ""),
                sup,
                self._conf_display(conf),
                result.get("po", "") or "MAM-0000",
                result.get("invoice", ""),
                result.get("usd", ""),
                result.get("mvr", ""),
                result.get("eur", ""),
                result.get("gbp", ""),
                result.get("sgd", ""),
                result.get("grn", ""),
            ),
        )
        self._dispatch_row_map[row_id] = result
        self._dispatch_all_rows.append(row_id)

    def _clear_dispatch_results(self):
        self._dispatch_results = []
        self._dispatch_row_map.clear()
        self._dispatch_all_rows.clear()
        for i in self._dispatch_tree.get_children():
            self._dispatch_tree.delete(i)
        self._set_status("Dispatch results cleared.")

    def _on_dispatch_tree_edit(self, row_id, col_index, old_val, new_val):
        vals = list(self._dispatch_tree.item(row_id, "values"))

        if col_index == 11:  # GRN
            if new_val:
                nv = str(new_val).strip().upper()
                if not nv.startswith("RC-MAM-"):
                    nv = "RC-MAM-" + nv
                new_val = nv

        if col_index == 5:  # Invoice #
            raw = str(new_val).strip()
            if raw.upper().startswith("IN "):
                raw = raw[3:].strip()
            new_val = raw

        if col_index == 2:  # Supplier
            new_val = str(new_val).strip().upper()

        vals[col_index] = new_val
        self._dispatch_tree.item(row_id, values=vals)

        result = self._dispatch_row_map.get(row_id)
        if result:
            keys = ["file", "date", "supplier", "_conf", "po", "invoice",
                    "usd", "mvr", "eur", "gbp", "sgd", "grn"]
            if col_index < len(keys) and keys[col_index] != "_conf":
                result[keys[col_index]] = new_val

        if col_index in (2, 5, 11) and result:
            try:
                pdf_path = result.get("raw_path", "")
                if not pdf_path or not os.path.exists(pdf_path):
                    folder = self.dirs.get("processed") or self.dirs.get("scanned")
                    if hasattr(self, "_dispatch_folder_var"):
                        ch = self._dispatch_folder_var.get().strip()
                        if ch:
                            folder = ch
                    pdf_path = os.path.join(folder, result.get("file", ""))

                if not pdf_path or not os.path.exists(pdf_path):
                    messagebox.showwarning(
                        "Cannot Rename",
                        f"File not found:\n{pdf_path or '(no path)'}",
                    )
                else:
                    supplier = (vals[2] or "").strip().upper() or "UNKNOWN SUPPLIER"
                    grn_val = (vals[11] or "").strip().upper() or "NO-GRN"
                    inv_raw = (vals[5] or "").strip()

                    if inv_raw:
                        inv_part = f"IN {inv_raw}"
                    else:
                        inv_part = "NO-INVOICE"

                    new_dir = os.path.dirname(pdf_path)
                    new_name = self._safe_filename(f"{supplier} GRN {grn_val} {inv_part}.pdf")
                    new_path = os.path.join(new_dir, new_name)

                    base, ext = os.path.splitext(new_path)
                    cnt = 1
                    while os.path.exists(new_path) and new_path.lower() != pdf_path.lower():
                        new_path = f"{base}_{cnt}{ext}"
                        cnt += 1

                    if new_path != pdf_path:
                        shutil.move(pdf_path, new_path)

                    result["raw_path"] = new_path
                    result["file"] = new_name
                    result["supplier"] = supplier
                    result["invoice"] = inv_raw
                    result["grn"] = grn_val

                    vals[0] = new_name
                    vals[2] = supplier
                    vals[5] = inv_raw
                    vals[11] = grn_val
                    self._dispatch_tree.item(row_id, values=vals)

                    self._set_status(f"File renamed to {new_name}", ACCENT2)

            except Exception as e:
                logging.error(f"Dispatch rename failed: {e}", exc_info=True)
                messagebox.showerror("Rename Failed", f"Could not rename file:\n\n{e}")

        if result:
            self._sync_dispatch_to_rename(result)
            # Recolour the row (a corrected supplier is no longer "unknown")
            # and mirror the edit back to the AI Extract result tree.
            self._refresh_dispatch_row_tag(row_id)
            self._sync_to_ai_extract(
                result.get("doc_id", ""),
                result.get("supplier", ""),
                result.get("grn", ""),
                result.get("invoice", ""),
                result.get("file", ""),
            )

        self._set_status("Dispatch cell updated. Re-export to Excel to save changes.")

    # ------------------------------------------------------------------
    # EXPORT TO EXCEL
    # ------------------------------------------------------------------
    def _export_dispatch_to_excel(self):
        if not self._dispatch_results:
            messagebox.showwarning("No Data", "Run GRN Extraction first.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "GRN Data"

        headers = [
            "INVOICE DATE", "SUPPLIER NAME", "PURCHASE ORDER #", "INVOICE #",
            "USD", "MVR", "EUR", "GBP", "SGD", "GRN NO."
        ]
        ws.append(headers)

        CURRENCY_COLS = {5: "E", 6: "F", 7: "G", 8: "H", 9: "I"}

        for row_id in self._dispatch_tree.get_children():
            vals = self._dispatch_tree.item(row_id, "values")
            if not vals:
                continue

            row_data = [
                vals[1],   # Date
                vals[2],   # Supplier
                vals[4],   # PO
                vals[5],   # Invoice
                vals[6],   # USD
                vals[7],   # MVR
                vals[8],   # EUR
                vals[9],   # GBP
                vals[10],  # SGD
                vals[11],  # GRN
            ]
            ws.append(row_data)
            rn = ws.max_row

            for ci, cl in CURRENCY_COLS.items():
                cell = ws[f"{cl}{rn}"]
                val = row_data[ci - 1]
                if val not in ("", None):
                    try:
                        cell.value = float(str(val).replace(",", ""))
                        cell.number_format = NUMBER_FMT
                    except (ValueError, TypeError):
                        cell.value = val
                else:
                    cell.value = ""

        font = Font(name="Arial Narrow", size=10)
        align = Alignment(horizontal="center", vertical="center")

        for col, w in zip("ABCDEFGHIJ", [14, 35, 18, 22, 12, 12, 12, 12, 12, 60]):
            ws.column_dimensions[col].width = w

        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = align

        # Each manual export makes a BRAND-NEW file - never overwrite. The
        # timestamp normally guarantees uniqueness; the counter covers the
        # rare case of two exports within the same second.
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = os.path.join(self.dirs["base"], f"GRN_OUTPUT_{ts}.xlsx")
        _cnt = 1
        while os.path.exists(out):
            out = os.path.join(self.dirs["base"], f"GRN_OUTPUT_{ts}_{_cnt}.xlsx")
            _cnt += 1
        wb.save(out)

        self._set_status(f"Excel exported: {out}", SUCCESS)
        self._notify("Maafushivaru — Export Complete", f"{os.path.basename(out)} saved.")

        if messagebox.askyesno("Export Successful", f"Saved:\n{out}\n\nOpen now?"):
            os.startfile(out)

    def _write_grn_excel(self, rows: List[Dict]) -> str:
        """Write GRN rows (list of result dicts) to a BRAND-NEW timestamped
        Excel file and return its path. Shared by the AI Extract "Export Excel"
        button and the 30-result "generate sheet now?" prompt. Never overwrites
        an existing file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "GRN Data"

        headers = [
            "INVOICE DATE", "SUPPLIER NAME", "PURCHASE ORDER #", "INVOICE #",
            "USD", "MVR", "EUR", "GBP", "SGD", "GRN NO."
        ]
        ws.append(headers)

        CURRENCY_COLS = {5: "E", 6: "F", 7: "G", 8: "H", 9: "I"}

        for r in rows:
            inv = r.get("invoice", "")
            if isinstance(inv, str) and inv.upper() == "NO-INVOICE":
                inv = ""
            row_data = [
                r.get("date", ""),
                r.get("supplier", "") or "UNKNOWN SUPPLIER",
                r.get("po", "") or "MAM-0000",
                inv,
                r.get("usd", ""),
                r.get("mvr", ""),
                r.get("eur", ""),
                r.get("gbp", ""),
                r.get("sgd", ""),
                r.get("grn", "") or "RC-MAM-0000",
            ]
            ws.append(row_data)
            rn = ws.max_row
            for ci, cl in CURRENCY_COLS.items():
                cell = ws[f"{cl}{rn}"]
                val = row_data[ci - 1]
                if val not in ("", None):
                    try:
                        cell.value = float(str(val).replace(",", ""))
                        cell.number_format = NUMBER_FMT
                    except (ValueError, TypeError):
                        cell.value = val
                else:
                    cell.value = ""

        font = Font(name="Arial Narrow", size=10)
        align = Alignment(horizontal="center", vertical="center")
        for col, w in zip("ABCDEFGHIJ", [14, 35, 18, 22, 12, 12, 12, 12, 12, 60]):
            ws.column_dimensions[col].width = w
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = align

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = os.path.join(self.dirs["base"], f"GRN_OUTPUT_{ts}.xlsx")
        _cnt = 1
        while os.path.exists(out):
            out = os.path.join(self.dirs["base"], f"GRN_OUTPUT_{ts}_{_cnt}.xlsx")
            _cnt += 1
        wb.save(out)
        return out

    # ------------------------------------------------------------------
    # VALIDATE GRN
    # ------------------------------------------------------------------
    def _validate_grn(self):
        rows = self._dispatch_tree.get_children()
        if not rows:
            messagebox.showinfo("Validate GRN", "No results to validate. Run GRN Extraction first.")
            return

        folder = self.dirs.get("processed") or self.dirs.get("scanned")
        if hasattr(self, "_dispatch_folder_var"):
            ch = self._dispatch_folder_var.get().strip()
            if ch:
                folder = ch

        if not os.path.isdir(folder):
            messagebox.showerror("Validate GRN", f"Folder not found:\n{folder}")
            return

        folder_files = {f.lower() for f in os.listdir(folder) if f.lower().endswith(".pdf")}
        self._dispatch_tree.tag_configure("missing", foreground=ERROR)
        self._dispatch_tree.tag_configure("present", foreground=SUCCESS)
        found = missing = 0

        for row_id in rows:
            vals = self._dispatch_tree.item(row_id, "values")
            filename = str(vals[0]).strip() if vals else ""
            if filename.lower() in folder_files:
                self._dispatch_tree.item(row_id, tags=("present",))
                found += 1
            else:
                self._dispatch_tree.item(row_id, tags=("missing",))
                missing += 1

        msg = (
            f"Validation complete.\n\n"
            f"Rows checked : {found + missing}\n"
            f"Files FOUND  : {found}  (green)\n"
            f"Files MISSING: {missing}  (red)\n\n"
            f"Folder: {folder}"
        )
        (messagebox.showinfo if missing == 0 else messagebox.showwarning)(
            "Validate GRN" + (" — All OK" if missing == 0 else " — Missing Files"), msg
        )
        self._set_status(f"GRN Validation: {found} found, {missing} missing.",
                         SUCCESS if missing == 0 else WARNING)

    # ------------------------------------------------------------------
    # UNIFIED EXTRACT & PROCESS
    # ------------------------------------------------------------------
    def _extract_core_fields_for_file(self, pdf_path: str, mode: str) -> Dict:
        text = self._extract_text(pdf_path)
        if len((text or "").strip()) < 40:
            ai_ocr_text = self._extract_text_via_ai_ocrspace(pdf_path)
            if ai_ocr_text:
                text = ai_ocr_text

        rr = self._extract_receiving_report_fields(pdf_path)

        supplier = None
        invoice_for_filename = ""
        invoice_for_dispatch = ""
        confidence = 0.0

        app_s = self.cfg.get("app_settings", {})
        ocr_correction_on  = app_s.get("ocr_word_correction_enabled", False)
        cross_match_on     = app_s.get("smart_cross_match_enabled", False)
        suppliers_list     = self.cfg.get("suppliers", [])
        aliases_map        = self.cfg.get("aliases", {})

        # ── Step 0: OCR Word Correction (optional) ────────────────────────
        working_text = text
        if ocr_correction_on and OCR_CORRECTOR_AVAILABLE:
            corrected, ocr_detected_supplier = correct_supplier_in_text(
                text, suppliers_list, aliases_map
            )
            if ocr_detected_supplier:
                working_text = corrected
                supplier     = ocr_detected_supplier
                confidence   = 78.0
                logging.info(f"[OCR-CORR] OCR word correction matched supplier: {supplier}")

        # ── Step 1: Normal supplier extraction ───────────────────────────
        if not supplier:
            if mode == "legacy":
                cands, aliases = self._legacy_load_suppliers()
                legacy_text    = self._legacy_extract_text(pdf_path)
                supplier       = self._legacy_extract_supplier(legacy_text, cands, aliases) or "UNKNOWN SUPPLIER"
                invoice_for_filename = self._extract_invoice_old(legacy_text) or "NO-INVOICE"
                invoice_for_dispatch = self._extract_invoice_old(legacy_text) or ""
                confidence = 70.0 if supplier != "UNKNOWN SUPPLIER" else 0.0

            elif mode == "mixed":
                method = app_s.get("supplier_extraction_method", "both")
                if method in ("receiving_report_field", "both"):
                    supplier = self._extract_supplier_from_field(working_text)
                    if supplier: confidence = 85.0
                if not supplier and method in ("header_company_name", "both"):
                    supplier = self._extract_company_from_invoice_pages(pdf_path)
                    if supplier: confidence = 80.0
                if not supplier:
                    supplier, confidence = self._match_supplier_with_confidence(working_text, os.path.basename(pdf_path))
                if not supplier or supplier == "UNKNOWN SUPPLIER":
                    supplier   = _get_last_good_supplier() or "UNKNOWN SUPPLIER"
                    confidence = 0.0
                else:
                    _set_last_good_supplier(supplier)
                legacy_text = self._legacy_extract_text(pdf_path)
                invoice_for_filename = self._extract_invoice_old(legacy_text) or "NO-INVOICE"
                invoice_for_dispatch = self._extract_invoice_old(legacy_text) or ""

            else:  # custom
                method = app_s.get("supplier_extraction_method", "both")
                if method in ("receiving_report_field", "both"):
                    supplier = self._extract_supplier_from_field(working_text)
                    if supplier: confidence = 85.0
                if not supplier and method in ("header_company_name", "both"):
                    supplier = self._extract_company_from_invoice_pages(pdf_path)
                    if supplier: confidence = 80.0
                if not supplier:
                    supplier, confidence = self._match_supplier_with_confidence(working_text, os.path.basename(pdf_path))
                if not supplier or supplier == "UNKNOWN SUPPLIER":
                    supplier   = _get_last_good_supplier() or "UNKNOWN SUPPLIER"
                    confidence = 0.0
                else:
                    _set_last_good_supplier(supplier)
                invoice_for_filename = self._extract_invoice(working_text, supplier_hint=supplier) or "NO-INVOICE"
                invoice_for_dispatch = self._extract_invoice_old(working_text) or ""

        # Extract invoices for legacy mode if OCR correction found supplier early
        if mode == "legacy" and supplier != "UNKNOWN SUPPLIER" and not invoice_for_filename:
            legacy_text = self._legacy_extract_text(pdf_path)
            invoice_for_filename = self._extract_invoice_old(legacy_text) or "NO-INVOICE"
            invoice_for_dispatch = self._extract_invoice_old(legacy_text) or ""

        # ── Step 2: Smart Cross-Matching (optional) ───────────────────────
        if cross_match_on and CROSS_MATCHER_AVAILABLE:
            # Feature B: if supplier still unknown, try to ID from invoice pattern
            if (not supplier or supplier == "UNKNOWN SUPPLIER") and (invoice_for_dispatch or invoice_for_filename):
                invoice_probe = invoice_for_dispatch or invoice_for_filename
                inferred_sup, inferred_conf = infer_supplier_from_invoice(
                    working_text + " " + invoice_probe,
                    suppliers_list,
                    aliases_map,
                )
                if inferred_sup and inferred_conf > 0:
                    supplier   = inferred_sup
                    confidence = inferred_conf
                    logging.info(f"[CROSS-MATCH] Invoice pattern → supplier: {supplier} ({confidence:.0f}%)")
                    _set_last_good_supplier(supplier)

            # Feature A: if supplier is now known, clean the invoice number
            if supplier and supplier != "UNKNOWN SUPPLIER":
                raw_inv = invoice_for_filename if invoice_for_filename not in ("NO-INVOICE", "") else invoice_for_dispatch
                if raw_inv and raw_inv != "NO-INVOICE":
                    cleaned_inv = infer_invoice_from_supplier(raw_inv, supplier)
                    if cleaned_inv != raw_inv:
                        logging.info(f"[CROSS-MATCH] Invoice corrected: {raw_inv} → {cleaned_inv}")
                    invoice_for_filename = cleaned_inv or "NO-INVOICE"
                    invoice_for_dispatch = cleaned_inv or ""

        grn    = rr.get("grn", "") or self._extract_grn_full(pdf_path, text) or "RC-MAM-0000"
        po     = rr.get("po",  "") or "MAM-0000"
        date_v = rr.get("date","") or ""
        totals = rr.get("totals", {"USD": "", "MVR": "", "EUR": "", "GBP": "", "SGD": ""})

        return {
            "supplier": supplier or "UNKNOWN SUPPLIER",
            "grn": grn, "po": po, "date": date_v,
            "confidence": confidence,
            "invoice_filename": invoice_for_filename or "NO-INVOICE",
            "invoice_dispatch": invoice_for_dispatch or "",
            "usd": totals.get("USD", ""), "mvr": totals.get("MVR", ""),
            "eur": totals.get("EUR", ""), "gbp": totals.get("GBP", ""),
            "sgd": totals.get("SGD", ""), "text": text,
        }

    def _build_dispatch_from_rename_result(self, rename_result: Dict, scan_index: int = 0) -> Dict:
        return {
            "doc_id":     rename_result.get("doc_id", ""),
            "file":       rename_result.get("new_name", "") or rename_result.get("file", ""),
            "date":       rename_result.get("date", ""),
            "supplier":   rename_result.get("supplier", "UNKNOWN SUPPLIER"),
            "po":         rename_result.get("po", "") or "MAM-0000",
            "invoice":    rename_result.get("invoice_dispatch", rename_result.get("invoice", "")),
            "usd":        rename_result.get("usd", ""),
            "mvr":        rename_result.get("mvr", ""),
            "eur":        rename_result.get("eur", ""),
            "gbp":        rename_result.get("gbp", ""),
            "sgd":        rename_result.get("sgd", ""),
            "grn":        rename_result.get("grn", ""),
            "confidence": rename_result.get("confidence", 0.0),
            "is_valid":   rename_result.get("status") in ("success", "simulated"),
            "errors":     rename_result.get("errors", ""),
            "raw_path":   rename_result.get("dest_path", ""),
            "scan_index": scan_index,
        }

    def _rename_and_build_dispatch_single(self, pdf_path, processed, failed, mode, scan_index=0):
        src_name = os.path.basename(pdf_path)
        doc_id   = f"{src_name}|{scan_index}|{datetime.now().timestamp()}"

        rename_result = {
            "doc_id": doc_id, "file": src_name, "new_name": "",
            "supplier": "UNKNOWN SUPPLIER", "grn": "NO-GRN",
            "invoice": "", "invoice_dispatch": "", "date": "", "po": "MAM-0000",
            "usd": "", "mvr": "", "eur": "", "gbp": "", "sgd": "",
            "status": "error", "dest_path": "", "errors": "",
            "duplicate_warning": "", "confidence": 0.0,
        }

        try:
            core = self._extract_core_fields_for_file(pdf_path, mode)

            supplier         = core["supplier"] or "UNKNOWN SUPPLIER"
            grn              = core["grn"] or "RC-MAM-0000"
            invoice_filename = core["invoice_filename"] or "NO-INVOICE"
            invoice_dispatch = core["invoice_dispatch"] or ""
            date_val         = core["date"] or ""
            po               = core["po"] or "MAM-0000"
            confidence       = core.get("confidence", 0.0)

            rename_result.update({
                "supplier": supplier, "grn": grn, "confidence": confidence,
                "invoice": invoice_filename if invoice_filename != "NO-INVOICE" else "",
                "invoice_dispatch": invoice_dispatch, "date": date_val, "po": po,
                "usd": core["usd"], "mvr": core["mvr"], "eur": core["eur"],
                "gbp": core["gbp"], "sgd": core["sgd"],
            })

            dupes = self._find_duplicate_invoice(invoice_filename, processed)
            if dupes:
                rename_result["duplicate_warning"] = f"Invoice already exists: {dupes[0]}"

            inv_part = f"IN {invoice_filename}" if invoice_filename and invoice_filename != "NO-INVOICE" else "NO-INVOICE"
            dest = os.path.join(processed, self._safe_filename(f"{supplier} GRN {grn} {inv_part}.pdf"))

            cnt = 1
            base, ext = os.path.splitext(dest)
            while os.path.exists(dest):
                dest = f"{base}_{cnt}{ext}"
                cnt += 1

            dry_run = self.cfg.get("app_settings", {}).get("dry_run", False)
            if not dry_run:
                shutil.move(pdf_path, dest)
                rename_result["status"]    = "success"
                rename_result["dest_path"] = dest
            else:
                rename_result["status"]    = "simulated"
                rename_result["dest_path"] = pdf_path

            rename_result["new_name"] = os.path.basename(dest)
            dispatch_result = self._build_dispatch_from_rename_result(rename_result, scan_index)
            return rename_result, dispatch_result

        except Exception as e:
            rename_result["errors"] = str(e)
            rename_result["status"] = "error"
            logging.error(f"Unified process failed [{src_name}]: {e}", exc_info=True)
            try:
                shutil.move(pdf_path, os.path.join(failed, src_name))
            except Exception:
                pass
            dispatch_result = self._build_dispatch_from_rename_result(rename_result, scan_index)
            dispatch_result["is_valid"] = False
            dispatch_result["errors"]   = str(e)
            return rename_result, dispatch_result

    def _start_extract_and_process(self):
        if self._worker_running or self._dispatch_running:
            return

        _reset_last_good_supplier()
        self._worker_running = True
        self._dispatch_running = True
        self._cancel_requested = False

        if not getattr(self, "_ep_was_cancelled", False):
            self._rename_results = []
            self._dispatch_results = []
            self._rename_row_map.clear()
            self._dispatch_row_map.clear()
            self._rename_all_rows.clear()
            self._dispatch_all_rows.clear()

            for i in self._rename_tree.get_children():
                self._rename_tree.delete(i)
            for i in self._dispatch_tree.get_children():
                self._dispatch_tree.delete(i)

        self._ep_was_cancelled = False

        self._set_buttons_state("disabled")
        self._set_status("Extract & Process started...", ACCENT)
        mode = self.cfg.get("app_settings", {}).get("processing_mode", "legacy")

        def worker():
            n_success = 0
            n_failed = 0
            failed_names = []

            try:
                scanned = self.dirs.get("scanned")
                processed = self.dirs.get("processed")
                failed = self.dirs.get("failed")

                os.makedirs(scanned, exist_ok=True)
                os.makedirs(processed, exist_ok=True)
                os.makedirs(failed, exist_ok=True)

                files = self._get_pdf_files_strict_order(scanned)
                logging.info("Strict processing order: %s", [os.path.basename(f) for f in files])

                total = len(files)
                if total == 0:
                    self._set_status("SCANNED folder is empty — nothing to process.", WARNING)
                    self.after(0, lambda: self._progress.configure(maximum=1, value=0))
                    self.after(0, lambda: self._dispatch_progress.configure(maximum=1, value=0))
                    return

                self.after(0, lambda: self._progress.configure(maximum=max(total, 1), value=0))
                self.after(0, lambda: self._dispatch_progress.configure(maximum=max(total, 1), value=0))

                # Force sequential processing to preserve exact order
                for idx, pdf_path in enumerate(files, 1):
                    if self._cancel_requested:
                        self._set_status("Extract & Process cancelled.", WARNING)
                        break

                    fname = os.path.basename(pdf_path)
                    self._set_status(f"[{mode.upper()}] {fname} [{idx}/{total}]", ACCENT)

                    rr, dr = self._rename_and_build_dispatch_single(
                        pdf_path, processed, failed, mode, idx - 1
                    )

                    self._rename_results.append(rr)
                    self._dispatch_results.append(dr)

                    if rr.get("status") == "success":
                        n_success += 1
                    else:
                        n_failed += 1
                        failed_names.append(fname)

                    self.after(0, lambda r=rr: self._add_rename_tree_row(r))
                    self.after(0, lambda r=dr: self._add_dispatch_tree_row(r))
                    self.after(0, lambda v=idx: self._progress.configure(value=v))
                    self.after(0, lambda v=idx: self._dispatch_progress.configure(value=v))

                if self._cancel_requested:
                    self._ep_was_cancelled = True

                if not self._cancel_requested:
                    summary = f"Extract & Process complete — {n_success} succeeded, {n_failed} failed."
                    self._set_status(summary, SUCCESS if n_failed == 0 else WARNING)

                    notif = summary
                    if n_failed and failed_names:
                        notif += f"\nFailed: {', '.join(failed_names[:3])}"
                        if len(failed_names) > 3:
                            notif += f" +{len(failed_names) - 3} more"

                    self._notify("Maafushivaru — Extract & Process Complete", notif)

            except Exception as e:
                logging.error(f"Extract & Process error: {e}\n{traceback.format_exc()}")
                self.after(0, lambda m=str(e): messagebox.showerror("Extract & Process Error", f"Error:\n\n{m}"))
                self._set_status(f"Extract & Process failed: {e}", ERROR)

            finally:
                self._worker_running = False
                self._dispatch_running = False
                self.after(0, self._set_buttons_state, "normal")
                self.after(0, self._refresh_dashboard_stats)

        threading.Thread(target=worker, daemon=True).start()

    def _extract_text_via_ai_ocrspace(self, pdf_path: str) -> str:
        """
        Optional OCR.space extraction path.
        Uses OCR.space if available and configured.
        Returns upper-cased text or empty string on failure.
        """
        if not AI_MATCHER_AVAILABLE or self._ai_matcher is None:
            return ""

        try:
            self._append_ai_log(f"Trying OCR.space extraction for: {os.path.basename(pdf_path)}")
            text, err = self._ai_matcher.extract_text(file_path=pdf_path)
            if err:
                self._append_ai_log(f"OCR.space error: {err}")
                return ""
            self._append_ai_log(f"OCR.space extraction success: {len(text)} chars")
            return (text or "").upper()
        except Exception as e:
            self._append_ai_log(f"OCR.space extraction exception: {e}")
            return ""

    # ------------------------------------------------------------------
    # LINK HELPERS
    # ------------------------------------------------------------------
    def _find_dispatch_row_by_doc_id(self, doc_id: str):
        for row_id, data in self._dispatch_row_map.items():
            if data.get("doc_id") == doc_id:
                return row_id, data
        return None, None

    def _find_rename_row_by_doc_id(self, doc_id: str):
        for row_id, data in self._rename_row_map.items():
            if data.get("doc_id") == doc_id:
                return row_id, data
        return None, None

    def _apply_ai_extract_edit(self, ai_result: Dict) -> bool:
        """Propagate an edit made on the AI Extract tab to the OCR Renamer and
        GRN Dispatch tabs, and rename the file on disk to match.

        Only acts if this document has ALREADY been sent to the tabs (matched by
        doc_id). Returns True if it propagated, False if there was nothing to do.
        """
        doc_id = ai_result.get("doc_id", "")
        if not doc_id:
            return False
        r_row, rename_result = self._find_rename_row_by_doc_id(doc_id)
        if not r_row or not rename_result:
            return False  # not sent to tabs yet -> nothing to update

        supplier = (ai_result.get("supplier") or "UNKNOWN SUPPLIER").strip().upper()
        grn = (ai_result.get("grn") or "RC-MAM-0000").strip().upper()
        inv_raw = (ai_result.get("invoice") or "").strip()
        inv_internal = "" if (not inv_raw or inv_raw.upper() == "NO-INVOICE") else inv_raw

        # --- Rename the renamed file on disk to reflect the edit ---
        dp = rename_result.get("dest_path", "")
        if dp and os.path.exists(dp):
            fn_inv = f"IN {inv_internal}" if inv_internal else "NO-INVOICE"
            nd = os.path.join(
                os.path.dirname(dp),
                self._safe_filename(f"{supplier} GRN {grn} {fn_inv}.pdf"),
            )
            if os.path.abspath(nd) != os.path.abspath(dp):
                base, ext = os.path.splitext(nd)
                cnt = 1
                while os.path.exists(nd):
                    nd = f"{base}_{cnt}{ext}"
                    cnt += 1
                try:
                    shutil.move(dp, nd)
                    rename_result["dest_path"] = nd
                except Exception as e:
                    logging.error(f"[AI EDIT] rename failed: {e}", exc_info=True)
                    messagebox.showerror("Rename Failed", f"Could not rename file:\n\n{e}")
                    return False

        new_name = os.path.basename(rename_result.get("dest_path", "")) or rename_result.get("file", "")
        rename_result.update({
            "file": new_name,
            "new_name": new_name,
            "supplier": supplier,
            "grn": grn,
            "invoice": inv_internal,
            "invoice_dispatch": inv_internal,
            "date": ai_result.get("date", rename_result.get("date", "")),
            "po": ai_result.get("po", "") or "MAM-0000",
            "usd": ai_result.get("usd", ""),
            "mvr": ai_result.get("mvr", ""),
            "eur": ai_result.get("eur", ""),
            "gbp": ai_result.get("gbp", ""),
            "sgd": ai_result.get("sgd", ""),
            "status": "corrected",
        })

        # --- Update OCR Renamer tree row ---
        # rename columns: 0 file, 1 new_name, 2 supplier, 3 conf, 4 grn,
        #                 5 invoice, 6 status, 7 dupe
        vals = list(self._rename_tree.item(r_row, "values"))
        if vals:
            inv_disp = f"IN {inv_internal}" if inv_internal else "NO-INVOICE"
            vals[0] = new_name
            vals[1] = new_name
            vals[2] = supplier
            vals[4] = grn
            vals[5] = inv_disp
            if len(vals) > 6:
                vals[6] = "corrected"
            self._rename_tree.item(r_row, values=vals, tags=("corrected",))
            self._rename_tree.tag_configure("corrected", foreground=ACCENT2)

        # --- Propagate to GRN Dispatch (file/supplier/invoice/grn) ---
        self._sync_rename_to_dispatch(rename_result)

        # --- Also push date / PO / currency totals to the dispatch row ---
        # dispatch columns: 0 file,1 date,2 supplier,3 conf,4 po,5 invoice,
        #                   6 usd,7 mvr,8 eur,9 gbp,10 sgd,11 grn
        d_row, dispatch_result = self._find_dispatch_row_by_doc_id(doc_id)
        if d_row and dispatch_result:
            dvals = list(self._dispatch_tree.item(d_row, "values"))
            if dvals and len(dvals) >= 12:
                dvals[1] = rename_result["date"]
                dvals[4] = rename_result["po"]
                dvals[6] = rename_result["usd"]
                dvals[7] = rename_result["mvr"]
                dvals[8] = rename_result["eur"]
                dvals[9] = rename_result["gbp"]
                dvals[10] = rename_result["sgd"]
                self._dispatch_tree.item(d_row, values=dvals)
            dispatch_result.update({
                "date": rename_result["date"],
                "po": rename_result["po"],
                "usd": rename_result["usd"],
                "mvr": rename_result["mvr"],
                "eur": rename_result["eur"],
                "gbp": rename_result["gbp"],
                "sgd": rename_result["sgd"],
            })

        self._refresh_dashboard_stats()
        logging.info(f"[AI EDIT] Propagated edit + rename for doc {doc_id} -> {new_name}")
        return True

    def _sync_rename_to_dispatch(self, rename_result: Dict):
        doc_id = rename_result.get("doc_id", "")
        row_id, dispatch_result = self._find_dispatch_row_by_doc_id(doc_id)
        if not row_id or not dispatch_result:
            return

        vals = list(self._dispatch_tree.item(row_id, "values"))
        if not vals:
            return

        new_file = rename_result.get("new_name", "") or rename_result.get("file", "")
        new_supplier = rename_result.get("supplier", "")
        new_invoice = rename_result.get("invoice_dispatch", rename_result.get("invoice", ""))
        new_grn = rename_result.get("grn", "")

        vals[0] = new_file
        vals[2] = new_supplier
        vals[5] = new_invoice
        vals[11] = new_grn

        self._dispatch_tree.item(row_id, values=vals)

        dispatch_result.update({
            "file": new_file,
            "supplier": new_supplier,
            "invoice": new_invoice,
            "grn": new_grn,
            "raw_path": rename_result.get("dest_path", dispatch_result.get("raw_path", "")),
        })

    def _sync_to_ai_extract(self, doc_id: str, supplier: str, grn: str,
                            invoice_raw: str, file_name: str):
        """Propagate an edit made on the OCR Renamer / GRN Dispatch tab back to
        the matching row in the AI Extract result tree, so all three tabs stay
        consistent (e.g. an "UNKNOWN SUPPLIER" row corrected in Dispatch is no
        longer shown as unknown in AI Extract either)."""
        tree = getattr(self, "_aix_tree", None)
        row_map = getattr(self, "_aix_row_map", None)
        if tree is None or row_map is None or not doc_id:
            return
        for rid, res in list(row_map.items()):
            if res.get("doc_id") != doc_id:
                continue
            res["supplier"] = supplier
            res["grn"] = grn
            res["invoice"] = invoice_raw
            if file_name:
                res["file"] = file_name
            try:
                vals = list(tree.item(rid, "values"))
                if vals and len(vals) >= 12:
                    if file_name:
                        vals[0] = file_name
                    vals[2] = supplier            # Supplier
                    vals[5] = invoice_raw          # Invoice #
                    vals[11] = grn                 # GRN
                    tree.item(rid, values=vals)
            except Exception:
                pass
            break

    def _refresh_dispatch_row_tag(self, row_id: str):
        """Recompute the colour tag of a GRN Dispatch row after an inline edit
        so a corrected supplier no longer shows the orange 'unknown' colour."""
        result = self._dispatch_row_map.get(row_id)
        if not result:
            return
        conf = result.get("confidence", 0.0)
        conf_thr = self.cfg.get("app_settings", {}).get("confidence_warn_threshold", 80)
        sup = (result.get("supplier", "") or "").upper()
        if not result.get("is_valid", True):
            tag = "invalid"
        elif sup == "UNKNOWN SUPPLIER" or not sup:
            tag = "unknown"
        elif conf and conf < conf_thr:
            tag = "low_conf"
        else:
            tag = "valid"
        try:
            self._dispatch_tree.item(row_id, tags=(tag,))
        except Exception:
            pass

    def _sync_dispatch_to_rename(self, dispatch_result: Dict):
        doc_id = dispatch_result.get("doc_id", "")
        row_id, rename_result = self._find_rename_row_by_doc_id(doc_id)
        if not row_id or not rename_result:
            return

        vals = list(self._rename_tree.item(row_id, "values"))
        if not vals:
            return

        new_file = dispatch_result.get("file", "")
        new_supplier = dispatch_result.get("supplier", "")
        new_grn = dispatch_result.get("grn", "")
        new_invoice_raw = dispatch_result.get("invoice", "")

        vals[0] = new_file
        vals[1] = new_file
        vals[2] = new_supplier
        vals[4] = new_grn
        vals[5] = f"IN {new_invoice_raw}" if new_invoice_raw else "NO-INVOICE"

        self._rename_tree.item(row_id, values=vals)

        rename_result.update({
            "file": new_file,
            "new_name": new_file,
            "supplier": new_supplier,
            "grn": new_grn,
            "invoice": new_invoice_raw,
            "invoice_dispatch": new_invoice_raw,
            "dest_path": dispatch_result.get("raw_path", rename_result.get("dest_path", "")),
        })

    # ------------------------------------------------------------------
    # SUPPLIER CONFIG UI
    # ------------------------------------------------------------------
    def _populate_supplier_tree(self):
        for i in self._sup_tree.get_children():
            self._sup_tree.delete(i)
        self._sup_all_rows = []
        suppliers = self.cfg.get("suppliers", [])
        aliases   = self.cfg.get("aliases", {})
        for sup in suppliers:
            alias_str = ", ".join(aliases.get(sup, []))
            iid = self._sup_tree.insert("", "end", values=(sup, alias_str))
            self._sup_all_rows.append(iid)

    def _on_sup_tree_edit(self, row_id, col_index, old_val, new_val):
        vals = list(self._sup_tree.item(row_id, "values"))
        vals[col_index] = new_val
        self._sup_tree.item(row_id, values=vals)

    def _add_supplier_row(self):
        self._sup_tree.insert("", "end", values=("NEW SUPPLIER", ""))

    def _remove_supplier_row(self):
        selected = self._sup_tree.selection()
        if not selected:
            messagebox.showinfo("Remove Supplier", "Select a row first.")
            return
        for row_id in selected:
            vals = self._sup_tree.item(row_id, "values")
            name = vals[0] if vals else ""
            if messagebox.askyesno("Remove Supplier", f"Remove '{name}'?"):
                self._sup_tree.delete(row_id)

    def _save_suppliers(self):
        suppliers = []
        aliases   = {}
        for row_id in self._sup_tree.get_children():
            vals = self._sup_tree.item(row_id, "values")
            name = str(vals[0]).strip().upper() if vals else ""
            alias_raw = str(vals[1]).strip() if len(vals) > 1 else ""
            if not name or name == "NEW SUPPLIER":
                continue
            suppliers.append(name)
            if alias_raw:
                alias_list = [a.strip().upper() for a in alias_raw.split(",") if a.strip()]
                if alias_list:
                    aliases[name] = alias_list
        self.cfg["suppliers"] = suppliers
        self.cfg["aliases"]   = aliases
        self._save_config()
        messagebox.showinfo("Suppliers Saved", f"Saved {len(suppliers)} supplier(s).")

    # ------------------------------------------------------------------
    # TEST SINGLE PDF DEBUG
    # ------------------------------------------------------------------
    def _test_single_pdf_debug(self):
        path = filedialog.askopenfilename(
            title="Select PDF to test",
            filetypes=[("PDF files", "*.pdf")],
            initialdir=self.dirs.get("scanned", "."),
        )
        if not path:
            return

        win = tk.Toplevel(self)
        win.title(f"Debug — {os.path.basename(path)}")
        win.geometry("1020x740")
        win.configure(bg=BG)

        hdr = tk.Frame(win, bg=PANEL2, height=44)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)
        tk.Label(hdr, text=f"Debug: {os.path.basename(path)}",
                 bg=PANEL2, fg=TEXT, font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT, padx=16, pady=12)
        status_lbl = tk.Label(hdr, text="Extracting...", bg=PANEL2, fg=MUTED, font=("Segoe UI", 9))
        status_lbl.pack(side=tk.RIGHT, padx=16)

        nb = ttk.Notebook(win)
        nb.pack(fill=tk.BOTH, expand=True, padx=12, pady=(8, 12))

        def make_text_tab(title):
            f = ttk.Frame(nb)
            nb.add(f, text=f"  {title}  ")
            txt = tk.Text(f, bg=PANEL, fg=TEXT, font=("Consolas", 9),
                          wrap="word", relief="flat", insertbackground=TEXT)
            sb  = ttk.Scrollbar(f, orient="vertical", command=txt.yview)
            txt.configure(yscrollcommand=sb.set)
            txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            sb.pack(side=tk.RIGHT, fill=tk.Y)
            return txt

        raw_text   = make_text_tab("Raw Text")
        match_text = make_text_tab("Extraction Results")

        for tag, fg_color in [
            ("header",  ACCENT),  ("found",   SUCCESS),
            ("missing", ERROR),   ("label",   WARNING),
            ("value",   TEXT),
        ]:
            match_text.tag_configure(tag, foreground=fg_color)
        for tag, bg_color in [
            ("supplier_hl", "#1d4ed8"), ("grn_hl", "#065f46"),
            ("invoice_hl",  "#7c3aed"), ("po_hl",  "#92400e"),
        ]:
            raw_text.tag_configure(tag, background=bg_color, foreground="white")

        def run_debug():
            mode = self.cfg.get("app_settings", {}).get("processing_mode", "legacy")
            try:
                core = self._extract_core_fields_for_file(path, mode)
            except Exception as e:
                win.after(0, lambda: status_lbl.configure(text=f"Error: {e}"))
                win.after(0, lambda: raw_text.insert("end", traceback.format_exc()))
                return

            extracted_text = core.get("text", "")
            supplier       = core.get("supplier", "")
            grn            = core.get("grn", "")
            po             = core.get("po", "")
            inv_fn         = core.get("invoice_filename", "")
            inv_disp       = core.get("invoice_dispatch", "")
            date_val       = core.get("date", "")
            confidence     = core.get("confidence", 0.0)
            totals = {k: core.get(k.lower(), "") for k in ("USD", "MVR", "EUR", "GBP", "SGD")}

            def update_ui():
                status_lbl.configure(text="Done.")
                raw_text.delete("1.0", "end")
                raw_text.insert("end", extracted_text or "(no text extracted)")

                def highlight(term, tag):
                    if not term or len(term) < 3:
                        return
                    start = "1.0"
                    while True:
                        idx = raw_text.search(term, start, nocase=True, stopindex="end")
                        if not idx:
                            break
                        raw_text.tag_add(tag, idx, f"{idx}+{len(term)}c")
                        start = f"{idx}+{len(term)}c"

                highlight(supplier, "supplier_hl")
                highlight(grn.replace("RC-MAM-", "").split("-")[0] if grn else "", "grn_hl")
                highlight(inv_disp, "invoice_hl")
                highlight(po, "po_hl")

                match_text.delete("1.0", "end")

                def ins(text, tag=None):
                    match_text.insert("end", text, tag or ())

                ins(f"=== DEBUG: {os.path.basename(path)} ===\n\n", "header")
                ins(f"Mode: {mode.upper()}\n\n", "label")

                for label, val in [
                    ("Supplier",          supplier),
                    ("Confidence",        f"{confidence:.1f}%"),
                    ("GRN",               grn),
                    ("Invoice (filename)",inv_fn),
                    ("Invoice (dispatch)",inv_disp),
                    ("PO Number",         po),
                    ("Date",              date_val),
                ]:
                    ins(f"  {label:<28}", "label")
                    ok = val and val not in ("NO-GRN", "NO-INVOICE", "UNKNOWN SUPPLIER", "", "0.0%")
                    ins(f"{val or '(not found)'}\n", "found" if ok else "missing")

                ins("\nCurrency Totals:\n", "header")
                for cur, amt in totals.items():
                    ins(f"  {cur:<28}", "label")
                    ins(f"{amt if amt else '(none)'}\n", "found" if amt else "missing")

                ins("\nProposed filename:\n", "header")
                ins(f"  {supplier} GRN {grn} {inv_fn}.pdf\n", "found")

            win.after(0, update_ui)

        threading.Thread(target=run_debug, daemon=True).start()

    # ------------------------------------------------------------------
    # TEST CURRENT OCR ENGINE
    # ------------------------------------------------------------------
    def _test_current_engine(self):
        engine = self.cfg["app_settings"].get("ocr_engine", "tesseract")
        if not self._check_engine(engine) and engine != "tesseract":
            messagebox.showerror("Not Installed",
                                 f"{ENGINE_LABELS[engine]} not installed.\n{ENGINE_INSTALL[engine]}")
            return

        def run():
            try:
                canvas = np.full((100, 450), 255, dtype=np.uint8)
                cv2.putText(canvas, "INVOICE NO: INV-2024-001",   (10, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.75, 0, 2)
                cv2.putText(canvas, "RC-MAM-000000001",           (10, 80), cv2.FONT_HERSHEY_SIMPLEX, 0.75, 0, 2)
                r  = self._run_ocr(canvas)
                ok = bool(r.strip())
                self.after(0, lambda: messagebox.showinfo(
                    "Engine Test",
                    f"{ENGINE_LABELS[engine]}\n\n{r.strip() or '(empty)'}\n\n{'✓ OK' if ok else '✗ Empty output'}",
                ))
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Engine Test Failed", str(e)))

        threading.Thread(target=run, daemon=True).start()


# ---------------------------------------------------------------------------
# ENTRY POINT
# ---------------------------------------------------------------------------
def main():
    app = MaafushivaruHub()
    app.mainloop()


if __name__ == "__main__":
    main()


